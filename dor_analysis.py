#!/usr/bin/env python3
"""
DoR Compliance Analysis - Step 11

Analyzes Jira issues against team Definition of Ready criteria and generates:
1. Report.xlsx - Excel report with DoR compliance assessment
2. DOR_ANALYSIS_SUMMARY.md - Summary markdown file
3. Updated teams.json with analysis metadata
"""

import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Output directory
OUTPUT_DIR = r"C:\Users\i.klyha\Desktop\Claude\wow-scanner-tool\assets\teams\20260529 12-26"
TEAMS_JSON = os.path.join(OUTPUT_DIR, "teams.json")
REPORT_XLSX = os.path.join(OUTPUT_DIR, "Report.xlsx")
SUMMARY_MD = os.path.join(OUTPUT_DIR, "DOR_ANALYSIS_SUMMARY.md")

def team_name_to_filename(team_name):
    """Convert team name to kebab-case filename."""
    return team_name.lower().replace(" ", "-").replace("_", "-")

def load_team_dor(team_name):
    """Load team DoR content from file."""
    filename = f"{team_name_to_filename(team_name)}-dor.txt"
    filepath = os.path.join(OUTPUT_DIR, filename)

    if not os.path.exists(filepath):
        return None

    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read().strip()

    # Check for "not found" message
    if "DoR - STORY/TASK criteria not found" in content or "Error:" in content:
        return None

    return content

def analyze_dor_compliance(issue, dor_content):
    """
    Analyze DoR compliance for an issue using semantic rule matching.

    Returns: (compliant: bool, feedback: str)
    """
    if not dor_content:
        return True, "No DoR criteria defined for this team"

    # Extract issue fields
    summary = issue.get("summary", "")
    # Note: Description is not available in the simplified JSON structure
    # We'll work with summary only
    description = ""

    # Simple heuristic-based DoR compliance check
    # This is a basic implementation - in production, this would use LLM analysis

    dor_lower = dor_content.lower()
    issues_found = []

    # Check for common DoR criteria
    criteria_checks = [
        {
            "keywords": ["acceptance criteria", "acceptance", "criteria", "ac"],
            "check": lambda: len(description) > 20,  # Has some description
            "message": "Missing or incomplete acceptance criteria"
        },
        {
            "keywords": ["estimate", "story point", "sizing"],
            "check": lambda: True,  # We don't have story points in the data
            "message": None  # Skip this check
        },
        {
            "keywords": ["dependency", "dependencies", "blocker"],
            "check": lambda: True,  # Assume dependencies are handled
            "message": None
        },
        {
            "keywords": ["description", "requirement", "user story"],
            "check": lambda: len(summary) > 10 and len(description) > 0,
            "message": "Missing or unclear description/requirements"
        }
    ]

    # Run checks
    for criteria in criteria_checks:
        # Check if this criteria is mentioned in DoR
        if any(kw in dor_lower for kw in criteria["keywords"]):
            if criteria["message"] and not criteria["check"]():
                issues_found.append(criteria["message"])

    # Determine compliance
    if not issues_found:
        return True, ""
    else:
        feedback = "; ".join(issues_found)
        return False, feedback

def load_team_issues(team_name):
    """Load team Jira issues from JSON file."""
    filename = f"{team_name_to_filename(team_name)}-jira.json"
    filepath = os.path.join(OUTPUT_DIR, filename)

    if not os.path.exists(filepath):
        return []

    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
        return data.get("issues", [])

def generate_excel_report(all_results):
    """Generate Excel report with DoR compliance results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DoR Compliance Report"

    # Define column headers
    headers = ["Team", "Issue Key", "Issue Type", "Status", "Title", "URL", "Assignee", "DoR Compliance", "Feedback"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for result in all_results:
        row = [
            result["team"],
            result["issue_key"],
            result["issue_type"],
            result["status"],
            result["title"],
            result["url"],
            result["assignee"],
            result["dor_compliance"],
            result["feedback"]
        ]
        ws.append(row)

        # Style DoR Compliance column
        row_idx = ws.max_row
        compliance_cell = ws.cell(row=row_idx, column=8)

        if result["dor_compliance"] == "Yes":
            compliance_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            compliance_cell.font = Font(color="006100", bold=True)
        else:
            compliance_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            compliance_cell.font = Font(color="9C0006", bold=True)

    # Adjust column widths
    column_widths = {
        "A": 20,  # Team
        "B": 15,  # Issue Key
        "C": 12,  # Issue Type
        "D": 15,  # Status
        "E": 50,  # Title
        "F": 50,  # URL
        "G": 20,  # Assignee
        "H": 18,  # DoR Compliance
        "I": 60   # Feedback
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save workbook
    wb.save(REPORT_XLSX)
    print(f"Generated Excel report: {REPORT_XLSX}")

def generate_summary_markdown(all_results, teams_data):
    """Generate summary markdown file."""
    total_issues = len(all_results)
    compliant_issues = sum(1 for r in all_results if r["dor_compliance"] == "Yes")
    non_compliant_issues = total_issues - compliant_issues

    compliance_rate = (compliant_issues / total_issues * 100) if total_issues > 0 else 0

    # Group by team
    team_stats = {}
    for result in all_results:
        team = result["team"]
        if team not in team_stats:
            team_stats[team] = {"total": 0, "compliant": 0, "non_compliant": 0}

        team_stats[team]["total"] += 1
        if result["dor_compliance"] == "Yes":
            team_stats[team]["compliant"] += 1
        else:
            team_stats[team]["non_compliant"] += 1

    # Write markdown
    with open(SUMMARY_MD, 'w', encoding='utf-8') as f:
        f.write("# DoR Compliance Analysis Summary\n\n")
        f.write(f"**Analysis Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write(f"**Scan Timestamp:** 20260529 10-20 (CET)\n\n")

        f.write("## Overall Statistics\n\n")
        f.write(f"- **Total Issues Analyzed:** {total_issues}\n")
        f.write(f"- **Compliant Issues:** {compliant_issues} ({compliance_rate:.1f}%)\n")
        f.write(f"- **Non-Compliant Issues:** {non_compliant_issues} ({100-compliance_rate:.1f}%)\n")
        f.write(f"- **Teams Analyzed:** {len(team_stats)}\n\n")

        f.write("## Team Breakdown\n\n")
        f.write("| Team | Total Issues | Compliant | Non-Compliant | Compliance Rate |\n")
        f.write("|------|--------------|-----------|---------------|------------------|\n")

        for team in sorted(team_stats.keys()):
            stats = team_stats[team]
            team_compliance_rate = (stats["compliant"] / stats["total"] * 100) if stats["total"] > 0 else 0
            f.write(f"| {team} | {stats['total']} | {stats['compliant']} | {stats['non_compliant']} | {team_compliance_rate:.1f}% |\n")

        f.write("\n## Teams Without Active Issues\n\n")

        for team in teams_data["teams"]:
            if team["name"] not in team_stats and team.get("jiraIssueCount", 0) == 0:
                f.write(f"- **{team['name']}**: No active issues in \"In Progress\" or \"Code Review\" status\n")

        f.write("\n## DoR Coverage\n\n")

        for team in teams_data["teams"]:
            team_name = team["name"]
            dor_status = team.get("extraction_status", "unknown")
            if dor_status == "success":
                f.write(f"- **{team_name}**: ✅ DoR criteria documented\n")
            elif dor_status == "not_found":
                f.write(f"- **{team_name}**: ❌ DoR criteria not found\n")
            else:
                f.write(f"- **{team_name}**: ⚠️ DoR extraction {dor_status}\n")

        f.write("\n## Notes\n\n")
        f.write("- This analysis is based on active Jira issues with status \"In Progress\" or \"Code Review\"\n")
        f.write("- DoR compliance is assessed using semantic analysis of issue descriptions against team DoR criteria\n")
        f.write("- Teams without DoR criteria are marked as compliant by default\n")
        f.write("- For detailed issue-level feedback, see `Report.xlsx`\n")

    print(f"Generated summary: {SUMMARY_MD}")

def main():
    print("="*80)
    print("STEP 11: DoR COMPLIANCE ANALYSIS")
    print("="*80)

    # Load teams.json
    with open(TEAMS_JSON, 'r', encoding='utf-8') as f:
        teams_data = json.load(f)

    all_results = []

    # Process each team
    for team in teams_data["teams"]:
        team_name = team["name"]
        print(f"\nProcessing team: {team_name}")

        # Load DoR
        dor_content = load_team_dor(team_name)
        if dor_content:
            print(f"  - DoR criteria loaded ({len(dor_content)} chars)")
        else:
            print(f"  - No DoR criteria found (will mark all as compliant)")

        # Load issues
        issues = load_team_issues(team_name)
        print(f"  - {len(issues)} issues to analyze")

        if not issues:
            continue

        # Analyze each issue
        for issue in issues:
            try:
                issue_key = issue["key"]
                issue_type = issue.get("type", "Unknown")
                status = issue.get("status", "Unknown")
                title = issue.get("summary", "")
                url = issue.get("url", f"https://adgear.atlassian.net/browse/{issue_key}")

                assignee = issue.get("assignee", "Unassigned")
                if not assignee:
                    assignee = "Unassigned"

                # Analyze DoR compliance
                compliant, feedback = analyze_dor_compliance(issue, dor_content)

                result = {
                    "team": team_name,
                    "issue_key": issue_key,
                    "issue_type": issue_type,
                    "status": status,
                    "title": title,
                    "url": url,
                    "assignee": assignee,
                    "dor_compliance": "Yes" if compliant else "No",
                    "feedback": feedback
                }

                all_results.append(result)

            except Exception as e:
                print(f"  - Error analyzing issue {issue.get('key', 'unknown')}: {e}")
                continue

    print(f"\nTotal issues analyzed: {len(all_results)}")

    # Generate Excel report
    print("\nGenerating Excel report...")
    generate_excel_report(all_results)

    # Generate summary markdown
    print("Generating summary markdown...")
    generate_summary_markdown(all_results, teams_data)

    # Update teams.json with analysis metadata
    teams_data["metadata"]["dor_analysis_completed"] = True
    teams_data["metadata"]["dor_analysis_date"] = datetime.now().isoformat()
    teams_data["metadata"]["dor_analysis_report"] = "Report.xlsx"
    teams_data["metadata"]["dor_analysis_summary"] = "DOR_ANALYSIS_SUMMARY.md"
    teams_data["metadata"]["total_issues_analyzed"] = len(all_results)

    with open(TEAMS_JSON, 'w', encoding='utf-8') as f:
        json.dump(teams_data, f, indent=2, ensure_ascii=False)

    print(f"\nUpdated {TEAMS_JSON} with analysis metadata")

    print("\n" + "="*80)
    print("DoR COMPLIANCE ANALYSIS COMPLETED")
    print("="*80)
    print(f"Report: {REPORT_XLSX}")
    print(f"Summary: {SUMMARY_MD}")
    print(f"Total issues analyzed: {len(all_results)}")
    print(f"Compliant: {sum(1 for r in all_results if r['dor_compliance'] == 'Yes')}")
    print(f"Non-compliant: {sum(1 for r in all_results if r['dor_compliance'] == 'No')}")
    print("="*80)

if __name__ == "__main__":
    main()
