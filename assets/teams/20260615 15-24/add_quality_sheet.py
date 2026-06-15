#!/usr/bin/env python3
"""Step 12: Add 'DoR quality' sheet to Report-DoR.xlsx (3rd sheet)"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"C:\Users\i.klyha\Desktop\Claude\wow-scanner-tool\assets\teams\20260615 15-24"

results = [
    {"team": "Polonium UF", "overall": 84, "note": "Strong coverage; could improve AC format specifics and ownership clarity"},
    {"team": "Bigos", "overall": 72, "note": "Missing explicit risk identification and testing strategy"},
    {"team": "Zurek", "overall": 72, "note": "Missing explicit risk identification and testing strategy"},
    {"team": "Europium", "overall": 71, "note": "Missing estimation criteria and testing strategy"},
    {"team": "Mouflons", "overall": 71, "note": "Missing estimation criteria, scope/sprint fit, and technical feasibility"},
    {"team": "Wolves", "overall": 71, "note": "Missing estimation criteria, scope/sprint fit, and technical feasibility"},
    {"team": "Capybaras", "overall": 60, "note": "Missing design/UX, risks, scope criteria; too much focus on labeling"},
    {"team": "Igni", "overall": 60, "note": "Missing scope/sprint fit, design requirements, risks, testing strategy"},
    {"team": "Abyss", "overall": 59, "note": "Missing scope/sprint fit, risks, testing strategy, stakeholder alignment"},
    {"team": "Copernicium", "overall": 56, "note": "Missing scope/sprint fit, risks, testing strategy; criteria lack specificity"},
    {"team": "EP Core", "overall": 56, "note": "Missing scope, risks, testing; purely passive list with no behavioral guidance"},
    {"team": "Radium", "overall": 52, "note": "Vague criteria ('well described'), missing scope/sprint fit, risks, testing"},
    {"team": "ML Platform Pandas", "overall": 49, "note": "Minimal detail, no ownership defined, criteria too generic to be actionable"},
]

average_quality = round(sum(r["overall"] for r in results) / len(results))

wb = load_workbook(f"{OUTPUT_DIR}/Report-DoR.xlsx")
print(f"Before: {wb.sheetnames}")

ws = wb.create_sheet("DoR quality")

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
orange_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Row 1: KPI
ws.cell(1, 1).value = "DoR quality lvl"
ws.cell(1, 1).font = Font(bold=True, size=11)
ws.cell(1, 2).value = f"{average_quality}%"
ws.cell(1, 2).font = Font(bold=True, size=11, color="CC6600")  # 64% = orange

# Row 3: Headers
for col, (header, width) in enumerate([("Team", 25), ("Quality", 12), ("Note", 60)], 1):
    cell = ws.cell(3, col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width

# Data rows
for row_idx, result in enumerate(results, 4):
    ws.cell(row_idx, 1).value = result["team"]
    ws.cell(row_idx, 1).border = thin_border
    ws.cell(row_idx, 1).alignment = Alignment(vertical='center')

    quality_cell = ws.cell(row_idx, 2)
    quality_cell.value = f"{result['overall']}%"
    quality_cell.border = thin_border
    quality_cell.alignment = Alignment(horizontal='center', vertical='center')
    if result['overall'] >= 70:
        quality_cell.fill = green_fill
    elif result['overall'] >= 40:
        quality_cell.fill = orange_fill
    else:
        quality_cell.fill = red_fill

    note_cell = ws.cell(row_idx, 3)
    note_cell.value = result["note"]
    note_cell.border = thin_border
    note_cell.alignment = Alignment(wrap_text=True, vertical='center')

ws.freeze_panes = 'A4'
wb.save(f"{OUTPUT_DIR}/Report-DoR.xlsx")

print(f"After: {wb.sheetnames}")
print(f"Average DoR Quality: {average_quality}%")
print(f"[SUCCESS] Report-DoR.xlsx now has 3 sheets")
