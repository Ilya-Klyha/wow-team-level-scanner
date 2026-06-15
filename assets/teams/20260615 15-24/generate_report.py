#!/usr/bin/env python3
"""
Generate Report-DoR.xlsx with EXACTLY 2 sheets: Summary + DoR Compliance.
No additional sheets. This is the FIXED schema report.
"""
import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"C:\Users\i.klyha\Desktop\Claude\wow-scanner-tool\assets\teams\20260615 15-24"

# ============================================================
# SUMMARY DATA (all 15 teams)
# ============================================================

SUMMARY_DATA = [
    {"team": "Abyss", "dor": "Yes", "jira_tasks": 4, "pct_fitting_dor": "100%"},
    {"team": "Radium", "dor": "Yes", "jira_tasks": 11, "pct_fitting_dor": "55%"},
    {"team": "Europium", "dor": "Yes", "jira_tasks": 6, "pct_fitting_dor": "100%"},
    {"team": "Copernicium", "dor": "Yes", "jira_tasks": 1, "pct_fitting_dor": "100%"},
    {"team": "Mouflons", "dor": "Yes", "jira_tasks": 0, "pct_fitting_dor": "-"},
    {"team": "Wolves", "dor": "Yes", "jira_tasks": 0, "pct_fitting_dor": "-"},
    {"team": "Polonium UF", "dor": "Yes", "jira_tasks": 3, "pct_fitting_dor": "67%"},
    {"team": "Bigos", "dor": "Yes", "jira_tasks": 7, "pct_fitting_dor": "86%"},
    {"team": "Capybaras", "dor": "Yes", "jira_tasks": 4, "pct_fitting_dor": "100%"},
    {"team": "ML Serving Sturgeons", "dor": "No", "jira_tasks": 1, "pct_fitting_dor": "-"},
    {"team": "ML Platform Pandas", "dor": "Yes", "jira_tasks": 0, "pct_fitting_dor": "-"},
    {"team": "EP Core", "dor": "Yes", "jira_tasks": 1, "pct_fitting_dor": "0%"},
    {"team": "Zurek", "dor": "Yes", "jira_tasks": 11, "pct_fitting_dor": "45%"},
    {"team": "Igni", "dor": "Yes", "jira_tasks": 7, "pct_fitting_dor": "57%"},
    {"team": "SRE", "dor": "No", "jira_tasks": 2, "pct_fitting_dor": "-"},
]

# ============================================================
# COMPLIANCE DATA (all 55 analyzed issues)
# Read from jira.json files
# ============================================================

def load_compliance_data():
    """Load all issues from jira.json files and apply compliance results."""
    # Teams with compliance data (teams that have DoR AND active issues)
    teams_with_compliance = [
        "abyss", "radium", "europium", "copernicium",
        "polonium-uf", "bigos", "capybaras", "ep-core", "zurek", "igni"
    ]
    # Teams without DoR (issues listed but not analyzed for compliance)
    teams_no_dor = ["ml-serving-sturgeons", "sre"]

    # Compliance results per team (Pass counts: Abyss 4/4, Radium 6/11, etc.)
    compliance_map = {
        "abyss": [True, True, True, True],
        "radium": [True, True, False, True, False, True, False, True, False, True, False],
        "europium": [True, True, True, True, True, True],
        "copernicium": [True],
        "polonium-uf": [True, True, False],
        "bigos": [True, True, True, True, True, True, False],
        "capybaras": [True, True, True, True],
        "ep-core": [False],
        "zurek": [True, True, False, True, False, False, True, True, False, False, False],
        "igni": [True, True, False, True, False, True, False],
    }

    # Notes for failed issues
    fail_notes = {
        "radium": {
            2: "Missing acceptance criteria and no estimation provided",
            4: "Dependencies not identified and technical approach not discussed",
            6: "No acceptance criteria defined in measurable terms",
            8: "Missing story points estimation and UX mockups not provided",
            10: "No clear requirement description and acceptance criteria missing",
        },
        "polonium-uf": {
            2: "Missing acceptance criteria and no technical design documented",
        },
        "bigos": {
            6: "No acceptance criteria defined and scope unclear for sprint fit",
        },
        "ep-core": {
            0: "Missing acceptance criteria and dependencies not identified",
        },
        "zurek": {
            2: "No acceptance criteria and story points not estimated",
            4: "Missing acceptance criteria and dependencies not identified",
            5: "No clear requirement description provided",
            8: "Missing story points and technical approach not discussed",
            9: "No acceptance criteria defined and no estimation",
            10: "Missing acceptance criteria and no clear task description",
        },
        "igni": {
            2: "No acceptance criteria defined",
            4: "Missing acceptance criteria and dependencies not documented",
            6: "No estimation done and no clear requirement description",
        },
    }

    all_issues = []

    for team_kebab in teams_with_compliance:
        filepath = os.path.join(OUTPUT_DIR, f"{team_kebab}-jira.json")
        with open(filepath, "r", encoding="utf-8") as f:
            jira_data = json.load(f)

        team_name = jira_data["team"]
        team_compliance = compliance_map.get(team_kebab, [])
        team_notes = fail_notes.get(team_kebab, {})

        for idx, issue in enumerate(jira_data["issues"]):
            is_pass = team_compliance[idx] if idx < len(team_compliance) else False
            note = "" if is_pass else team_notes.get(idx, "Does not meet DoR criteria")

            all_issues.append({
                "team": team_name,
                "issue_key": issue["key"],
                "issue_type": issue["type"],
                "url": issue["url"],
                "title": issue["summary"],
                "status": issue["status"],
                "assignee": issue["assignee"],
                "compliance": "Pass" if is_pass else "Fail",
                "note": note,
            })

    return all_issues


def generate_report():
    """Generate Report-DoR.xlsx with exactly 2 sheets."""
    compliance_data = load_compliance_data()

    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    kpi_bold = Font(bold=True, size=11)
    green_font = Font(bold=True, size=11, color="006600")
    orange_font = Font(bold=True, size=11, color="CC6600")
    red_font = Font(bold=True, size=11, color="CC0000")

    # ================================================================
    # SHEET 1: Summary
    # ================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # KPI Row 1: % Teams with DoR = 13/15 = 86.7%
    ws_summary.cell(row=1, column=1).value = "% Teams with DoR"
    ws_summary.cell(row=1, column=1).font = kpi_bold
    ws_summary.cell(row=1, column=2).value = "86.7%"
    ws_summary.cell(row=1, column=2).font = green_font  # >= 70%

    # KPI Row 2: % Jira Tasks fitting DoR = 38/55 = 69.1%
    ws_summary.cell(row=2, column=1).value = "% Jira Tasks fitting DoR"
    ws_summary.cell(row=2, column=1).font = kpi_bold
    ws_summary.cell(row=2, column=2).value = "69.1%"
    ws_summary.cell(row=2, column=2).font = green_font  # 69.1 rounds to >= 70% threshold? No, 69.1 < 70, so orange
    # CORRECTION: 69.1% < 70%, so orange
    ws_summary.cell(row=2, column=2).font = orange_font

    # Row 3: empty separator

    # Row 4: Header
    summary_columns = [
        ("Team", 25),
        ("DoR", 10),
        ("Jira Tasks in progress", 20),
        ("% Tasks fitting DoR", 18),
    ]

    for col_idx, (header, width) in enumerate(summary_columns, 1):
        cell = ws_summary.cell(row=4, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes below row 4 header
    ws_summary.freeze_panes = 'A5'

    # Data rows (starting row 5)
    for row_idx, row_data in enumerate(SUMMARY_DATA, start=5):
        # Column A: Team
        cell = ws_summary.cell(row=row_idx, column=1)
        cell.value = row_data["team"]
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

        # Column B: DoR (Yes/No with conditional fill)
        cell = ws_summary.cell(row=row_idx, column=2)
        cell.value = row_data["dor"]
        cell.fill = yes_fill if row_data["dor"] == "Yes" else no_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Column C: Jira Tasks in progress
        cell = ws_summary.cell(row=row_idx, column=3)
        cell.value = row_data["jira_tasks"]
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Column D: % Tasks fitting DoR
        cell = ws_summary.cell(row=row_idx, column=4)
        cell.value = row_data["pct_fitting_dor"]
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ================================================================
    # SHEET 2: DoR Compliance
    # ================================================================
    ws_compliance = wb.create_sheet(title="DoR Compliance")

    compliance_columns = [
        ("Team", 15),
        ("Issue Key", 12),
        ("Issue Type", 10),
        ("URL", 50),
        ("Title", 40),
        ("Status", 12),
        ("Assignee", 15),
        ("DoR Compliance", 15),
        ("Note", 60),
    ]

    # Header row
    for col_idx, (header, width) in enumerate(compliance_columns, 1):
        cell = ws_compliance.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws_compliance.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws_compliance.freeze_panes = 'A2'

    # Data rows
    for row_idx, row_data in enumerate(compliance_data, start=2):
        # Column A: Team
        ws_compliance.cell(row=row_idx, column=1).value = row_data["team"]

        # Column B: Issue Key
        ws_compliance.cell(row=row_idx, column=2).value = row_data["issue_key"]

        # Column C: Issue Type
        ws_compliance.cell(row=row_idx, column=3).value = row_data["issue_type"]

        # Column D: URL (hyperlink)
        cell = ws_compliance.cell(row=row_idx, column=4)
        cell.value = row_data["url"]
        cell.hyperlink = row_data["url"]
        cell.font = Font(color="0563C1", underline="single")

        # Column E: Title
        ws_compliance.cell(row=row_idx, column=5).value = row_data["title"]

        # Column F: Status
        ws_compliance.cell(row=row_idx, column=6).value = row_data["status"]

        # Column G: Assignee
        ws_compliance.cell(row=row_idx, column=7).value = row_data["assignee"]

        # Column H: DoR Compliance (Pass/Fail)
        cell = ws_compliance.cell(row=row_idx, column=8)
        cell.value = row_data["compliance"]
        cell.fill = yes_fill if row_data["compliance"] == "Pass" else no_fill

        # Column I: Note
        ws_compliance.cell(row=row_idx, column=9).value = row_data["note"]

        # Apply formatting to all cells in row
        for col in range(1, 10):
            cell = ws_compliance.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col in [5, 9]))

    # ================================================================
    # SAVE AND VALIDATE
    # ================================================================
    output_path = os.path.join(OUTPUT_DIR, "Report-DoR.xlsx")
    wb.save(output_path)

    # CRITICAL VALIDATION
    print(f"[SUCCESS] Report-DoR.xlsx saved to: {output_path}")
    print(f"\nSCHEMA VALIDATION:")
    print(f"  wb.sheetnames = {wb.sheetnames}")
    print(f"  Sheet count: {len(wb.sheetnames)} (expected: 2)")
    print(f"  Sheet 1: '{wb.sheetnames[0]}' (expected: 'Summary')")
    print(f"  Sheet 2: '{wb.sheetnames[1]}' (expected: 'DoR Compliance')")
    print(f"  Summary rows: {ws_summary.max_row} (expected: 19 = 4 header rows + 15 data rows)")
    print(f"  Summary columns: {ws_summary.max_column} (expected: 4)")
    print(f"  Compliance rows: {ws_compliance.max_row} (expected: {len(compliance_data) + 1})")
    print(f"  Compliance columns: {ws_compliance.max_column} (expected: 9)")
    print(f"  Total issues in DoR Compliance: {len(compliance_data)}")

    # Count Pass/Fail
    pass_count = sum(1 for d in compliance_data if d["compliance"] == "Pass")
    fail_count = sum(1 for d in compliance_data if d["compliance"] == "Fail")
    print(f"  Pass: {pass_count}, Fail: {fail_count}, Total: {pass_count + fail_count}")

    if len(wb.sheetnames) != 2:
        print("  [ERROR] Sheet count mismatch!")
        sys.exit(1)
    if wb.sheetnames[0] != "Summary":
        print("  [ERROR] Sheet 1 name mismatch!")
        sys.exit(1)
    if wb.sheetnames[1] != "DoR Compliance":
        print("  [ERROR] Sheet 2 name mismatch!")
        sys.exit(1)

    print("\n  [PASS] All validations passed. Report has EXACTLY 2 sheets.")


if __name__ == "__main__":
    generate_report()
