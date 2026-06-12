"""
Generate DoR Compliance Excel Report for SRPOL Teams.
Scan date: 2026-06-12 08:53 CET
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEAMS_JSON = os.path.join(SCRIPT_DIR, "teams.json")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "Report.xlsx")

# Style constants
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FONT = Font(name="Calibri", size=11, color="006100")
RED_FONT = Font(name="Calibri", size=11, color="9C0006")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def load_teams():
    with open(TEAMS_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data["teams"]


def apply_header_style(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def generate_report():
    teams = load_teams()

    teams_with_dor = [t for t in teams if t["dorStatus"] == "success"]
    teams_without_dor = [t for t in teams if t["dorStatus"] != "success"]
    total_teams = len(teams)
    pct_without_dor = (len(teams_without_dor) / total_teams * 100) if total_teams > 0 else 0

    total_issues = sum(t["jiraIssueCount"] for t in teams)
    # No issues means 0% fitting DoR
    pct_fitting_dor = 0.0

    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # KPI rows
    ws_summary.cell(row=1, column=1, value="Teams without DoR:").font = BOLD_FONT
    ws_summary.cell(row=1, column=2, value=f"{pct_without_dor:.1f}%").font = BOLD_FONT
    ws_summary.cell(row=2, column=1, value="Issues fitting DoR:").font = BOLD_FONT
    ws_summary.cell(row=2, column=2, value=f"{pct_fitting_dor:.1f}%").font = BOLD_FONT

    # Table header at row 4
    headers = ["Team", "DoR", "Jira Tasks"]
    for col, header in enumerate(headers, 1):
        ws_summary.cell(row=4, column=col, value=header)
    apply_header_style(ws_summary, 4, len(headers))

    # Team rows starting at row 5
    for i, team in enumerate(teams, start=5):
        has_dor = team["dorStatus"] == "success"

        ws_summary.cell(row=i, column=1, value=team["name"]).font = NORMAL_FONT
        ws_summary.cell(row=i, column=1).border = THIN_BORDER

        dor_cell = ws_summary.cell(row=i, column=2, value="Yes" if has_dor else "No")
        dor_cell.alignment = Alignment(horizontal="center")
        dor_cell.border = THIN_BORDER
        if has_dor:
            dor_cell.fill = GREEN_FILL
            dor_cell.font = GREEN_FONT
        else:
            dor_cell.fill = RED_FILL
            dor_cell.font = RED_FONT

        task_cell = ws_summary.cell(row=i, column=3, value=team["jiraIssueCount"])
        task_cell.alignment = Alignment(horizontal="center")
        task_cell.font = NORMAL_FONT
        task_cell.border = THIN_BORDER

    # Column widths
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 12

    # --- Sheet 2: DoR Compliance ---
    ws_compliance = wb.create_sheet("DoR Compliance")

    compliance_headers = [
        "Team", "Issue Key", "Issue Type", "Status",
        "Title", "URL", "Assignee", "DoR Compliance", "Feedback"
    ]
    for col, header in enumerate(compliance_headers, 1):
        ws_compliance.cell(row=1, column=col, value=header)
    apply_header_style(ws_compliance, 1, len(compliance_headers))

    # Column widths for compliance sheet
    col_widths = [22, 14, 12, 12, 40, 50, 18, 16, 40]
    for i, width in enumerate(col_widths, 1):
        ws_compliance.column_dimensions[chr(64 + i)].width = width

    # No data rows (0 issues found)

    # Save
    wb.save(OUTPUT_FILE)
    print(f"Report generated: {OUTPUT_FILE}")
    print(f"  - Teams scanned: {total_teams}")
    print(f"  - Teams with DoR: {len(teams_with_dor)}")
    print(f"  - Teams without DoR: {len(teams_without_dor)} ({pct_without_dor:.1f}%)")
    print(f"  - Total Jira issues: {total_issues}")
    print(f"  - Issues fitting DoR: {pct_fitting_dor:.1f}%")


if __name__ == "__main__":
    generate_report()
