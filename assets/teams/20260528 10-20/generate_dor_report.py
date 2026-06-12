#!/usr/bin/env python3
"""
Generate DoR Compliance Excel Report.
Analyzes each Jira issue against team's DoR criteria.
"""

import json
import os
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Attempting to install...")
    import subprocess
    subprocess.check_call(['py', '-m', 'pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True

OUTPUT_DIR = r"C:\Users\i.klyha\Desktop\Claude\wow-scanner-tool\assets\teams\20260528 10-20"

# Teams with both DoR and active issues (need analysis)
TEAMS_TO_ANALYZE = [
    ("PE-WAW-Abyss", "pe-waw-abyss"),
    ("Radium", "radium"),
    ("Europium", "europium"),
    ("Copernicium", "copernicium"),
    ("Polonium UF", "polonium-uf"),
    ("Capybaras", "capybaras"),
    ("EP Core", "ep-core"),
    ("Igni", "igni"),
]

def load_dor(team_slug):
    """Load DoR criteria from file."""
    dor_path = os.path.join(OUTPUT_DIR, f'{team_slug}-dor.txt')
    if not os.path.exists(dor_path):
        return None

    with open(dor_path, 'r', encoding='utf-8') as f:
        content = f.read()

    if 'DoR - STORY/TASK criteria not found' in content:
        return None

    return content

def load_jira_issues(team_slug):
    """Load Jira issues from JSON file."""
    jira_path = os.path.join(OUTPUT_DIR, f'{team_slug}-jira.json')
    if not os.path.exists(jira_path):
        return []

    with open(jira_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    return data.get('issues', [])

def extract_dor_criteria(dor_content):
    """Extract individual DoR criteria from content."""
    if not dor_content:
        return []

    # Simple heuristic: split by lines, look for bullet points or numbered items
    criteria = []
    lines = dor_content.split('\n')
    current_criterion = []

    for line in lines:
        line = line.strip()
        if not line:
            if current_criterion:
                criteria.append(' '.join(current_criterion))
                current_criterion = []
            continue

        # Check if line starts a new criterion (bullet, number, or capital letter)
        if (line.startswith('-') or line.startswith('*') or
            line.startswith('•') or (line[0].isdigit() and line[1:3] in ['. ', ') ']) or
            (len(line) > 0 and line[0].isupper() and current_criterion == [])):

            if current_criterion:
                criteria.append(' '.join(current_criterion))

            # Remove bullet/number prefix
            clean_line = line.lstrip('-*•0123456789. )')
            current_criterion = [clean_line]
        else:
            # Continuation of current criterion
            current_criterion.append(line)

    if current_criterion:
        criteria.append(' '.join(current_criterion))

    # Filter out very short criteria (likely section headers)
    criteria = [c for c in criteria if len(c) > 15]

    return criteria

def analyze_issue_compliance(issue, dor_criteria):
    """
    Analyze if an issue meets DoR criteria.
    This is a simplified heuristic-based analysis.
    In production, this would use LLM-based semantic analysis.
    """

    # Get issue fields
    summary = issue.get('summary', '').lower()
    issue_type = issue.get('type', '')
    status = issue.get('status', '')
    assignee = issue.get('assignee', '')

    # Check for description (we don't have it in our data, so we'll note this)
    has_description = len(summary) > 10  # Placeholder

    # Analyze compliance with common DoR patterns
    findings = []
    compliance_score = 0
    total_criteria = len(dor_criteria)

    for criterion in dor_criteria:
        criterion_lower = criterion.lower()
        is_met = False
        feedback = ""

        # Check for common DoR aspects
        if 'acceptance criteria' in criterion_lower or 'ac' in criterion_lower:
            # We don't have AC data, so we'll mark as "needs verification"
            feedback = "[VERIFY] Acceptance Criteria: Verify AC is defined in issue description"
            is_met = None

        elif 'estimate' in criterion_lower or 'story point' in criterion_lower:
            # We don't have estimate data
            feedback = "[VERIFY] Estimates: Verify story points are assigned"
            is_met = None

        elif 'clear' in criterion_lower and ('description' in criterion_lower or 'requirement' in criterion_lower):
            if has_description:
                feedback = "[OK] Has summary/description"
                is_met = True
                compliance_score += 1
            else:
                feedback = "[MISSING] Missing description"
                is_met = False

        elif 'assignee' in criterion_lower or 'assigned' in criterion_lower:
            if assignee and assignee != 'Unassigned':
                feedback = f"[OK] Assigned to {assignee}"
                is_met = True
                compliance_score += 1
            else:
                feedback = "[MISSING] Not assigned"
                is_met = False

        elif 'dependencies' in criterion_lower:
            feedback = "[VERIFY] Dependencies: Verify dependencies are documented"
            is_met = None

        elif 'techspec' in criterion_lower or 'tech spec' in criterion_lower:
            feedback = "[VERIFY] TechSpec: Verify TechSpec approval if required"
            is_met = None

        elif 'contact' in criterion_lower:
            feedback = "[VERIFY] Contact: Verify contact information for external tasks"
            is_met = None

        elif 'monitoring' in criterion_lower or 'alerting' in criterion_lower:
            feedback = "[VERIFY] Monitoring: Verify monitoring configured if applicable"
            is_met = None

        else:
            # Generic criterion
            feedback = f"[VERIFY] {criterion[:50]}... - Needs verification"
            is_met = None

        findings.append({
            'criterion': criterion,
            'is_met': is_met,
            'feedback': feedback
        })

    # Calculate overall compliance
    verifiable_criteria = sum(1 for f in findings if f['is_met'] is not None)
    if verifiable_criteria > 0:
        compliance_percentage = int((compliance_score / verifiable_criteria) * 100)
    else:
        compliance_percentage = 0

    # Determine overall status
    if compliance_percentage >= 80:
        overall_status = "COMPLIANT"
    elif compliance_percentage >= 50:
        overall_status = "PARTIAL"
    else:
        overall_status = "NEEDS WORK"

    return {
        'overall_status': overall_status,
        'compliance_percentage': compliance_percentage,
        'findings': findings,
        'summary': f"{compliance_score}/{verifiable_criteria} verifiable criteria met"
    }

def create_excel_report():
    """Create Excel report with DoR compliance analysis."""

    print("Generating DoR Compliance Excel Report...")
    print()

    wb = Workbook()
    ws = wb.active
    ws.title = "DoR Compliance"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    compliant_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    needs_work_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write header
    headers = [
        "Team",
        "Issue Key",
        "Issue Type",
        "Status",
        "Summary",
        "Assignee",
        "Priority",
        "DoR Status",
        "Compliance %",
        "Findings Summary",
        "Actionable Feedback"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Adjust column widths
    column_widths = [15, 12, 10, 12, 40, 15, 10, 12, 10, 30, 50]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Process each team
    row = 2
    total_issues = 0
    compliant_issues = 0
    partial_issues = 0
    needs_work_issues = 0

    for team_name, team_slug in TEAMS_TO_ANALYZE:
        print(f"Analyzing {team_name}...")

        # Load DoR
        dor_content = load_dor(team_slug)
        if not dor_content:
            print(f"  Warning: No DoR found for {team_name}")
            continue

        dor_criteria = extract_dor_criteria(dor_content)
        print(f"  Found {len(dor_criteria)} DoR criteria")

        # Load issues
        issues = load_jira_issues(team_slug)
        if not issues:
            print(f"  No issues found for {team_name}")
            continue

        print(f"  Analyzing {len(issues)} issues...")

        # Analyze each issue
        for issue in issues:
            analysis = analyze_issue_compliance(issue, dor_criteria)

            # Write row
            ws.cell(row=row, column=1).value = team_name
            ws.cell(row=row, column=2).value = issue['key']
            ws.cell(row=row, column=3).value = issue['type']
            ws.cell(row=row, column=4).value = issue['status']
            ws.cell(row=row, column=5).value = issue['summary']
            ws.cell(row=row, column=6).value = issue['assignee']
            ws.cell(row=row, column=7).value = issue['priority']
            ws.cell(row=row, column=8).value = analysis['overall_status']
            ws.cell(row=row, column=9).value = analysis['compliance_percentage']
            ws.cell(row=row, column=10).value = analysis['summary']

            # Build actionable feedback
            feedback_items = []
            for finding in analysis['findings']:
                if finding['is_met'] == False:
                    feedback_items.append(finding['feedback'])
                elif finding['is_met'] is None:
                    feedback_items.append(finding['feedback'])

            ws.cell(row=row, column=11).value = '\n'.join(feedback_items[:5]) if feedback_items else "Review all DoR criteria"

            # Apply status color
            status_cell = ws.cell(row=row, column=8)
            if analysis['overall_status'] == 'COMPLIANT':
                status_cell.fill = compliant_fill
                compliant_issues += 1
            elif analysis['overall_status'] == 'PARTIAL':
                status_cell.fill = partial_fill
                partial_issues += 1
            else:
                status_cell.fill = needs_work_fill
                needs_work_issues += 1

            # Apply borders and alignment
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

            row += 1
            total_issues += 1

    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["DoR Compliance Analysis Summary"])
    summary_ws.append([])
    summary_ws.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_ws.append(["Total Teams Analyzed:", len(TEAMS_TO_ANALYZE)])
    summary_ws.append(["Total Issues Analyzed:", total_issues])
    summary_ws.append([])
    summary_ws.append(["Compliance Breakdown:"])
    summary_ws.append(["  Compliant (≥80%):", compliant_issues, f"{int(compliant_issues/total_issues*100) if total_issues > 0 else 0}%"])
    summary_ws.append(["  Partial (50-79%):", partial_issues, f"{int(partial_issues/total_issues*100) if total_issues > 0 else 0}%"])
    summary_ws.append(["  Needs Work (<50%):", needs_work_issues, f"{int(needs_work_issues/total_issues*100) if total_issues > 0 else 0}%"])

    # Style summary sheet
    summary_ws['A1'].font = Font(bold=True, size=14)
    for row_idx in range(3, 11):
        summary_ws.cell(row=row_idx, column=1).font = Font(bold=True)

    # Save workbook
    output_path = os.path.join(OUTPUT_DIR, 'Report.xlsx')
    wb.save(output_path)

    print()
    print(f"[SUCCESS] Excel report created: Report.xlsx")
    print(f"  Total issues analyzed: {total_issues}")
    print(f"  Compliant: {compliant_issues} ({int(compliant_issues/total_issues*100) if total_issues > 0 else 0}%)")
    print(f"  Partial: {partial_issues} ({int(partial_issues/total_issues*100) if total_issues > 0 else 0}%)")
    print(f"  Needs Work: {needs_work_issues} ({int(needs_work_issues/total_issues*100) if total_issues > 0 else 0}%)")

def main():
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl library is required")
        print("Install with: py -m pip install openpyxl")
        return

    create_excel_report()
    print()
    print("Done!")

if __name__ == '__main__':
    main()
