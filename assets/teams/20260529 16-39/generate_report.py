#!/usr/bin/env python3
"""
Generate DoR Compliance Excel Report - FIXED SCHEMA
DO NOT MODIFY THIS SCHEMA
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# FIXED SCHEMA - DO NOT MODIFY
SHEET_NAME = "DoR Compliance"
COLUMNS = [
    ("Team", 15),
    ("Issue Key", 12),
    ("Issue Type", 10),
    ("Status", 12),
    ("Title", 40),
    ("URL", 50),
    ("Assignee", 15),
    ("DoR Compliance", 15),
    ("Feedback", 60)
]

def generate_report(data_json, output_path):
    """Generate Excel report with FIXED schema."""
    data = json.loads(data_json)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write header row
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add data rows - FIXED SCHEMA (9 columns)
    for row_idx, row_data in enumerate(data, start=2):
        # Column A: Team
        ws.cell(row=row_idx, column=1).value = row_data["team"]

        # Column B: Issue Key
        ws.cell(row=row_idx, column=2).value = row_data["issue_key"]

        # Column C: Issue Type
        ws.cell(row=row_idx, column=3).value = row_data["issue_type"]

        # Column D: Status
        ws.cell(row=row_idx, column=4).value = row_data["status"]

        # Column E: Title
        ws.cell(row=row_idx, column=5).value = row_data["title"]

        # Column F: URL (as hyperlink)
        ws.cell(row=row_idx, column=6).value = row_data["url"]
        ws.cell(row=row_idx, column=6).hyperlink = row_data["url"]
        ws.cell(row=row_idx, column=6).style = "Hyperlink"

        # Column G: Assignee
        ws.cell(row=row_idx, column=7).value = row_data["assignee"]

        # Column H: DoR Compliance (Yes/No ONLY)
        compliance_value = row_data["dor_compliance"]
        ws.cell(row=row_idx, column=8).value = compliance_value
        ws.cell(row=row_idx, column=8).fill = yes_fill if compliance_value == "Yes" else no_fill

        # Column I: Feedback (empty if Yes, reason if No)
        ws.cell(row=row_idx, column=9).value = row_data.get("feedback", "")

        # Apply formatting to all cells in row
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            # Wrap text for Title and Feedback columns
            cell.alignment = Alignment(vertical='top', wrap_text=(col in [5, 9]))

    # Save workbook
    wb.save(output_path)
    print(f"[SUCCESS] Excel report saved to: {output_path}")

    # Validate schema
    print("\nSCHEMA VALIDATION:")
    print(f"  - Sheet count: {len(wb.sheetnames)} (expected: 1)")
    print(f"  - Sheet name: '{wb.sheetnames[0]}' (expected: '{SHEET_NAME}')")
    print(f"  - Column count: {ws.max_column} (expected: 9)")
    print(f"  - Data rows: {ws.max_row - 1} (header row excluded)")

    if len(wb.sheetnames) != 1:
        print("  [WARNING] Schema violation: Expected exactly 1 sheet")
    if wb.sheetnames[0] != SHEET_NAME:
        print(f"  [WARNING] Schema violation: Sheet name should be '{SHEET_NAME}'")
    if ws.max_column != 9:
        print(f"  [WARNING] Schema violation: Expected 9 columns, found {ws.max_column}")

if __name__ == "__main__":
    generate_report(sys.argv[1], sys.argv[2])
