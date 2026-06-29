#!/usr/bin/env python3
"""
DoD Scanner - Canonical Report Generator
Produces Report-DoD.xlsx and Report-DoD.html from JSON data files.

Usage:
    python3 generate_dod_reports.py <summary_data.json> <quality_data.json> <output_dir> <scan_date>

This script is the SINGLE SOURCE OF TRUTH for DoD report formatting.
It must NEVER be modified by the LLM during skill execution.
"""
import json
import sys
import os
from pathlib import Path

TEMPLATE_VERSION = "1.0"

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl not installed. Excel report will not be generated.")


# ============================================================
# FIXED SCHEMA CONSTANTS
# ============================================================

SUMMARY_SHEET_NAME = "Summary"
QUALITY_SHEET_NAME = "DoD quality"

# Reference: assets/teams/20260615 15-43/generate-report-dod.py
SUMMARY_COLUMNS = [
    ("Team", 28),
    ("DoD", 12),
]

# Simplified 3-column format (consistent with DoR quality sheet pattern)
QUALITY_COLUMNS = [
    ("Team", 25),
    ("Quality", 12),
    ("Note", 60),
]

# Thresholds from reference scripts
# Summary KPI: green >= 80, yellow >= 50, red < 50
SUMMARY_KPI_GREEN = 80
SUMMARY_KPI_YELLOW = 50

# Quality scores: green >= 75, yellow >= 50, red < 50
QUALITY_GREEN = 75
QUALITY_YELLOW = 50


# ============================================================
# EXCEL REPORT GENERATION
# ============================================================

def generate_excel(summary_data, quality_data, output_path):
    """Generate Report-DoD.xlsx with 2 sheets matching reference format."""
    if not HAS_OPENPYXL:
        print("[ERROR] Cannot generate Excel report without openpyxl.")
        return False

    wb = Workbook()

    # Reference styles from generate-report-dod.py
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    kpi_label_font = Font(name="Calibri", bold=True, size=11)
    kpi_value_font = Font(name="Calibri", bold=True, size=14)
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yes_font = Font(name="Calibri", color="006100")
    no_font = Font(name="Calibri", color="9C0006")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Calculate KPI
    total_teams = len(summary_data)
    teams_with_dod = sum(1 for t in summary_data if t["dod"] == "Yes")
    pct = round(teams_with_dod / total_teams * 100, 1) if total_teams > 0 else 0

    # === SHEET 1: Summary ===
    ws = wb.active
    ws.title = SUMMARY_SHEET_NAME

    # KPI row (reference: merged A1:B1, value in C1 with conditional fill)
    ws.merge_cells("A1:B1")
    ws["A1"] = "% Teams with DoD"
    ws["A1"].font = kpi_label_font
    ws["C1"] = f"{pct}%"
    ws["C1"].font = kpi_value_font
    if pct >= SUMMARY_KPI_GREEN:
        ws["C1"].fill = yes_fill
    elif pct >= SUMMARY_KPI_YELLOW:
        ws["C1"].fill = yellow_fill
    else:
        ws["C1"].fill = no_fill

    # Row 2: empty separator

    # Row 3: Table header
    for col_idx, (header, width) in enumerate(SUMMARY_COLUMNS, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Extra column width for KPI value
    ws.column_dimensions["C"].width = 12

    # Data rows (starting at row 4)
    for row_idx, team in enumerate(summary_data, 4):
        ws.cell(row=row_idx, column=1, value=team["team"]).border = thin_border
        dod_cell = ws.cell(row=row_idx, column=2, value=team["dod"])
        dod_cell.border = thin_border
        dod_cell.alignment = Alignment(horizontal="center")
        if team["dod"] == "Yes":
            dod_cell.fill = yes_fill
            dod_cell.font = yes_font
        else:
            dod_cell.fill = no_fill
            dod_cell.font = no_font

    ws.freeze_panes = "A4"

    # === SHEET 2: DoD quality ===
    ws_q = wb.create_sheet(title=QUALITY_SHEET_NAME)

    # Quality score styles (from add_dod_quality_sheet.py thresholds)
    green_fill_q = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    orange_fill_q = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill_q = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_font_q = Font(name="Calibri", color="006100")
    orange_font_q = Font(name="Calibri", color="9C5700")
    red_font_q = Font(name="Calibri", color="9C0006")
    na_font = Font(name="Calibri", italic=True, color="808080")

    # KPI Row 1
    if quality_data:
        avg_quality = round(sum(r["overall"] for r in quality_data) / len(quality_data), 1)
    else:
        avg_quality = 0

    ws_q.cell(1, 1).value = "DoD quality lvl"
    ws_q.cell(1, 1).font = kpi_label_font
    ws_q.cell(1, 2).value = f"{avg_quality}%"
    ws_q.cell(1, 2).font = kpi_value_font
    if avg_quality >= QUALITY_GREEN:
        ws_q.cell(1, 2).fill = green_fill_q
    elif avg_quality >= QUALITY_YELLOW:
        ws_q.cell(1, 2).fill = orange_fill_q
    else:
        ws_q.cell(1, 2).fill = red_fill_q

    # Row 3: Headers
    for col_idx, (header, width) in enumerate(QUALITY_COLUMNS, 1):
        cell = ws_q.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws_q.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows (sorted by overall descending - already sorted in input)
    sorted_quality = sorted(quality_data, key=lambda x: x["overall"], reverse=True)

    for row_idx, result in enumerate(sorted_quality, 4):
        ws_q.cell(row_idx, 1).value = result["team"]
        ws_q.cell(row_idx, 1).border = thin_border
        ws_q.cell(row_idx, 1).alignment = Alignment(vertical="center")

        quality_cell = ws_q.cell(row_idx, 2)
        quality_cell.value = f"{result['overall']}%"
        quality_cell.border = thin_border
        quality_cell.alignment = Alignment(horizontal="center", vertical="center")
        score = result["overall"]
        if score >= QUALITY_GREEN:
            quality_cell.fill = green_fill_q
            quality_cell.font = green_font_q
        elif score >= QUALITY_YELLOW:
            quality_cell.fill = orange_fill_q
            quality_cell.font = orange_font_q
        else:
            quality_cell.fill = red_fill_q
            quality_cell.font = red_font_q

        note_cell = ws_q.cell(row_idx, 3)
        note_cell.value = result["note"]
        note_cell.border = thin_border
        note_cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Teams without DoD (shown as N/A)
    teams_with_dod_names = {r["team"] for r in quality_data}
    no_dod_teams = [t["team"] for t in summary_data if t["dod"] == "No"]
    next_row = len(sorted_quality) + 4

    for team_name in no_dod_teams:
        ws_q.cell(next_row, 1).value = team_name
        ws_q.cell(next_row, 1).border = thin_border
        for col in range(2, 4):
            cell = ws_q.cell(next_row, col)
            cell.value = "N/A"
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = na_font
        next_row += 1

    # Average row
    if quality_data:
        next_row += 1
        ws_q.cell(next_row, 1).value = "Average (teams with DoD)"
        ws_q.cell(next_row, 1).font = Font(name="Calibri", bold=True)
        ws_q.cell(next_row, 1).border = thin_border
        avg_cell = ws_q.cell(next_row, 2)
        avg_cell.value = f"{avg_quality}%"
        avg_cell.font = Font(name="Calibri", bold=True, size=12)
        avg_cell.border = thin_border
        avg_cell.alignment = Alignment(horizontal="center")
        if avg_quality >= QUALITY_GREEN:
            avg_cell.fill = green_fill_q
        elif avg_quality >= QUALITY_YELLOW:
            avg_cell.fill = orange_fill_q
        else:
            avg_cell.fill = red_fill_q

    ws_q.freeze_panes = "A4"

    # Save
    wb.save(output_path)
    return True


# ============================================================
# HTML REPORT GENERATION
# ============================================================

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>DoD Scanner Report - SRPOL Teams</title>
  <style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
      background: #f0f2f5;
      color: #1a1a2e;
      line-height: 1.5;
      padding: 24px;
    }}
    .container {{ max-width: 1400px; margin: 0 auto; }}
    header {{
      background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
      color: #fff;
      padding: 32px 40px;
      border-radius: 12px;
      margin-bottom: 24px;
      box-shadow: 0 4px 20px rgba(0,0,0,0.15);
    }}
    header h1 {{ font-size: 28px; font-weight: 700; margin-bottom: 4px; }}
    header .subtitle {{ font-size: 14px; opacity: 0.8; }}
    .kpi-grid {{
      display: grid;
      grid-template-columns: repeat(2, 1fr);
      gap: 20px;
      margin-bottom: 32px;
    }}
    .kpi-card {{
      background: #fff;
      border-radius: 10px;
      padding: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.06);
      border-top: 4px solid #ccc;
    }}
    .kpi-card.good {{ border-top-color: #27ae60; }}
    .kpi-card.warn {{ border-top-color: #f39c12; }}
    .kpi-card.bad {{ border-top-color: #e74c3c; }}
    .kpi-card .kpi-label {{
      font-size: 13px;
      text-transform: uppercase;
      letter-spacing: 0.5px;
      color: #666;
      margin-bottom: 8px;
      font-weight: 600;
    }}
    .kpi-card .kpi-value {{
      font-size: 36px;
      font-weight: 800;
    }}
    .kpi-card .kpi-detail {{
      font-size: 12px;
      color: #888;
      margin-top: 4px;
    }}
    .kpi-value.good {{ color: #27ae60; }}
    .kpi-value.warn {{ color: #f39c12; }}
    .kpi-value.bad {{ color: #e74c3c; }}
    .section {{
      background: #fff;
      border-radius: 10px;
      padding: 24px;
      margin-bottom: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.06);
    }}
    .section h2 {{
      font-size: 18px;
      font-weight: 700;
      margin-bottom: 16px;
      padding-bottom: 12px;
      border-bottom: 2px solid #f0f2f5;
      color: #1a1a2e;
    }}
    .section h2 .count {{
      font-size: 14px;
      font-weight: 400;
      color: #888;
      margin-left: 8px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }}
    thead th {{
      background: #1a1a2e;
      color: #fff;
      padding: 10px 12px;
      text-align: left;
      font-weight: 600;
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: 0.3px;
      white-space: nowrap;
    }}
    thead th.center {{ text-align: center; }}
    tbody td {{
      padding: 10px 12px;
      border-bottom: 1px solid #eee;
      vertical-align: middle;
    }}
    tbody tr:hover {{ background: #f8f9fc; }}
    tbody tr:last-child td {{ border-bottom: none; }}
    .center {{ text-align: center; }}
    .team-name {{ font-weight: 600; white-space: nowrap; }}
    .note-col {{ font-size: 12px; color: #555; }}
    .badge {{
      display: inline-block;
      padding: 3px 10px;
      border-radius: 12px;
      font-size: 11px;
      font-weight: 700;
      text-transform: uppercase;
    }}
    .badge-yes {{ background: #d4edda; color: #155724; }}
    .badge-no {{ background: #f8d7da; color: #721c24; }}
    .score {{ font-weight: 700; }}
    .score.good {{ color: #27ae60; }}
    .score.warn {{ color: #f39c12; }}
    .score.bad {{ color: #e74c3c; }}
    .bar-container {{
      width: 100%;
      height: 20px;
      background: #eee;
      border-radius: 4px;
      overflow: hidden;
    }}
    .bar {{
      height: 100%;
      border-radius: 4px;
    }}
    .bar.good {{ background: linear-gradient(90deg, #27ae60, #2ecc71); }}
    .bar.warn {{ background: linear-gradient(90deg, #f39c12, #f1c40f); }}
    .bar.bad {{ background: linear-gradient(90deg, #e74c3c, #c0392b); }}
    a {{ color: #2980b9; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    footer {{
      text-align: center;
      padding: 16px;
      font-size: 12px;
      color: #999;
      margin-top: 8px;
    }}
    @media (max-width: 768px) {{
      .kpi-grid {{ grid-template-columns: 1fr; }}
      body {{ padding: 12px; }}
      header {{ padding: 20px; }}
      .section {{ padding: 16px; }}
      table {{ font-size: 11px; }}
    }}
  </style>
</head>
<body>
  <div class="container">
    <header>
      <h1>DoD Scanner Report</h1>
      <div class="subtitle">SRPOL Teams - Definition of Done Analysis | Scan Date: {scan_date} | Teams: {team_count}</div>
    </header>

    <div class="kpi-grid">
      <div class="kpi-card {kpi1_card_class}">
        <div class="kpi-label">Teams with DoD Defined</div>
        <div class="kpi-value {kpi1_value_class}">{kpi1_value}</div>
        <div class="kpi-detail">{kpi1_detail}</div>
      </div>
      <div class="kpi-card {kpi2_card_class}">
        <div class="kpi-label">DoD Quality Level</div>
        <div class="kpi-value {kpi2_value_class}">{kpi2_value}</div>
        <div class="kpi-detail">{kpi2_detail}</div>
      </div>
    </div>

    <div class="section">
      <h2>Team Overview<span class="count">({team_count} teams)</span></h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th class="center">DoD</th>
          </tr>
        </thead>
        <tbody>
{team_overview_rows}
        </tbody>
      </table>
    </div>

    <div class="section">
      <h2>DoD Quality Assessment<span class="count">({quality_count} teams assessed)</span></h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th class="center">Score</th>
            <th>Quality</th>
            <th>Note</th>
          </tr>
        </thead>
        <tbody>
{quality_rows}
        </tbody>
      </table>
    </div>

    <footer>
      DoD Scanner v{version} | Generated by wow-dod-scanner skill | Data source: SRPOL Teams Confluence
    </footer>
  </div>
</body>
</html>"""


def _color_class(value):
    """Return CSS class based on threshold: good >= 70, warn 40-69, bad < 40."""
    if value >= 70:
        return "good"
    elif value >= 40:
        return "warn"
    else:
        return "bad"


def _escape_html(text):
    """Escape HTML special characters."""
    if not text:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def generate_html(summary_data, quality_data, scan_date, output_path):
    """Generate Report-DoD.html using the fixed template."""
    team_count = len(summary_data)

    # KPI 1: Teams with DoD
    teams_with_dod = sum(1 for t in summary_data if t["dod"] == "Yes")
    pct_teams = round(teams_with_dod / team_count * 100, 1) if team_count > 0 else 0
    kpi1_card_class = _color_class(pct_teams)
    kpi1_value_class = _color_class(pct_teams)
    kpi1_value = f"{pct_teams}%"
    kpi1_detail = f"{teams_with_dod} of {team_count} teams"

    # KPI 2: DoD Quality Level
    if quality_data:
        avg_quality = round(sum(r["overall"] for r in quality_data) / len(quality_data), 1)
        kpi2_card_class = _color_class(avg_quality)
        kpi2_value_class = _color_class(avg_quality)
        kpi2_value = f"{avg_quality}%"
        kpi2_detail = f"Across {len(quality_data)} assessed teams"
    else:
        kpi2_card_class = ""
        kpi2_value_class = ""
        kpi2_value = "-"
        kpi2_detail = "No teams assessed"

    # Team Overview rows
    overview_lines = []
    for team in summary_data:
        dod_badge = "badge-yes" if team["dod"] == "Yes" else "badge-no"
        dod_text = team["dod"]
        overview_lines.append(
            f'        <tr>\n'
            f'          <td class="team-name">{_escape_html(team["team"])}</td>\n'
            f'          <td class="center"><span class="badge {dod_badge}">{dod_text}</span></td>\n'
            f'        </tr>'
        )
    team_overview_rows = "\n".join(overview_lines)

    # Quality Assessment rows (sorted by overall desc)
    sorted_quality = sorted(quality_data, key=lambda x: x["overall"], reverse=True)
    quality_count = len(sorted_quality)
    quality_lines = []
    for q in sorted_quality:
        score = q["overall"]
        bar_class = _color_class(score)
        score_cls = _color_class(score)
        quality_lines.append(
            f'        <tr>\n'
            f'          <td class="team-name">{_escape_html(q["team"])}</td>\n'
            f'          <td class="center"><strong class="score {score_cls}">{score}%</strong></td>\n'
            f'          <td>\n'
            f'            <div class="bar-container">\n'
            f'              <div class="bar {bar_class}" style="width: {score}%"></div>\n'
            f'            </div>\n'
            f'          </td>\n'
            f'          <td class="note-col">{_escape_html(q["note"])}</td>\n'
            f'        </tr>'
        )
    quality_rows = "\n".join(quality_lines)

    # Render template
    html = HTML_TEMPLATE.format(
        scan_date=scan_date,
        team_count=team_count,
        version=TEMPLATE_VERSION,
        kpi1_card_class=kpi1_card_class,
        kpi1_value_class=kpi1_value_class,
        kpi1_value=kpi1_value,
        kpi1_detail=kpi1_detail,
        kpi2_card_class=kpi2_card_class,
        kpi2_value_class=kpi2_value_class,
        kpi2_value=kpi2_value,
        kpi2_detail=kpi2_detail,
        team_overview_rows=team_overview_rows,
        quality_count=quality_count,
        quality_rows=quality_rows,
    )

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return True


# ============================================================
# MAIN ENTRY POINT
# ============================================================

def main():
    print(f"[INFO] DoD Report Generator v{TEMPLATE_VERSION}")

    if len(sys.argv) < 5:
        print("Usage: python3 generate_dod_reports.py <summary.json> <quality.json> <output_dir> <scan_date>")
        print("Example: python3 generate_dod_reports.py summary_data.json quality_data.json . 2026-06-29")
        sys.exit(1)

    summary_path = sys.argv[1]
    quality_path = sys.argv[2]
    output_dir = sys.argv[3]
    scan_date = sys.argv[4]

    # Load data
    with open(summary_path, 'r', encoding='utf-8') as f:
        summary_data = json.load(f)
    with open(quality_path, 'r', encoding='utf-8') as f:
        quality_data = json.load(f)

    print(f"[INFO] Loaded data: {len(summary_data)} teams, {len(quality_data)} quality assessments")

    # Generate Excel
    xlsx_path = os.path.join(output_dir, "Report-DoD.xlsx")
    if generate_excel(summary_data, quality_data, xlsx_path):
        print(f"[SUCCESS] Excel report: {xlsx_path}")
        print(f"  Sheets: [{SUMMARY_SHEET_NAME}, {QUALITY_SHEET_NAME}]")
    else:
        print(f"[WARNING] Excel report not generated (openpyxl missing)")

    # Generate HTML
    html_path = os.path.join(output_dir, "Report-DoD.html")
    if generate_html(summary_data, quality_data, scan_date, html_path):
        print(f"[SUCCESS] HTML report: {html_path}")
    else:
        print(f"[ERROR] HTML report generation failed")

    # Summary
    total_teams = len(summary_data)
    teams_with_dod = sum(1 for t in summary_data if t["dod"] == "Yes")
    pct = round(teams_with_dod / total_teams * 100, 1) if total_teams > 0 else 0
    avg_quality = round(sum(r["overall"] for r in quality_data) / len(quality_data), 1) if quality_data else 0

    print(f"\n=== Report Generation Complete ===")
    print(f"  Teams: {total_teams}")
    print(f"  Teams with DoD: {teams_with_dod} ({pct}%)")
    print(f"  Avg DoD Quality: {avg_quality}%")

    # Schema validation
    print(f"\nSCHEMA VALIDATION:")
    errors = 0
    if len(summary_data) != 15:
        print(f"  [WARNING] summary_data has {len(summary_data)} entries (expected 15)")
        errors += 1
    for t in summary_data:
        if t.get("dod") not in ("Yes", "No"):
            print(f"  [WARNING] Team '{t.get('team')}' has invalid dod value: '{t.get('dod')}'")
            errors += 1
    for q in quality_data:
        if not (0 <= q.get("overall", -1) <= 100):
            print(f"  [WARNING] Team '{q.get('team')}' has invalid overall score: {q.get('overall')}")
            errors += 1
    if errors == 0:
        print(f"  [OK] All validations passed")
    else:
        print(f"  [{errors} warning(s)]")


if __name__ == "__main__":
    main()
