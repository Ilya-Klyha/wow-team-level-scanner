#!/usr/bin/env python3
"""
DoR Scanner - Canonical Report Generator
Produces Report-DoR.xlsx and Report-DoR.html from JSON data files.

Usage:
    python3 generate_reports.py <summary_data.json> <compliance_data.json> <quality_data.json> <output_dir> <scan_date>

This script is the SINGLE SOURCE OF TRUTH for report formatting.
It must NEVER be modified by the LLM during skill execution.
"""
import json
import sys
import os
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl not installed. Excel report will not be generated.")


# ============================================================
# FIXED SCHEMA CONSTANTS
# ============================================================

SUMMARY_SHEET_NAME = "Summary"
COMPLIANCE_SHEET_NAME = "DoR Compliance"
QUALITY_SHEET_NAME = "DoR quality"

SUMMARY_COLUMNS = [
    ("Team", 25),
    ("DoR", 10),
    ("Jira Tasks in progress", 20),
    ("% Tasks fitting DoR", 18),
]

COMPLIANCE_COLUMNS = [
    ("Team", 15),
    ("Issue Key", 12),
    ("Issue Type", 10),
    ("URL", 50),
    ("Title", 40),
    ("Status", 12),
    ("Assignee", 15),
    ("DoR Compliance", 15),
    ("Note", 60),
]

QUALITY_COLUMNS = [
    ("Team", 25),
    ("Coverage", 12),
    ("Note", 60),
]


# ============================================================
# EXCEL REPORT GENERATION
# ============================================================

def generate_excel(compliance_data, summary_data, quality_data, output_path):
    """Generate Report-DoR.xlsx with 3 sheets matching fixed schema."""
    if not HAS_OPENPYXL:
        print("[ERROR] Cannot generate Excel report without openpyxl.")
        return False

    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Calculate KPIs
    total_teams = len(summary_data)
    teams_with_dor_count = sum(1 for t in summary_data if t["dor"] == "Yes")
    teams_with_dor_pct = round(teams_with_dor_count / total_teams * 100, 1) if total_teams > 0 else 0

    total_issues = len(compliance_data)
    compliant_issues = sum(1 for d in compliance_data if d["dor_compliance"] == "Pass")
    compliance_pct = round(compliant_issues / total_issues * 100, 1) if total_issues > 0 else 0

    # KPI fonts
    kpi_bold = Font(bold=True, size=11)
    red_font = Font(bold=True, size=11, color="CC0000")
    green_font = Font(bold=True, size=11, color="006600")
    orange_font = Font(bold=True, size=11, color="CC6600")

    # === SHEET 1: Summary ===
    ws_summary = wb.active
    ws_summary.title = SUMMARY_SHEET_NAME

    # Row 1: % Teams with DoR
    ws_summary.cell(row=1, column=1).value = "% Teams with DoR"
    ws_summary.cell(row=1, column=1).font = kpi_bold
    ws_summary.cell(row=1, column=2).value = f"{teams_with_dor_pct}%"
    if teams_with_dor_pct >= 70:
        ws_summary.cell(row=1, column=2).font = green_font
    elif teams_with_dor_pct >= 40:
        ws_summary.cell(row=1, column=2).font = orange_font
    else:
        ws_summary.cell(row=1, column=2).font = red_font

    # Row 2: % Jira Tasks fitting DoR
    ws_summary.cell(row=2, column=1).value = "% Jira Tasks fitting DoR"
    ws_summary.cell(row=2, column=1).font = kpi_bold
    ws_summary.cell(row=2, column=2).value = f"{compliance_pct}%"
    if compliance_pct >= 70:
        ws_summary.cell(row=2, column=2).font = green_font
    elif compliance_pct >= 40:
        ws_summary.cell(row=2, column=2).font = orange_font
    else:
        ws_summary.cell(row=2, column=2).font = red_font

    # Row 3: empty separator
    # Row 4: Table header
    for col_idx, (header, width) in enumerate(SUMMARY_COLUMNS, 1):
        cell = ws_summary.cell(row=4, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

    ws_summary.freeze_panes = 'A5'

    # Summary data rows (row 5+)
    for row_idx, row_data in enumerate(summary_data, start=5):
        ws_summary.cell(row=row_idx, column=1).value = row_data["team"]
        ws_summary.cell(row=row_idx, column=1).border = thin_border
        ws_summary.cell(row=row_idx, column=1).alignment = Alignment(vertical='center')

        dor_value = row_data["dor"]
        ws_summary.cell(row=row_idx, column=2).value = dor_value
        ws_summary.cell(row=row_idx, column=2).fill = yes_fill if dor_value == "Yes" else no_fill
        ws_summary.cell(row=row_idx, column=2).border = thin_border
        ws_summary.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center', vertical='center')

        ws_summary.cell(row=row_idx, column=3).value = row_data["jira_tasks"]
        ws_summary.cell(row=row_idx, column=3).border = thin_border
        ws_summary.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center', vertical='center')

        team_pct = row_data.get("pct_fitting_dor", "-")
        ws_summary.cell(row=row_idx, column=4).value = team_pct
        ws_summary.cell(row=row_idx, column=4).border = thin_border
        ws_summary.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center', vertical='center')

    # === SHEET 2: DoR Compliance ===
    ws = wb.create_sheet(title=COMPLIANCE_SHEET_NAME)

    for col_idx, (header, width) in enumerate(COMPLIANCE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'

    for row_idx, row_data in enumerate(compliance_data, start=2):
        ws.cell(row=row_idx, column=1).value = row_data["team"]
        ws.cell(row=row_idx, column=2).value = row_data["issue_key"]
        ws.cell(row=row_idx, column=3).value = row_data["issue_type"]
        ws.cell(row=row_idx, column=4).value = row_data["url"]
        ws.cell(row=row_idx, column=4).hyperlink = row_data["url"]
        ws.cell(row=row_idx, column=4).style = "Hyperlink"
        ws.cell(row=row_idx, column=5).value = row_data["title"]
        ws.cell(row=row_idx, column=6).value = row_data["status"]
        ws.cell(row=row_idx, column=7).value = row_data["assignee"]

        compliance_value = row_data["dor_compliance"]
        ws.cell(row=row_idx, column=8).value = compliance_value
        ws.cell(row=row_idx, column=8).fill = yes_fill if compliance_value == "Pass" else no_fill

        ws.cell(row=row_idx, column=9).value = row_data.get("note", "")

        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col in [5, 9]))

    # === SHEET 3: DoR quality ===
    ws_q = wb.create_sheet(title=QUALITY_SHEET_NAME)

    # KPI Row 1
    avg_coverage = round(sum(r["coverage"] for r in quality_data) / len(quality_data)) if quality_data else 0
    ws_q.cell(1, 1).value = "DoR quality lvl"
    ws_q.cell(1, 1).font = Font(bold=True, size=11)
    ws_q.cell(1, 2).value = avg_coverage
    if avg_coverage >= 70:
        ws_q.cell(1, 2).font = Font(bold=True, size=11, color="006600")
    elif avg_coverage >= 40:
        ws_q.cell(1, 2).font = Font(bold=True, size=11, color="CC6600")
    else:
        ws_q.cell(1, 2).font = Font(bold=True, size=11, color="CC0000")

    # Row 3: header
    for col, (header, width) in enumerate(QUALITY_COLUMNS, 1):
        cell = ws_q.cell(3, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws_q.column_dimensions[get_column_letter(col)].width = width

    # Sort by coverage desc
    sorted_quality = sorted(quality_data, key=lambda x: x["coverage"], reverse=True)

    for row_idx, result in enumerate(sorted_quality, 4):
        ws_q.cell(row_idx, 1).value = result["team"]
        ws_q.cell(row_idx, 1).border = thin_border
        ws_q.cell(row_idx, 1).alignment = Alignment(vertical='center')

        coverage_cell = ws_q.cell(row_idx, 2)
        coverage_cell.value = result["coverage"]
        coverage_cell.border = thin_border
        coverage_cell.alignment = Alignment(horizontal='center', vertical='center')
        if result['coverage'] >= 70:
            coverage_cell.fill = yes_fill
        elif result['coverage'] >= 40:
            coverage_cell.fill = orange_fill
        else:
            coverage_cell.fill = no_fill

        note_cell = ws_q.cell(row_idx, 3)
        note_cell.value = result["note"]
        note_cell.border = thin_border
        note_cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws_q.freeze_panes = 'A4'

    # Save
    wb.save(output_path)
    return True


# ============================================================
# HTML REPORT GENERATION
# ============================================================

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>DoR Scanner Report - SRPOL Teams</title>
  <style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
      background: #f0f2f5;
      color: #1a1a2e;
      line-height: 1.5;
      padding: 24px;
    }}
    .container {{ max-width: 1400px; margin: 0 auto; }}
    header {{
      background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
      color: #fff;
      padding: 32px 40px;
      border-radius: 12px;
      margin-bottom: 24px;
      box-shadow: 0 4px 20px rgba(0,0,0,0.15);
    }}
    header h1 {{ font-size: 28px; font-weight: 700; margin-bottom: 4px; }}
    header .subtitle {{ font-size: 14px; opacity: 0.8; }}
    .kpi-grid {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 20px;
      margin-bottom: 32px;
    }}
    .kpi-card {{
      background: #fff;
      border-radius: 10px;
      padding: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.06);
      border-top: 4px solid #ccc;
    }}
    .kpi-card.good {{ border-top-color: #27ae60; }}
    .kpi-card.warn {{ border-top-color: #f39c12; }}
    .kpi-card.bad {{ border-top-color: #e74c3c; }}
    .kpi-card .kpi-label {{
      font-size: 13px;
      text-transform: uppercase;
      letter-spacing: 0.5px;
      color: #666;
      margin-bottom: 8px;
      font-weight: 600;
    }}
    .kpi-card .kpi-value {{
      font-size: 36px;
      font-weight: 800;
    }}
    .kpi-card .kpi-detail {{
      font-size: 12px;
      color: #888;
      margin-top: 4px;
    }}
    .kpi-value.good {{ color: #27ae60; }}
    .kpi-value.warn {{ color: #f39c12; }}
    .kpi-value.bad {{ color: #e74c3c; }}
    .section {{
      background: #fff;
      border-radius: 10px;
      padding: 24px;
      margin-bottom: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.06);
    }}
    .section h2 {{
      font-size: 18px;
      font-weight: 700;
      margin-bottom: 16px;
      padding-bottom: 12px;
      border-bottom: 2px solid #f0f2f5;
      color: #1a1a2e;
    }}
    .section h2 .count {{
      font-size: 14px;
      font-weight: 400;
      color: #888;
      margin-left: 8px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }}
    thead th {{
      background: #1a1a2e;
      color: #fff;
      padding: 10px 12px;
      text-align: left;
      font-weight: 600;
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: 0.3px;
      white-space: nowrap;
    }}
    thead th.center {{ text-align: center; }}
    tbody td {{
      padding: 10px 12px;
      border-bottom: 1px solid #eee;
      vertical-align: middle;
    }}
    tbody tr:hover {{ background: #f8f9fc; }}
    tbody tr:last-child td {{ border-bottom: none; }}
    .center {{ text-align: center; }}
    .team-name {{ font-weight: 600; white-space: nowrap; }}
    .title-col {{ max-width: 350px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
    .note-col {{ font-size: 12px; color: #555; }}
    .badge {{
      display: inline-block;
      padding: 3px 10px;
      border-radius: 12px;
      font-size: 11px;
      font-weight: 700;
      text-transform: uppercase;
    }}
    .badge-yes {{ background: #d4edda; color: #155724; }}
    .badge-no {{ background: #f8d7da; color: #721c24; }}
    .score {{ font-weight: 700; }}
    .score.good {{ color: #27ae60; }}
    .score.warn {{ color: #f39c12; }}
    .score.bad {{ color: #e74c3c; }}
    .score.na {{ color: #aaa; }}
    .bar-container {{
      width: 100%;
      height: 20px;
      background: #eee;
      border-radius: 4px;
      overflow: hidden;
    }}
    .bar {{
      height: 100%;
      border-radius: 4px;
    }}
    .bar.good {{ background: linear-gradient(90deg, #27ae60, #2ecc71); }}
    .bar.warn {{ background: linear-gradient(90deg, #f39c12, #f1c40f); }}
    .bar.bad {{ background: linear-gradient(90deg, #e74c3c, #c0392b); }}
    a {{ color: #2980b9; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    footer {{
      text-align: center;
      padding: 16px;
      font-size: 12px;
      color: #999;
      margin-top: 8px;
    }}
    @media (max-width: 768px) {{
      .kpi-grid {{ grid-template-columns: 1fr; }}
      body {{ padding: 12px; }}
      header {{ padding: 20px; }}
      .section {{ padding: 16px; }}
      table {{ font-size: 11px; }}
    }}
  </style>
</head>
<body>
  <div class="container">
    <header>
      <h1>DoR Scanner Report</h1>
      <div class="subtitle">SRPOL Teams - Definition of Ready Analysis | Scan Date: {scan_date} | Teams: {team_count}</div>
    </header>

    <div class="kpi-grid">
      <div class="kpi-card {kpi1_card_class}">
        <div class="kpi-label">Teams with DoR Defined</div>
        <div class="kpi-value {kpi1_value_class}">{kpi1_value}</div>
        <div class="kpi-detail">{kpi1_detail}</div>
      </div>
      <div class="kpi-card {kpi2_card_class}">
        <div class="kpi-label">Tasks Fitting DoR</div>
        <div class="kpi-value {kpi2_value_class}">{kpi2_value}</div>
        <div class="kpi-detail">{kpi2_detail}</div>
      </div>
      <div class="kpi-card {kpi3_card_class}">
        <div class="kpi-label">DoR Quality Level</div>
        <div class="kpi-value {kpi3_value_class}">{kpi3_value}</div>
        <div class="kpi-detail">{kpi3_detail}</div>
      </div>
    </div>

    <div class="section">
      <h2>Team Overview</h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th class="center">DoR</th>
            <th class="center">Jira Tasks in progress</th>
            <th class="center">% Tasks fitting DoR</th>
          </tr>
        </thead>
        <tbody>
{team_overview_rows}
        </tbody>
      </table>
    </div>

    <div class="section">
      <h2>DoR Compliance - Failed Issues<span class="count">({failed_count} issues)</span></h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th>Issue Key</th>
            <th>Issue Type</th>
            <th>Title</th>
            <th>Status</th>
            <th>Assignee</th>
            <th>Note</th>
          </tr>
        </thead>
        <tbody>
{failed_issues_rows}
        </tbody>
      </table>
    </div>

    <div class="section">
      <h2>DoR Quality Assessment</h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th class="center">Coverage</th>
            <th>Score</th>
            <th>Note</th>
          </tr>
        </thead>
        <tbody>
{quality_rows}
        </tbody>
      </table>
    </div>

    <footer>
      DoR Scanner v2.0 | Generated by wow-dor-scanner skill | Data source: SRPOL Teams Confluence
    </footer>
  </div>
</body>
</html>"""


def _color_class(value):
    """Return CSS class based on threshold: good >= 70, warn 40-69, bad < 40."""
    if value >= 70:
        return "good"
    elif value >= 40:
        return "warn"
    else:
        return "bad"


def _score_class(pct_str):
    """Return CSS class for a percentage string like '73%' or '-'."""
    if pct_str == "-":
        return "na"
    try:
        val = int(pct_str.replace("%", ""))
        if val >= 80:
            return "good"
        elif val >= 50:
            return "warn"
        else:
            return "bad"
    except (ValueError, AttributeError):
        return "na"


def _escape_html(text):
    """Escape HTML special characters."""
    if not text:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def generate_html(compliance_data, summary_data, quality_data, scan_date, output_path):
    """Generate Report-DoR.html using the fixed template."""
    team_count = len(summary_data)

    # KPI 1: Teams with DoR
    teams_with_dor = sum(1 for t in summary_data if t["dor"] == "Yes")
    pct_teams = round(teams_with_dor / team_count * 100, 1) if team_count > 0 else 0
    kpi1_card_class = _color_class(pct_teams)
    kpi1_value_class = _color_class(pct_teams)
    kpi1_value = f"{pct_teams}%"
    kpi1_detail = f"{teams_with_dor} of {team_count} teams"

    # KPI 2: Tasks Fitting DoR
    total_issues = len(compliance_data)
    compliant_issues = sum(1 for d in compliance_data if d["dor_compliance"] == "Pass")
    if total_issues > 0:
        pct_compliance = round(compliant_issues / total_issues * 100, 1)
        kpi2_card_class = _color_class(pct_compliance)
        kpi2_value_class = _color_class(pct_compliance)
        kpi2_value = f"{pct_compliance}%"
        kpi2_detail = f"{compliant_issues} of {total_issues} tasks compliant"
    else:
        kpi2_card_class = ""
        kpi2_value_class = ""
        kpi2_value = "-"
        kpi2_detail = "No active issues analyzed"

    # KPI 3: DoR Quality Level
    if quality_data:
        avg_quality = round(sum(r["coverage"] for r in quality_data) / len(quality_data))
        kpi3_card_class = _color_class(avg_quality)
        kpi3_value_class = _color_class(avg_quality)
        kpi3_value = f"{avg_quality}/100"
        kpi3_detail = f"Across {len(quality_data)} assessed teams"
    else:
        kpi3_card_class = ""
        kpi3_value_class = ""
        kpi3_value = "-"
        kpi3_detail = "No teams assessed"

    # Team Overview rows
    overview_lines = []
    for team in summary_data:
        dor_badge = "badge-yes" if team["dor"] == "Yes" else "badge-no"
        dor_text = team["dor"]
        pct_str = team["pct_fitting_dor"]
        score_cls = _score_class(pct_str)
        overview_lines.append(
            f'        <tr>\n'
            f'          <td class="team-name">{_escape_html(team["team"])}</td>\n'
            f'          <td class="center"><span class="badge {dor_badge}">{dor_text}</span></td>\n'
            f'          <td class="center">{team["jira_tasks"]}</td>\n'
            f'          <td class="center"><span class="score {score_cls}">{pct_str}</span></td>\n'
            f'        </tr>'
        )
    team_overview_rows = "\n".join(overview_lines)

    # Failed Issues rows (only Fail items)
    failed_items = [d for d in compliance_data if d["dor_compliance"] == "Fail"]
    failed_count = len(failed_items)
    failed_lines = []
    for item in failed_items:
        failed_lines.append(
            f'        <tr>\n'
            f'          <td class="team-name">{_escape_html(item["team"])}</td>\n'
            f'          <td><a href="{_escape_html(item["url"])}" target="_blank">{_escape_html(item["issue_key"])}</a></td>\n'
            f'          <td>{_escape_html(item["issue_type"])}</td>\n'
            f'          <td class="title-col">{_escape_html(item["title"])}</td>\n'
            f'          <td>{_escape_html(item["status"])}</td>\n'
            f'          <td>{_escape_html(item["assignee"])}</td>\n'
            f'          <td class="note-col">{_escape_html(item["note"])}</td>\n'
            f'        </tr>'
        )
    failed_issues_rows = "\n".join(failed_lines)

    # Quality Assessment rows (sorted by coverage desc)
    sorted_quality = sorted(quality_data, key=lambda x: x["coverage"], reverse=True)
    quality_lines = []
    for q in sorted_quality:
        cov = q["coverage"]
        bar_class = _color_class(cov)
        score_cls = _color_class(cov)
        quality_lines.append(
            f'        <tr>\n'
            f'          <td class="team-name">{_escape_html(q["team"])}</td>\n'
            f'          <td class="center"><strong class="score {score_cls}">{cov}</strong></td>\n'
            f'          <td>\n'
            f'            <div class="bar-container">\n'
            f'              <div class="bar {bar_class}" style="width: {cov}%"></div>\n'
            f'            </div>\n'
            f'          </td>\n'
            f'          <td class="note-col">{_escape_html(q["note"])}</td>\n'
            f'        </tr>'
        )
    quality_rows = "\n".join(quality_lines)

    # Render template
    html = HTML_TEMPLATE.format(
        scan_date=scan_date,
        team_count=team_count,
        kpi1_card_class=kpi1_card_class,
        kpi1_value_class=kpi1_value_class,
        kpi1_value=kpi1_value,
        kpi1_detail=kpi1_detail,
        kpi2_card_class=kpi2_card_class,
        kpi2_value_class=kpi2_value_class,
        kpi2_value=kpi2_value,
        kpi2_detail=kpi2_detail,
        kpi3_card_class=kpi3_card_class,
        kpi3_value_class=kpi3_value_class,
        kpi3_value=kpi3_value,
        kpi3_detail=kpi3_detail,
        team_overview_rows=team_overview_rows,
        failed_count=failed_count,
        failed_issues_rows=failed_issues_rows,
        quality_rows=quality_rows,
    )

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return True


# ============================================================
# MAIN ENTRY POINT
# ============================================================

def main():
    if len(sys.argv) < 6:
        print("Usage: python3 generate_reports.py <summary.json> <compliance.json> <quality.json> <output_dir> <scan_date>")
        print("Example: python3 generate_reports.py summary_data.json compliance_data.json quality_data.json . 2026-06-26")
        sys.exit(1)

    summary_path = sys.argv[1]
    compliance_path = sys.argv[2]
    quality_path = sys.argv[3]
    output_dir = sys.argv[4]
    scan_date = sys.argv[5]

    # Load data
    with open(summary_path, 'r', encoding='utf-8') as f:
        summary_data = json.load(f)
    with open(compliance_path, 'r', encoding='utf-8') as f:
        compliance_data = json.load(f)
    with open(quality_path, 'r', encoding='utf-8') as f:
        quality_data = json.load(f)

    print(f"[INFO] Loaded data: {len(summary_data)} teams, {len(compliance_data)} issues, {len(quality_data)} quality scores")

    # Generate Excel
    xlsx_path = os.path.join(output_dir, "Report-DoR.xlsx")
    if generate_excel(compliance_data, summary_data, quality_data, xlsx_path):
        print(f"[SUCCESS] Excel report: {xlsx_path}")
        print(f"  Sheets: [{SUMMARY_SHEET_NAME}, {COMPLIANCE_SHEET_NAME}, {QUALITY_SHEET_NAME}]")
    else:
        print(f"[WARNING] Excel report not generated (openpyxl missing)")

    # Generate HTML
    html_path = os.path.join(output_dir, "Report-DoR.html")
    if generate_html(compliance_data, summary_data, quality_data, scan_date, html_path):
        print(f"[SUCCESS] HTML report: {html_path}")
    else:
        print(f"[ERROR] HTML report generation failed")

    # Summary
    total_issues = len(compliance_data)
    compliant = sum(1 for d in compliance_data if d["dor_compliance"] == "Pass")
    failed = total_issues - compliant
    avg_quality = round(sum(r["coverage"] for r in quality_data) / len(quality_data)) if quality_data else 0

    print(f"\n=== Report Generation Complete ===")
    print(f"  Teams: {len(summary_data)}")
    print(f"  Issues analyzed: {total_issues} (Pass: {compliant}, Fail: {failed})")
    print(f"  Avg DoR Quality: {avg_quality}/100")


if __name__ == "__main__":
    main()
