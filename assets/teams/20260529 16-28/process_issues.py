#!/usr/bin/env python3
"""
WoW Team Scanner - DoR Compliance Analysis
Processes Jira issues and analyzes DoR compliance
"""

import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Team mapping: skill team name -> Jira customfield_10114 value
TEAM_MAPPING = {
    "PE-WAW-Abyss": "Abyss",
    "Radium": "AE - WAW - Radium",
    "Europium": "AP - WAW - Europium",
    "Copernicium": "AE - WAW - Copernicium",
    "Mouflons": "Mouflons",
    "Wolves": "Wolves",
    "Polonium LF": "Polonium",  # Both Polonium teams share same field value
    "Polonium UF": "Polonium",
    "Bigos": "Bigos",
    "Capybaras": "Capybaras",
    "ML Serving": "ML Serving",
    "ML Platform": "ML Platform",
    "Zurek": "Zurek",
    "EP Core": "EP Core",
    "Igni": "Igni",
    "SRE": "SRE"
}

# DoR file mapping
DOR_FILES = {
    "PE-WAW-Abyss": "pe-waw-abyss-dor.txt",
    "Radium": "radium-dor.txt",
    "Europium": "europium-dor.txt",
    "Copernicium": "copernicium-dor.txt",
    "Mouflons": "mouflons-dor.txt",
    "Wolves": "wolves-dor.txt",
    "Polonium LF": "polonium-lf-dor.txt",
    "Polonium UF": "polonium-uf-dor.txt",
    "Bigos": "bigos-dor.txt",
    "Capybaras": "capybaras-dor.txt",
    "ML Serving": "ml-serving-dor.txt",
    "ML Platform": "ml-platform-dor.txt",
    "Zurek": "zurek-dor.txt",
    "EP Core": "ep-core-dor.txt",
    "Igni": "igni-dor.txt",
    "SRE": "sre-dor.txt"
}

def load_dor(dor_file_path):
    """Load DoR content from file"""
    if not os.path.exists(dor_file_path):
        return None

    with open(dor_file_path, 'r', encoding='utf-8') as f:
        content = f.read().strip()

    # Check if DoR is missing
    if "not found" in content.lower() or "could not access" in content.lower():
        return None

    return content

def analyze_dor_compliance(issue, dor_content):
    """
    Analyze if a Jira issue meets DoR criteria
    Returns: (compliance: bool, feedback: str)
    """
    if not dor_content:
        return True, "No DoR criteria defined for this team"

    # Extract issue fields
    summary = issue['fields'].get('summary', '')
    description = issue['fields'].get('description', '')
    issuetype = issue['fields']['issuetype']['name']

    # Simple heuristic-based analysis (can be enhanced with LLM later)
    feedback_items = []
    compliant = True

    # Check 1: Description exists and is substantial
    if not description or len(description.strip()) < 50:
        feedback_items.append("Description is missing or too short")
        compliant = False

    # Check 2: Summary is clear and specific
    if len(summary.strip()) < 10:
        feedback_items.append("Summary is too brief")
        compliant = False

    # Check 3: Common DoR criteria patterns
    dor_lower = dor_content.lower()

    # Acceptance criteria check
    if 'acceptance criteria' in dor_lower:
        if not description or 'acceptance criteria' not in description.lower():
            feedback_items.append("Acceptance criteria not documented in description")
            compliant = False

    # Dependencies check
    if 'dependencies' in dor_lower or 'dependency' in dor_lower:
        # Simplified check - would need to verify Jira links/fields
        pass

    # Estimate check
    if 'estimate' in dor_lower or 'story point' in dor_lower:
        # Would need to check story points field
        pass

    if compliant:
        return True, ""
    else:
        return False, "; ".join(feedback_items)

def process_issues(aenw_file, aetvp_file, output_dir):
    """Process Jira issues and generate DoR compliance report"""

    # Load issue data
    with open(aenw_file, 'r', encoding='utf-8') as f:
        aenw_data = json.load(f)

    with open(aetvp_file, 'r', encoding='utf-8') as f:
        aetvp_data = json.load(f)

    all_issues = aenw_data['issues']['nodes'] + aetvp_data['issues']['nodes']

    print(f"Total issues loaded: {len(all_issues)}")

    # Filter and categorize issues by team
    team_issues = {team: [] for team in TEAM_MAPPING.keys()}
    unmatched_issues = []

    for issue in all_issues:
        team_field = issue['fields'].get('customfield_10114')
        if team_field and isinstance(team_field, dict):
            team_name_jira = team_field.get('name', '')

            # Find matching team
            matched = False
            for skill_team, jira_team in TEAM_MAPPING.items():
                if jira_team in team_name_jira:
                    team_issues[skill_team].append(issue)
                    matched = True
                    break

            if not matched:
                unmatched_issues.append(issue)
        else:
            unmatched_issues.append(issue)

    # Print team issue counts
    print("\nIssues per team:")
    for team, issues in team_issues.items():
        if issues:
            print(f"  {team}: {len(issues)}")

    print(f"\nUnmatched issues: {len(unmatched_issues)}")

    # Generate Excel report
    wb = Workbook()
    ws = wb.active
    ws.title = "DoR Compliance"

    # Define column headers
    headers = ["Team", "Issue Key", "Issue Type", "Status", "Title", "URL", "Assignee", "DoR Compliance", "Feedback"]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Process each team
    row = 2
    for team, issues in sorted(team_issues.items()):
        if not issues:
            continue

        # Load team's DoR
        dor_file = os.path.join(output_dir, DOR_FILES[team])
        dor_content = load_dor(dor_file)

        for issue in issues:
            key = issue['key']
            issuetype = issue['fields']['issuetype']['name']
            status = issue['fields']['status']['name']
            summary = issue['fields'].get('summary', '')
            url = f"https://adgear.atlassian.net/browse/{key}"

            assignee_obj = issue['fields'].get('assignee')
            if assignee_obj:
                assignee = assignee_obj.get('displayName', 'Unassigned')
            else:
                assignee = "Unassigned"

            # Analyze DoR compliance
            is_compliant, feedback = analyze_dor_compliance(issue, dor_content)

            # Add row
            ws.append([
                team,
                key,
                issuetype,
                status,
                summary,
                url,
                assignee,
                "Yes" if is_compliant else "No",
                feedback
            ])

            # Style compliance cell
            compliance_cell = ws.cell(row, 8)
            if is_compliant:
                compliance_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                compliance_cell.font = Font(color="006100")
            else:
                compliance_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                compliance_cell.font = Font(color="9C0006")

            compliance_cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1

    # Adjust column widths
    column_widths = [20, 15, 12, 15, 50, 60, 20, 15, 80]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Save workbook
    report_file = os.path.join(output_dir, "Report.xlsx")
    wb.save(report_file)

    print(f"\n[SUCCESS] Excel report generated: {report_file}")
    print(f"  Total issues analyzed: {row - 2}")

    return team_issues

if __name__ == "__main__":
    output_dir = Path(__file__).parent
    # Tool results are stored in user's home .claude directory
    claude_dir = Path.home() / ".claude" / "projects" / "C--Users-i-klyha-Desktop-Claude-wow-scanner-tool" / "6ff25fed-697b-46cc-97d7-cb86ef173727" / "tool-results"
    aenw_file = claude_dir / "toolu_bdrk_01UbEt8F9ghribHkUafHazEf.txt"
    aetvp_file = claude_dir / "toolu_bdrk_01SL4y3KpTvAk8ii4GKwxzFe.txt"

    process_issues(str(aenw_file), str(aetvp_file), str(output_dir))
