#!/usr/bin/env python3
"""Step 10: Add 'DoD quality' sheet to Report-DoD.xlsx"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"C:\Users\i.klyha\Desktop\Claude\wow-scanner-tool\assets\teams\20260615 15-43"

results = [
    {"team": "Europium", "overall": 77, "note": "Missing explicit performance/security validation and PO acceptance"},
    {"team": "Mouflons", "overall": 66, "note": "Missing deployment, monitoring, and performance criteria"},
    {"team": "Wolves", "overall": 66, "note": "Missing deployment, monitoring, and performance criteria"},
    {"team": "Igni", "overall": 65, "note": "Missing code review, monitoring, PO acceptance; documentation TBD"},
    {"team": "Abyss", "overall": 64, "note": "Missing explicit code review and PO acceptance; criteria bundled"},
    {"team": "Polonium UF", "overall": 64, "note": "Missing CI/pipeline, monitoring, defect tracking, PO acceptance"},
    {"team": "ML Serving Sturgeons", "overall": 62, "note": "Missing deployment, monitoring, defect tracking criteria"},
    {"team": "ML Platform Pandas", "overall": 62, "note": "Missing deployment, monitoring, defect tracking criteria"},
    {"team": "Copernicium", "overall": 60, "note": "AC criterion bundles too many items; missing CI, defects, PO acceptance"},
    {"team": "Zurek", "overall": 58, "note": "Missing code review, monitoring, PO acceptance; vague test requirements"},
    {"team": "Bigos", "overall": 56, "note": "Missing code review, tests, CI, documentation, monitoring criteria"},
    {"team": "Capybaras", "overall": 55, "note": "Tests criterion TBD; missing code review, CI, monitoring, PO acceptance"},
    {"team": "Radium", "overall": 54, "note": "Vague criteria ('??'), missing documentation, monitoring, PO acceptance"},
    {"team": "EP Core", "overall": 46, "note": "Very incomplete - missing code review, tests, CI, deployment, monitoring"},
]

average_quality = round(sum(r["overall"] for r in results) / len(results))

wb = load_workbook(f"{OUTPUT_DIR}/Report-DoD.xlsx")
print(f"Before: {wb.sheetnames}")

ws = wb.create_sheet("DoD quality")

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
orange_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Row 1: KPI
ws.cell(1, 1).value = "DoD quality lvl"
ws.cell(1, 1).font = Font(bold=True, size=11)
ws.cell(1, 2).value = f"{average_quality}%"
if average_quality >= 70:
    ws.cell(1, 2).font = Font(bold=True, size=11, color="006600")
elif average_quality >= 40:
    ws.cell(1, 2).font = Font(bold=True, size=11, color="CC6600")
else:
    ws.cell(1, 2).font = Font(bold=True, size=11, color="CC0000")

# Row 3: Headers
for col, (header, width) in enumerate([("Team", 25), ("Quality", 12), ("Note", 60)], 1):
    cell = ws.cell(3, col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width

# Data rows
for row_idx, result in enumerate(results, 4):
    ws.cell(row_idx, 1).value = result["team"]
    ws.cell(row_idx, 1).border = thin_border
    ws.cell(row_idx, 1).alignment = Alignment(vertical='center')

    quality_cell = ws.cell(row_idx, 2)
    quality_cell.value = f"{result['overall']}%"
    quality_cell.border = thin_border
    quality_cell.alignment = Alignment(horizontal='center', vertical='center')
    if result['overall'] >= 70:
        quality_cell.fill = green_fill
    elif result['overall'] >= 40:
        quality_cell.fill = orange_fill
    else:
        quality_cell.fill = red_fill

    note_cell = ws.cell(row_idx, 3)
    note_cell.value = result["note"]
    note_cell.border = thin_border
    note_cell.alignment = Alignment(wrap_text=True, vertical='center')

ws.freeze_panes = 'A4'
wb.save(f"{OUTPUT_DIR}/Report-DoD.xlsx")

print(f"After: {wb.sheetnames}")
print(f"Average DoD Quality: {average_quality}%")
print(f"[SUCCESS] Report-DoD.xlsx now has 2 sheets")
