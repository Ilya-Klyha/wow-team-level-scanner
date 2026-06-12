#!/usr/bin/env python3
"""
DoR Compliance Analysis Script
Analyzes Jira issues against team Definition of Ready criteria
and generates an Excel report with compliance results.
"""

import json
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define analyzable teams (teams with both DoR and active issues)
ANALYZABLE_TEAMS = [
    {
        "name": "PE-WAW-Abyss",
        "dor_file": "pe-waw-abyss-dor.txt",
        "jira_file": "pe-waw-abyss-jira.json"
    },
    {
        "name": "Radium",
        "dor_file": "radium-dor.txt",
        "jira_file": "radium-jira.json"
    },
    {
        "name": "Europium",
        "dor_file": "europium-dor.txt",
        "jira_file": "europium-jira.json"
    },
    {
        "name": "Copernicium",
        "dor_file": "copernicium-dor.txt",
        "jira_file": "copernicium-jira.json"
    },
    {
        "name": "Polonium-Upper-Funnel",
        "dor_file": "polonium-upper-funnel-dor.txt",
        "jira_file": "polonium-upper-funnel-jira.json"
    },
    {
        "name": "Capybaras",
        "dor_file": "capybaras-dor.txt",
        "jira_file": "capybaras-jira.json"
    },
    {
        "name": "EP-Core",
        "dor_file": "ep-core-dor.txt",
        "jira_file": "ep-core-jira.json"
    },
    {
        "name": "Igni",
        "dor_file": "igni-dor.txt",
        "jira_file": "igni-jira.json"
    }
]

# Manual DoR compliance analysis results
# Based on thorough review of each issue against team's DoR criteria
DOR_ANALYSIS_RESULTS = {
    "PE-WAW-Abyss": [
        {
            "issue_key": "MAW-418",
            "meets_dor": False,
            "missing_criteria": ["No description provided", "Acceptance criteria missing", "Dependencies not identified"]
        },
        {
            "issue_key": "MAW-377",
            "meets_dor": True,
            "missing_criteria": []
        },
        {
            "issue_key": "MAW-376",
            "meets_dor": True,
            "missing_criteria": []
        },
        {
            "issue_key": "MAW-375",
            "meets_dor": True,
            "missing_criteria": []
        },
        {
            "issue_key": "MAW-374",
            "meets_dor": True,
            "missing_criteria": []
        },
        {
            "issue_key": "MAW-259",
            "meets_dor": True,
            "missing_criteria": []
        }
    ],
    "Radium": [
        {
            "issue_key": "AENW-939",
            "meets_dor": False,
            "missing_criteria": ["No user story format", "Acceptance criteria missing", "Figma mockup not provided"]
        },
        {
            "issue_key": "AENW-915",
            "meets_dor": False,
            "missing_criteria": ["Not well described", "Acceptance criteria vague"]
        },
        {
            "issue_key": "AENW-912",
            "meets_dor": False,
            "missing_criteria": ["Not well described", "Limited context provided"]
        },
        {
            "issue_key": "AENW-909",
            "meets_dor": False,
            "missing_criteria": ["Incomplete description", "Acceptance criteria not fully testable"]
        },
        {
            "issue_key": "AENW-906",
            "meets_dor": False,
            "missing_criteria": ["No description provided", "No acceptance criteria", "Not a development task"]
        },
        {
            "issue_key": "AENW-904",
            "meets_dor": False,
            "missing_criteria": ["Not well described", "No user story context", "Limited business justification"]
        },
        {
            "issue_key": "AENW-867",
            "meets_dor": True,
            "missing_criteria": []
        },
        {
            "issue_key": "AENW-843",
            "meets_dor": True,
            "missing_criteria": []
        },
        {
            "issue_key": "AENW-767",
            "meets_dor": False,
            "missing_criteria": ["No user story format", "Requirements not well described"]
        },
        {
            "issue_key": "AENW-579",
            "meets_dor": False,
            "missing_criteria": ["No description provided", "No acceptance criteria"]
        },
        {
            "issue_key": "AENW-574",
            "meets_dor": False,
            "missing_criteria": ["No acceptance criteria defined", "Missing Figma mockup link"]
        }
    ],
    "Europium": [
        {
            "issue_key": "AENW-895",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined", "Figma required but not linked"]
        },
        {
            "issue_key": "AENW-859",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined", "Tech spec approval status unclear"]
        },
        {
            "issue_key": "AENW-701",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined", "Open questions not resolved"]
        },
        {
            "issue_key": "AENW-685",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined", "Design not finalized"]
        }
    ],
    "Copernicium": [
        {
            "issue_key": "AETVP-584",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "AETVP-579",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "AETVP-564",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not clear"]
        },
        {
            "issue_key": "AETVP-528",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "AETVP-476",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        }
    ],
    "Polonium-Upper-Funnel": [
        {
            "issue_key": "RSW-1237",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined", "Testcases not defined"]
        },
        {
            "issue_key": "RSW-1214",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined", "Testcases not defined"]
        },
        {
            "issue_key": "RSW-1038",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined", "Mockups not provided"]
        },
        {
            "issue_key": "RSW-880",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined", "Dependencies not clear"]
        }
    ],
    "Capybaras": [
        {
            "issue_key": "RSW-1314",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not measurable", "Team tag missing"]
        },
        {
            "issue_key": "RSW-1306",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria format incorrect", "Labels incomplete"]
        },
        {
            "issue_key": "RSW-1305",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not in artifact form"]
        },
        {
            "issue_key": "RSW-1097",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined", "User story format incomplete"]
        },
        {
            "issue_key": "RSW-555",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria unclear", "Too long in progress"]
        }
    ],
    "EP-Core": [
        {
            "issue_key": "EPCW-232",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "EPCW-195",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not testable"]
        },
        {
            "issue_key": "EPCW-171",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "EPCW-163",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not clear"]
        },
        {
            "issue_key": "EPCW-158",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "EPCW-157",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "EPCW-153",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not clear"]
        },
        {
            "issue_key": "EPCW-148",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "EPCW-145",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "EPCW-138",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not testable"]
        },
        {
            "issue_key": "EPCW-131",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "EPCW-96",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not clear"]
        },
        {
            "issue_key": "EPCW-95",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "EPCW-22",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "EPCW-9",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "EPCW-6",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        }
    ],
    "Igni": [
        {
            "issue_key": "ASPW-997",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "ASPW-689",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not testable"]
        },
        {
            "issue_key": "ASPW-637",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Reproduction steps not clear"]
        },
        {
            "issue_key": "ASPW-560",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined", "Not a development task"]
        },
        {
            "issue_key": "ASPW-554",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined"]
        },
        {
            "issue_key": "ASPW-6",
            "meets_dor": False,
            "missing_criteria": ["Description missing", "Acceptance criteria not defined", "Monitoring requirements unclear"]
        }
    ]
}


def load_jira_issues(base_path):
    """Load all Jira issues from JSON files."""
    all_issues = {}

    for team in ANALYZABLE_TEAMS:
        jira_path = base_path / team["jira_file"]
        with open(jira_path, 'r', encoding='utf-8') as f:
            jira_data = json.load(f)
            all_issues[team["name"]] = jira_data.get("issues", [])

    return all_issues


def generate_excel_report(report_data, output_path):
    """Generate Excel report with DoR compliance analysis."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DoR Compliance"

    # Define columns
    columns = [
        ("Team", 18),
        ("Issue Key", 12),
        ("Issue Type", 10),
        ("Status", 12),
        ("Title", 50),
        ("URL", 55),
        ("Assignee", 20),
        ("DoR Compliance", 15),
        ("Feedback", 70)
    ]

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write header row
    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 25

    # Add data rows
    for row_idx, row_data in enumerate(report_data, start=2):
        ws.cell(row=row_idx, column=1).value = row_data["team"]
        ws.cell(row=row_idx, column=2).value = row_data["issue_key"]
        ws.cell(row=row_idx, column=3).value = row_data["issue_type"]
        ws.cell(row=row_idx, column=4).value = row_data["status"]
        ws.cell(row=row_idx, column=5).value = row_data["title"]
        ws.cell(row=row_idx, column=6).value = row_data["url"]
        ws.cell(row=row_idx, column=6).hyperlink = row_data["url"]
        ws.cell(row=row_idx, column=6).style = "Hyperlink"
        ws.cell(row=row_idx, column=7).value = row_data["assignee"]

        compliance_value = row_data["dor_compliance"]
        ws.cell(row=row_idx, column=8).value = compliance_value
        ws.cell(row=row_idx, column=8).fill = yes_fill if compliance_value == "Yes" else no_fill
        ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=row_idx, column=9).value = row_data.get("feedback", "")

        # Apply borders and alignment
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if col in [5, 9]:  # Title and Feedback columns
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(vertical='center')

    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")


def generate_report_data(all_issues):
    """Generate report data by combining issues with DoR analysis."""
    report_data = []

    for team_name, issues in all_issues.items():
        team_analysis = DOR_ANALYSIS_RESULTS.get(team_name, [])

        for issue in issues:
            analysis = next((a for a in team_analysis if a["issue_key"] == issue["key"]), None)

            if analysis:
                report_data.append({
                    "team": team_name,
                    "issue_key": issue["key"],
                    "issue_type": issue["type"],
                    "status": issue["status"],
                    "title": issue["summary"],
                    "url": issue["url"],
                    "assignee": issue.get("assignee", "Unassigned"),
                    "dor_compliance": "Yes" if analysis["meets_dor"] else "No",
                    "feedback": "; ".join(analysis["missing_criteria"]) if analysis["missing_criteria"] else ""
                })

    return report_data


def calculate_statistics(report_data):
    """Calculate statistics for the analysis."""
    total_issues = len(report_data)
    meets_dor = sum(1 for row in report_data if row["dor_compliance"] == "Yes")
    does_not_meet = total_issues - meets_dor
    compliance_rate = (meets_dor / total_issues * 100) if total_issues > 0 else 0

    # Team breakdown
    team_stats = {}
    for row in report_data:
        team = row["team"]
        if team not in team_stats:
            team_stats[team] = {"total": 0, "meets": 0}
        team_stats[team]["total"] += 1
        if row["dor_compliance"] == "Yes":
            team_stats[team]["meets"] += 1

    # Common missing criteria
    missing_criteria_count = {}
    for row in report_data:
        if row["feedback"]:
            criteria_list = [c.strip() for c in row["feedback"].split(";")]
            for criteria in criteria_list:
                missing_criteria_count[criteria] = missing_criteria_count.get(criteria, 0) + 1

    top_missing = sorted(missing_criteria_count.items(), key=lambda x: x[1], reverse=True)[:5]

    return {
        "total_issues": total_issues,
        "meets_dor": meets_dor,
        "does_not_meet": does_not_meet,
        "compliance_rate": compliance_rate,
        "team_stats": team_stats,
        "top_missing": top_missing
    }


def generate_summary_document(stats, output_path):
    """Generate summary markdown document."""
    now = datetime.now().isoformat()

    summary_lines = [
        "# DoR Compliance Analysis Summary",
        "",
        f"**Generated:** {now}",
        f"**Scan Directory:** {output_path.parent}",
        "",
        "---",
        "",
        "## Overall Statistics",
        "",
        f"- **Total Teams Analyzed:** {len(stats['team_stats'])}",
        f"- **Total Issues Analyzed:** {stats['total_issues']}",
        f"- **Issues Meeting DoR:** {stats['meets_dor']} ({stats['compliance_rate']:.1f}%)",
        f"- **Issues NOT Meeting DoR:** {stats['does_not_meet']} ({100 - stats['compliance_rate']:.1f}%)",
        "- **Teams Skipped (No DoR):** 4",
        "- **Teams Skipped (No Active Issues):** 4",
        "",
        "---",
        "",
        "## Breakdown by Team",
        "",
        "| Team | Total Issues | Meets DoR | Does Not Meet | Compliance Rate |",
        "|------|--------------|-----------|---------------|-----------------|"
    ]

    for team_name, team_data in sorted(stats['team_stats'].items()):
        total = team_data['total']
        meets = team_data['meets']
        does_not = total - meets
        rate = (meets / total * 100) if total > 0 else 0
        summary_lines.append(f"| {team_name} | {total} | {meets} | {does_not} | {rate:.1f}% |")

    summary_lines.extend([
        "",
        "---",
        "",
        "## Most Common DoR Gaps",
        ""
    ])

    for idx, (criteria, count) in enumerate(stats['top_missing'], 1):
        summary_lines.append(f"{idx}. **{criteria}** - {count} occurrences")

    summary_lines.extend([
        "",
        "---",
        "",
        "## Teams with No DoR Documentation",
        "",
        "- Polonium-LF",
        "- ML-Serving-Sturgeons",
        "- ML-Platform-Pandas",
        "- SRPOL-SRE",
        "",
        "---",
        "",
        "## Teams with DoR but No Active Issues",
        "",
        "- Mouflons",
        "- Wolves",
        "- Bigos",
        "- Ads-Reporting-Bigos-Zurek",
        "",
        "---",
        "",
        "## Files Generated",
        "",
        "- **Report.xlsx** - Full compliance report",
        "- **DOR_ANALYSIS_SUMMARY.md** - This summary",
        "- **teams.json** - Updated with analysis metadata",
        "",
        "---",
        "",
        "## Analysis Method",
        "",
        "- **Tool:** Claude Sonnet 4.5 (Manual semantic analysis)",
        "- **Approach:** Issue-by-issue review against team DoR criteria",
        "- **Criteria:** Strict compliance checking",
        "",
        "---",
        "",
        "## Recommendations",
        "",
        "1. Review all issues marked \"No\" in the report",
        "2. Address common DoR gaps identified above",
        "3. Teams without DoR should document their criteria",
        "4. Consider refining DoR criteria based on common gaps",
        "5. Establish DoR review process before sprint planning",
        "",
        "---",
        "",
        f"**Report Location:** {output_path.parent / 'Report.xlsx'}",
        ""
    ])

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(summary_lines))

    print(f"Summary document saved to: {output_path}")


def update_teams_json(stats, base_path):
    """Update teams.json with analysis metadata."""
    teams_json_path = base_path / "teams.json"

    with open(teams_json_path, 'r', encoding='utf-8') as f:
        teams_data = json.load(f)

    # Add metadata
    teams_data["metadata"]["dor_analysis"] = {
        "performed": True,
        "timestamp": datetime.now().isoformat(),
        "teams_analyzed": len(stats['team_stats']),
        "issues_analyzed": stats['total_issues'],
        "issues_meeting_dor": stats['meets_dor'],
        "issues_not_meeting_dor": stats['does_not_meet'],
        "compliance_rate": f"{stats['compliance_rate']:.1f}",
        "report_file": "Report.xlsx",
        "summary_file": "DOR_ANALYSIS_SUMMARY.md",
        "analysis_method": "manual_semantic"
    }

    # Update each team
    for team in teams_data["teams"]:
        team_name = team["name"]
        if team_name in stats['team_stats']:
            team_stats = stats['team_stats'][team_name]
            team["dor_analysis"] = {
                "analyzed": True,
                "issues_count": team_stats['total'],
                "meets_dor": team_stats['meets'],
                "does_not_meet": team_stats['total'] - team_stats['meets'],
                "compliance_rate": f"{(team_stats['meets'] / team_stats['total'] * 100):.1f}"
            }
        elif team.get("extraction_status") == "not_found":
            team["dor_analysis"] = {
                "analyzed": False,
                "reason": "no_dor_criteria"
            }
        elif team.get("jiraIssueCount", 0) == 0:
            team["dor_analysis"] = {
                "analyzed": False,
                "reason": "no_active_issues"
            }

    with open(teams_json_path, 'w', encoding='utf-8') as f:
        json.dump(teams_data, f, indent=2, ensure_ascii=False)

    print(f"teams.json updated with analysis metadata")


def main():
    base_path = Path(__file__).parent

    print("=== Starting DoR Compliance Analysis ===")
    print(f"Teams to analyze: {len(ANALYZABLE_TEAMS)}")

    # Load Jira issues
    print("\nLoading Jira issues...")
    all_issues = load_jira_issues(base_path)
    total_issues = sum(len(issues) for issues in all_issues.values())
    print(f"Total issues loaded: {total_issues}")

    # Generate report data
    print("\nGenerating report data...")
    report_data = generate_report_data(all_issues)

    # Calculate statistics
    print("Calculating statistics...")
    stats = calculate_statistics(report_data)

    # Generate Excel report
    print("\nGenerating Excel report...")
    excel_path = base_path / "Report.xlsx"
    generate_excel_report(report_data, excel_path)

    # Generate summary document
    print("Generating summary document...")
    summary_path = base_path / "DOR_ANALYSIS_SUMMARY.md"
    generate_summary_document(stats, summary_path)

    # Update teams.json
    print("Updating teams.json...")
    update_teams_json(stats, base_path)

    # Print final summary
    print("\n=== DoR Compliance Analysis Complete ===")
    print(f"Teams Analyzed: {len(stats['team_stats'])}")
    print(f"Issues Analyzed: {stats['total_issues']}")
    print(f"Meeting DoR: {stats['meets_dor']} ({stats['compliance_rate']:.1f}%)")
    print(f"Not Meeting DoR: {stats['does_not_meet']} ({100 - stats['compliance_rate']:.1f}%)")
    print(f"\nReport: {excel_path}")
    print(f"Summary: {summary_path}")


if __name__ == "__main__":
    main()
