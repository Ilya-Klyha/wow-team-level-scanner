"""
DoR Compliance Report Generator
Generates Report.xlsx with Summary and DoR Compliance sheets.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"C:\Users\i.klyha\Desktop\Claude\wow-scanner-tool\assets\teams\20260612 10-03"

# --- DoR Analysis Results ---
# Based on fetching and analyzing all 76 issue descriptions against each team's DoR criteria

dor_compliance_data = [
    # ABYSS (DoR: clarity, AC, monitoring if applicable, estimates, dependencies)
    ("Abyss", "MAW-376", "Story", "In Progress", "Enabling dynamic parameters",
     "https://adgear.atlassian.net/browse/MAW-376", "Rafal Stefanowicz", "Yes", ""),
    ("Abyss", "MAW-375", "Story", "In Progress", "Enabling standard pixel selection",
     "https://adgear.atlassian.net/browse/MAW-375", "Maciej Lewandowski", "Yes", ""),
    ("Abyss", "MAW-374", "Story", "In Progress", "Dynamic parameters support for pixels v1",
     "https://adgear.atlassian.net/browse/MAW-374", "Maksim Matchenia", "Yes", ""),

    # RADIUM (DoR: clarity/well described, AC testable with Figma for FE, estimates, dependencies fulfilled)
    ("Radium", "AENW-939", "Story", "Code Review",
     "[Remnant] On DSP form for 2023 immersive banner add left/right toggle and burn in logo depending on selection",
     "https://adgear.atlassian.net/browse/AENW-939", "Jakub Aniol", "No",
     "Missing explicit acceptance criteria; description states what to do but no testable AC defined"),
    ("Radium", "AENW-915", "Story", "In Progress", "Prepare serving services tasks for updated C2E Tech Spec",
     "https://adgear.atlassian.net/browse/AENW-915", "Marcin Al-Jawahiri", "Yes", ""),
    ("Radium", "AENW-912", "Story", "In Progress",
     "Verify status of rtb-delivery feature/aenw-543-docker-compose-for-e2e-testing branch for Alex",
     "https://adgear.atlassian.net/browse/AENW-912", "Marcin Al-Jawahiri", "Yes", ""),
    ("Radium", "AENW-906", "Task", "In Progress", "Deliver TVs to Cracow",
     "https://adgear.atlassian.net/browse/AENW-906", "Patryk Palyska", "No",
     "No description provided; missing requirement clarity and acceptance criteria"),
    ("Radium", "AENW-904", "Story", "In Progress", "[SW Test + Fast Track] Marcin Al-Jawahiri",
     "https://adgear.atlassian.net/browse/AENW-904", "Marcin Al-Jawahiri", "Yes", ""),
    ("Radium", "AENW-867", "Story", "Code Review",
     "[Remnant] Test Left Aligned CTA button for Remnant 2023 on TV and using Ad Response generator",
     "https://adgear.atlassian.net/browse/AENW-867", "Krzysztof Urbanski", "Yes", ""),
    ("Radium", "AENW-843", "Story", "In Progress",
     "[Remnant] Knowledge gathering with MTV - how 1st screen Unified and Form Builder works in this context",
     "https://adgear.atlassian.net/browse/AENW-843", "Jakub Aniol", "Yes", ""),
    ("Radium", "AENW-767", "Task", "In Progress",
     "[DSP-JS] Adding a new click action type \"click-to-email\" on the UI side",
     "https://adgear.atlassian.net/browse/AENW-767", "Jakub Aniol", "Yes", ""),
    ("Radium", "AENW-579", "Task", "In Progress",
     "E2E Test Microsite for Spotlight Beacon and Spotlight Row on real TV",
     "https://adgear.atlassian.net/browse/AENW-579", "Marcin Matyszkowicz", "No",
     "No description provided; missing requirement clarity and acceptance criteria"),
    ("Radium", "AENW-574", "Story", "Code Review",
     "[DSP] Add functionality to select Microsite on the form, save it and show it in the details view",
     "https://adgear.atlassian.net/browse/AENW-574", "Marcin Matyszkowicz", "No",
     "Missing explicit acceptance criteria; only user story description without testable AC"),
    ("Radium", "AENW-127", "Story", "In Progress", "[DSP-UI] Add new click action \"QR Code Microsite\"",
     "https://adgear.atlassian.net/browse/AENW-127", "Unassigned", "Yes", ""),

    # EUROPIUM (DoR: clear problem statement, AC defined and testable, UX in Figma if applicable,
    # dependencies identified, scope fits sprint, open questions resolved)
    ("Europium", "AENW-895", "Story", "Code Review", "[Microsite P2] General Settings Section UI Fixes",
     "https://adgear.atlassian.net/browse/AENW-895", "Grzegorz Prasal", "Yes", ""),
    ("Europium", "AENW-882", "Story", "In Progress", "[Microsite UCS] Verify bucket and Kafka topic alignment",
     "https://adgear.atlassian.net/browse/AENW-882", "Michal Wiacek", "Yes", ""),
    ("Europium", "AENW-876", "Story", "Code Review",
     "[Microsite UCS] Timeout detection for stuck transcoding jobs",
     "https://adgear.atlassian.net/browse/AENW-876", "Michal Wiacek", "Yes", ""),
    ("Europium", "AENW-859", "Story", "In Progress", "[Microsite P2] Implement rtb-delivery endpoint and handler",
     "https://adgear.atlassian.net/browse/AENW-859", "Raman Krot", "Yes", ""),
    ("Europium", "AENW-701", "Story", "In Progress", "[Microsite p2] UseCase + Task - HTML Generation",
     "https://adgear.atlassian.net/browse/AENW-701", "Karol Strojek", "Yes", ""),
    ("Europium", "AENW-685", "Story", "In Progress", "[Micoristes P2] Create Micro Frontend",
     "https://adgear.atlassian.net/browse/AENW-685", "Karol Strojek", "Yes", ""),

    # COPERNICIUM (DoR: clarity, AC testable, monitoring if applicable, estimates, dependencies/UI mocks)
    ("Copernicium", "AETVP-584", "Task", "In Progress",
     "[rtb-delivery, adgear-data, log-converter] P&S reporting: Adjust implementation to use top-level property instead of misc",
     "https://adgear.atlassian.net/browse/AETVP-584", "Kamil Madejek", "Yes", ""),
    ("Copernicium", "AETVP-579", "Task", "In Progress",
     "Model, database, and creative type registration for separate Poll & Survey",
     "https://adgear.atlassian.net/browse/AETVP-579", "Yevgeniy Wyszynski", "Yes", ""),
    ("Copernicium", "AETVP-564", "Task", "In Progress",
     "[Spike] Investigate \"infoblob\" feasibility in InteractiveAd SDK custom trackers",
     "https://adgear.atlassian.net/browse/AETVP-564", "Kamil Madejek", "Yes", ""),
    ("Copernicium", "AETVP-528", "Task", "Code Review", "Run test campaign on TV in the office",
     "https://adgear.atlassian.net/browse/AETVP-528", "Jacek Wroblewski", "Yes", ""),
    ("Copernicium", "AETVP-476", "Task", "In Progress", "[vertica-rollups] Add iadpoll metrics aggregation",
     "https://adgear.atlassian.net/browse/AETVP-476", "Kamil Madejek", "Yes", ""),

    # MOUFLONS (DoR: clear description with 4 questions, measurable/SMART AC, collaborative progress)
    ("Mouflons", "PEPI-11393", "Task", "In Development",
     "Analyze preliminary result of A/B tests of MC EE models",
     "https://adgear.atlassian.net/browse/PEPI-11393", "Marek Kochaczyk", "No",
     "Missing measurable acceptance criteria; description lists campaigns but no clear definition of done or expected outcomes"),

    # POLONIUM UPPER FUNNEL (DoR: well described with AC related to Epic, UX if applicable,
    # connected to Epic, dependencies, estimated, fits sprint, test cases, mockups for UI, etc.)
    ("Polonium Upper Funnel", "RSW-1237", "Task", "In Progress",
     "[FE] Integrate Retry Button with dedicated BE endpoint",
     "https://adgear.atlassian.net/browse/RSW-1237", "Zbigniew Stolarski", "Yes", ""),
    ("Polonium Upper Funnel", "RSW-1038", "Task", "In Progress",
     "[BE] Add new table to DB for custom groups' apps presets",
     "https://adgear.atlassian.net/browse/RSW-1038", "Dmitriy Borowski", "Yes", ""),
    ("Polonium Upper Funnel", "RSW-880", "Task", "Code Review", "[BE] GenAI Service - session reset",
     "https://adgear.atlassian.net/browse/RSW-880", "Rafal Babinski", "No",
     "Missing clear acceptance criteria; description is a technical specification of steps without testable AC"),

    # BIGOS (DoR: clarity concrete+measurable, clear & testable AC with FF + release env,
    # mockups if required, TechSpecs approved, estimated with SP, dependencies resolved, decomposed)
    ("Bigos", "MAW-406", "Story", "In Progress", "Add Simple Bearer Authentication",
     "https://adgear.atlassian.net/browse/MAW-406", "Kirill Timoszenko", "Yes", ""),
    ("Bigos", "MAW-372", "Task", "In Progress", "Setup git Agent",
     "https://adgear.atlassian.net/browse/MAW-372", "Kirill Timoszenko", "Yes", ""),
    ("Bigos", "MAW-367", "Task", "In Progress", "[SRE] Conversion API project setup",
     "https://adgear.atlassian.net/browse/MAW-367", "Lukasz Kaminski", "No",
     "Missing explicit acceptance criteria; description is a TODO checklist without clear definition of done"),
    ("Bigos", "MAW-365", "Story", "Code Review", "Basic Error Handling",
     "https://adgear.atlassian.net/browse/MAW-365", "Rafal Parol", "Yes", ""),
    ("Bigos", "MAW-364", "Story", "In Progress", "Raw Data Persistence",
     "https://adgear.atlassian.net/browse/MAW-364", "Rahul Tomer", "Yes", ""),
    ("Bigos", "MAW-363", "Story", "Code Review", "Event Schema Validation",
     "https://adgear.atlassian.net/browse/MAW-363", "Rafal Parol", "Yes", ""),
    ("Bigos", "MAW-362", "Story", "Code Review", "Public REST API Endpoint",
     "https://adgear.atlassian.net/browse/MAW-362", "Rafal Parol", "Yes", ""),

    # CAPYBARAS (DoR: proper description as user story, tasks to be done, AC,
    # testable+measurable AC as artifact, estimates, dependencies highlighted)
    ("Capybaras", "RSW-1314", "Task", "In Progress",
     "[SRPOL-OR] Measure performance gains of new WH and snowpark code",
     "https://adgear.atlassian.net/browse/RSW-1314", "Cezary Kret", "Yes", ""),
    ("Capybaras", "RSW-1306", "Task", "In Progress",
     "[AU] E2E Segment Generation Validation - Pipeline Trigger & Size Consistency on PROD",
     "https://adgear.atlassian.net/browse/RSW-1306", "Tomasz Zarod", "Yes", ""),
    ("Capybaras", "RSW-1238", "Task", "Code Review",
     "[SRPOL-OR] Measure cost changes after economic warehouse type introduction",
     "https://adgear.atlassian.net/browse/RSW-1238", "Dariusz Bielawski", "Yes", ""),
    ("Capybaras", "RSW-1097", "Story", "In Progress",
     "[SRPOL-OR] Add dedicated cost tracking view for OR",
     "https://adgear.atlassian.net/browse/RSW-1097", "Piotr Marciniak", "Yes", ""),
    ("Capybaras", "RSW-555", "Story", "In Progress",
     "[SRPOL-OR-26Q1] Create design documentation (high level design) for recommended approach",
     "https://adgear.atlassian.net/browse/RSW-555", "Dariusz Bielawski", "Yes", ""),

    # STURGEONS (DoR: descriptive subject, clarity, specific testable AC, estimates, dependencies in Jira)
    ("Sturgeons", "PEPI-11580", "Task", "In Development", "Add alerting rules in alerting.yaml",
     "https://adgear.atlassian.net/browse/PEPI-11580", "Kornel Pietrzyk", "Yes", ""),

    # EP CORE (DoR: clarity concrete+measurable, clear & testable AC, docs/TechSpecs if needed,
    # estimated with SP, dependencies resolved, Lifeguard label if needed)
    ("EP Core", "EPCW-272", "Task", "In Progress",
     "Okta App creation, Testing and Cutover sonarqube-stg(eks-use1-ep-test-1) service",
     "https://adgear.atlassian.net/browse/EPCW-272", "Nitish Yadav", "Yes", ""),
    ("EP Core", "EPCW-197", "Task", "In Progress", "Upgrade Grafana DEV to 12.x",
     "https://adgear.atlassian.net/browse/EPCW-197", "Michal Mankowski", "Yes", ""),
    ("EP Core", "EPCW-195", "Task", "Code Review",
     "Create Documentation for SonarQube Migration and findings",
     "https://adgear.atlassian.net/browse/EPCW-195", "Nitish Yadav", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-171", "Task", "In Progress",
     "Force top 5 tenants to create cleanup mechanism for their docker images",
     "https://adgear.atlassian.net/browse/EPCW-171", "Michal Mankowski", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-163", "Task", "In Progress", "Create eks-euw1-se-tooling-dev-1",
     "https://adgear.atlassian.net/browse/EPCW-163", "Chandra Prakash", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-158", "Task", "In Progress",
     "Enforce ECR Security Compliance via SCP - Mandatory Security Tag, Scan on Push & KMS Encryption",
     "https://adgear.atlassian.net/browse/EPCW-158", "Stanislaw Cabala", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-155", "Task", "In Progress",
     "[POC] Enable EKS to JFrog connectivity via AWS PrivateLink",
     "https://adgear.atlassian.net/browse/EPCW-155", "Pagalavan", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-153", "Task", "In Progress",
     "Implement JFrog Kubelet Credential Provider on EKS (passwordless image pulls)",
     "https://adgear.atlassian.net/browse/EPCW-153", "Marek Barczyk", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-148", "Bug", "In Progress", "Cleanup NACL rancher configuration",
     "https://adgear.atlassian.net/browse/EPCW-148", "Lukasz Sutkowski", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-145", "Task", "Code Review",
     "Investigate: Alerting likely not working on EKS clusters",
     "https://adgear.atlassian.net/browse/EPCW-145", "Chandra Prakash", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-131", "Task", "In Progress",
     "Migrate Listmonk to EKS eks-use1-ep-prod-1 with domain rename",
     "https://adgear.atlassian.net/browse/EPCW-131", "Bartosz Brodecki", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-96", "Task", "In Progress",
     "Do a Poc testing migration path of ingress-NGINX using Traefik",
     "https://adgear.atlassian.net/browse/EPCW-96", "Juan Antonio Costa Bermudo", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-95", "Task", "In Progress", "Re-try recreation of all master nodes on RKE1",
     "https://adgear.atlassian.net/browse/EPCW-95", "Dariusz Nejbauer", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-22", "Task", "Code Review",
     "Grant EP Data team a dedicated ArgoCD project with admin access",
     "https://adgear.atlassian.net/browse/EPCW-22", "Pagalavan", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-9", "Task", "Code Review",
     "Migrate SonarQube STG on use1-rdev-stg to eks-use1-ep-test-1",
     "https://adgear.atlassian.net/browse/EPCW-9", "Nitish Yadav", "No",
     "Missing explicit testable acceptance criteria"),
    ("EP Core", "EPCW-6", "Task", "In Progress", "Migrate Golinks use1-rdev to eks-use1-ep-prod-1",
     "https://adgear.atlassian.net/browse/EPCW-6", "Lukasz Sutkowski", "No",
     "Missing explicit testable acceptance criteria"),

    # ADS REPORTING (DoR: Bigos+Zurek combined - clarity, AC defined, estimated, dependencies resolved,
    # techspec approved if needed)
    ("Ads Reporting", "PEA-4213", "Task", "In Development",
     "[AUDIENCE SERVICE][DKMS] Test staging env after deploy",
     "https://adgear.atlassian.net/browse/PEA-4213", "Anton Krasnikau", "No",
     "Missing explicit testable acceptance criteria"),
    ("Ads Reporting", "PEA-4205", "Task", "In Development",
     "[UDWMIG][LIFT][STANDARD] Migration instructions",
     "https://adgear.atlassian.net/browse/PEA-4205", "Piotr Jaromin", "No",
     "Missing explicit testable acceptance criteria"),
    ("Ads Reporting", "PEA-4204", "Task", "In Development",
     "[UDWMIG][LIFT][STANDARD] Merge feature branch to master_udw(audience-service)",
     "https://adgear.atlassian.net/browse/PEA-4204", "Piotr Jaromin", "No",
     "Missing explicit testable acceptance criteria"),
    ("Ads Reporting", "PEA-4203", "Task", "In Development",
     "[UDWMIG][LIFT][STANDARD] Merge feature branch to master",
     "https://adgear.atlassian.net/browse/PEA-4203", "Piotr Jaromin", "No",
     "Missing explicit testable acceptance criteria"),
    ("Ads Reporting", "PEA-4196", "Task", "In Development",
     "+ [AUDIENCE SERVICE][DKMS][API] Added exceptions to dkms",
     "https://adgear.atlassian.net/browse/PEA-4196", "Michal Koscielak", "No",
     "Missing explicit testable acceptance criteria"),
    ("Ads Reporting", "PEA-4191", "Task", "In Development",
     "[AUDIENCE SERVICE][JUMPHOST][SPIKE] Investigate when jumphost will be deprecated and prepare plan",
     "https://adgear.atlassian.net/browse/PEA-4191", "Anton Krasnikau", "No",
     "Missing explicit testable acceptance criteria"),
    ("Ads Reporting", "PEA-4161", "Story", "In Development",
     "[TEST] Support for declared IP in RF methodology",
     "https://adgear.atlassian.net/browse/PEA-4161", "Marcin Nowacki", "No",
     "Missing explicit testable acceptance criteria"),
    ("Ads Reporting", "PEA-4160", "Task", "In Development",
     "[UDWMIG][LIFT][STANDARD][SNOWFLAKE] Report data integration tests",
     "https://adgear.atlassian.net/browse/PEA-4160", "Michal Koscielak", "No",
     "Missing explicit testable acceptance criteria"),
    ("Ads Reporting", "PEA-4101", "Bug", "In Development",
     "[BUG] Monthly reach is exceeding the monthly active user count in India for PMP R&F report",
     "https://adgear.atlassian.net/browse/PEA-4101", "Anton Krasnikau", "No",
     "Missing explicit testable acceptance criteria"),
    ("Ads Reporting", "PEA-4067", "Task", "In Development",
     "[UDWMIG][LIFT] Improving deployment on dev schemas",
     "https://adgear.atlassian.net/browse/PEA-4067", "Michal Lis", "No",
     "Missing explicit testable acceptance criteria"),
    ("Ads Reporting", "PEA-3790", "Task", "In Development",
     "[UDWMIG][LIFT][SNOWFLAKE][PRE-AGGREGATION] Fact App Open Event - deployment plan",
     "https://adgear.atlassian.net/browse/PEA-3790", "Marcin Nowacki", "No",
     "Missing explicit testable acceptance criteria"),

    # IGNI (DoR: clarity with context, contact persons for external tasks, AC testable,
    # monitoring/alerting if applicable, estimates, dependencies)
    ("Igni", "ASPW-1126", "Story", "In Progress", "Fix RTB-Bidder pod starting sequence",
     "https://adgear.atlassian.net/browse/ASPW-1126", "Pawel Kwietniewski", "No",
     "Missing acceptance criteria and description context"),
    ("Igni", "ASPW-997", "Story", "In Progress", "Deploy ConsumerD new release",
     "https://adgear.atlassian.net/browse/ASPW-997", "Pawel Kwietniewski", "No",
     "Missing acceptance criteria and description context"),
    ("Igni", "ASPW-689", "Story", "Code Review",
     "CCPA - Transition of Metadata from Erlang Based Implementation to Golang BERT Encoder",
     "https://adgear.atlassian.net/browse/ASPW-689", "Alexander Andersson", "No",
     "Missing explicit testable acceptance criteria"),
    ("Igni", "ASPW-637", "Bug", "In Progress", "consumerd ready file logic is broken",
     "https://adgear.atlassian.net/browse/ASPW-637", "Adam Dolinski", "No",
     "Missing explicit acceptance criteria for bug resolution"),
    ("Igni", "ASPW-560", "Story", "In Progress",
     "Knowledge Transfer session about rewriting objects to new native Golang encoder",
     "https://adgear.atlassian.net/browse/ASPW-560", "Franciszek Strzezek", "No",
     "Missing acceptance criteria and clear definition of done"),
    ("Igni", "ASPW-554", "Story", "In Progress", "Rewrite \"stvd_optouts\" object to use new encoder",
     "https://adgear.atlassian.net/browse/ASPW-554", "Franciszek Strzezek", "No",
     "Missing explicit testable acceptance criteria"),
    ("Igni", "ASPW-6", "Story", "In Progress",
     "Data Activation - consumerd pod takes 5-10 minutes, need to optimize",
     "https://adgear.atlassian.net/browse/ASPW-6", "Alexander Andersson", "No",
     "Missing explicit testable acceptance criteria"),
]

# Summary data for all 15 teams
summary_data = [
    ("Abyss", "Yes", 3),
    ("Radium", "Yes", 11),
    ("Europium", "Yes", 6),
    ("Copernicium", "Yes", 5),
    ("Mouflons", "Yes", 1),
    ("Wolves", "Yes", 0),
    ("Polonium Upper Funnel", "Yes", 3),
    ("Bigos", "Yes", 7),
    ("Capybaras", "Yes", 5),
    ("Sturgeons", "Yes", 1),
    ("Pandas", "Yes", 0),
    ("EP Core", "Yes", 16),
    ("Ads Reporting", "Yes", 11),
    ("Igni", "Yes", 7),
    ("SRPOL SRE", "No", 0),
]

# --- Generate Excel Report ---
wb = Workbook()

# ============ Sheet 1: Summary ============
ws_summary = wb.active
ws_summary.title = "Summary"

# Header styling
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Summary headers
summary_headers = ["Team", "DoR", "Jira Tasks"]
for col_idx, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Summary data
for row_idx, (team, dor, tasks) in enumerate(summary_data, 2):
    ws_summary.cell(row=row_idx, column=1, value=team).border = thin_border
    ws_summary.cell(row=row_idx, column=2, value=dor).border = thin_border
    ws_summary.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
    ws_summary.cell(row=row_idx, column=3, value=tasks).border = thin_border
    ws_summary.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")

# Set column widths for Summary
ws_summary.column_dimensions["A"].width = 25
ws_summary.column_dimensions["B"].width = 10
ws_summary.column_dimensions["C"].width = 12

# Freeze panes
ws_summary.freeze_panes = "A2"

# ============ Sheet 2: DoR Compliance ============
ws_compliance = wb.create_sheet("DoR Compliance")

# Compliance headers
compliance_headers = [
    "Team", "Issue Key", "Issue Type", "Status", "Title", "URL", "Assignee",
    "DoR Compliance", "Feedback"
]
for col_idx, header in enumerate(compliance_headers, 1):
    cell = ws_compliance.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Compliance data
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for row_idx, row_data in enumerate(dor_compliance_data, 2):
    team, key, itype, status, title, url, assignee, compliance, feedback = row_data
    ws_compliance.cell(row=row_idx, column=1, value=team).border = thin_border
    ws_compliance.cell(row=row_idx, column=2, value=key).border = thin_border
    ws_compliance.cell(row=row_idx, column=3, value=itype).border = thin_border
    ws_compliance.cell(row=row_idx, column=4, value=status).border = thin_border
    ws_compliance.cell(row=row_idx, column=5, value=title).border = thin_border
    ws_compliance.cell(row=row_idx, column=6, value=url).border = thin_border
    ws_compliance.cell(row=row_idx, column=7, value=assignee).border = thin_border

    compliance_cell = ws_compliance.cell(row=row_idx, column=8, value=compliance)
    compliance_cell.border = thin_border
    compliance_cell.alignment = Alignment(horizontal="center")
    if compliance == "Yes":
        compliance_cell.fill = green_fill
    else:
        compliance_cell.fill = red_fill

    ws_compliance.cell(row=row_idx, column=9, value=feedback).border = thin_border

# Set column widths for Compliance
col_widths = [15, 12, 10, 12, 40, 50, 15, 15, 60]
for idx, width in enumerate(col_widths, 1):
    ws_compliance.column_dimensions[get_column_letter(idx)].width = width

# Freeze panes
ws_compliance.freeze_panes = "A2"

# Save workbook
output_path = f"{OUTPUT_DIR}\\Report.xlsx"
wb.save(output_path)
print(f"Report saved to: {output_path}")

# --- Statistics ---
total_issues = len(dor_compliance_data)
compliant = sum(1 for d in dor_compliance_data if d[7] == "Yes")
non_compliant = total_issues - compliant
compliance_rate = (compliant / total_issues * 100) if total_issues > 0 else 0

print(f"\nTotal issues analyzed: {total_issues}")
print(f"DoR Compliant: {compliant} ({compliance_rate:.1f}%)")
print(f"Non-Compliant: {non_compliant} ({100-compliance_rate:.1f}%)")

# Per-team stats
teams_analyzed = {}
for row in dor_compliance_data:
    team = row[0]
    if team not in teams_analyzed:
        teams_analyzed[team] = {"total": 0, "compliant": 0}
    teams_analyzed[team]["total"] += 1
    if row[7] == "Yes":
        teams_analyzed[team]["compliant"] += 1

print("\nPer-team compliance:")
for team, stats in teams_analyzed.items():
    rate = stats["compliant"] / stats["total"] * 100
    print(f"  {team}: {stats['compliant']}/{stats['total']} ({rate:.0f}%)")
