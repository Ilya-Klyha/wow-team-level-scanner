#!/usr/bin/env python3
"""
Regenerate Report.xlsx with FIXED schema:
- Sheet 1 "Summary": KPI section + Team/DoR/Jira Tasks table
- Sheet 2 "DoR Compliance": Team, Issue Key, Issue Type, URL, Title, Status, Assignee, DoR Compliance, Note
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "C:/Users/i.klyha/Desktop/Claude/wow-scanner-tool/assets/teams/20260612 11-13"

# Load teams.json
with open(os.path.join(OUTPUT_DIR, "teams.json"), "r") as f:
    teams_data = json.load(f)

# Load all issue data from jira json files
all_issues = []
for team in teams_data["teams"]:
    jira_file = os.path.join(OUTPUT_DIR, team.get("jira_file", ""))
    if os.path.exists(jira_file):
        with open(jira_file, "r") as f:
            jira_data = json.load(f)
            for issue in jira_data.get("issues", []):
                all_issues.append({"team_name": team["name"], **issue})

# Load existing Report.xlsx to get compliance data
from openpyxl import load_workbook
old_wb = load_workbook(os.path.join(OUTPUT_DIR, "Report.xlsx"))
old_ws = old_wb["DoR Compliance"]

# Build compliance lookup from old report
compliance_map = {}
for row in range(2, old_ws.max_row + 1):
    key = old_ws.cell(row, 2).value  # Issue Key
    compliance_pct = old_ws.cell(row, 7).value  # Compliance %
    notes = old_ws.cell(row, 12).value  # Notes

    if key:
        passes = compliance_pct == 100 if compliance_pct is not None else True
        # Build note from failing criteria
        note_text = ""
        if not passes:
            failing_criteria = []
            criteria_cols = {8: "Requirement Clarity", 9: "Acceptance Criteria", 10: "Estimates", 11: "Dependencies"}
            for col, name in criteria_cols.items():
                val = old_ws.cell(row, col).value
                if val and val != "PASS":
                    failing_criteria.append(name)
            if failing_criteria:
                note_text = "Missing: " + ", ".join(failing_criteria)
            elif notes and notes != "All criteria met":
                note_text = notes
            else:
                note_text = "Does not fully meet DoR criteria"

        compliance_map[key] = {"passes": passes, "note": note_text}

old_wb.close()

# --- Generate new Report.xlsx ---
wb = Workbook()

# Styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
kpi_bold = Font(bold=True, size=11)
red_font = Font(bold=True, size=11, color="CC0000")
green_font = Font(bold=True, size=11, color="006600")
orange_font = Font(bold=True, size=11, color="CC6600")

# === SHEET 1: Summary ===
ws_summary = wb.active
ws_summary.title = "Summary"

# Calculate KPIs
total_teams = len(teams_data["teams"])
teams_with_dor = sum(1 for t in teams_data["teams"] if t["extraction_status"] == "success")
teams_with_dor_pct = round(teams_with_dor / total_teams * 100, 1) if total_teams > 0 else 0

# Issues compliant = issues where compliance is Pass
total_analyzed_issues = len(compliance_map)
compliant_issues = sum(1 for v in compliance_map.values() if v["passes"])
issues_compliant_pct = round(compliant_issues / total_analyzed_issues * 100, 1) if total_analyzed_issues > 0 else 0

# Row 1: KPI - Teams with DoR
ws_summary.cell(row=1, column=1).value = "Teams with DoR:"
ws_summary.cell(row=1, column=1).font = kpi_bold
ws_summary.cell(row=1, column=2).value = f"{teams_with_dor_pct}%"
ws_summary.cell(row=1, column=2).font = green_font if teams_with_dor_pct >= 70 else (orange_font if teams_with_dor_pct >= 40 else red_font)

# Row 2: KPI - Issues compliant to DoR
ws_summary.cell(row=2, column=1).value = "Issues fitting DoR:"
ws_summary.cell(row=2, column=1).font = kpi_bold
ws_summary.cell(row=2, column=2).value = f"{issues_compliant_pct}%"
ws_summary.cell(row=2, column=2).font = green_font if issues_compliant_pct >= 70 else (orange_font if issues_compliant_pct >= 40 else red_font)

# Row 3: empty separator

# Row 4: Table header - Team | DoR | Jira Tasks
SUMMARY_COLUMNS = [("Team", 25), ("DoR", 10), ("Jira Tasks", 12)]
for col_idx, (header, width) in enumerate(SUMMARY_COLUMNS, 1):
    cell = ws_summary.cell(row=4, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

# Row 5+: Data rows
for row_idx, team in enumerate(teams_data["teams"], start=5):
    has_dor = team["extraction_status"] == "success"
    dor_value = "Yes" if has_dor else "No"
    jira_count = team.get("jira_issues_count", 0)

    # Column A: Team
    ws_summary.cell(row=row_idx, column=1).value = team["name"]
    ws_summary.cell(row=row_idx, column=1).border = thin_border
    ws_summary.cell(row=row_idx, column=1).alignment = Alignment(vertical='center')

    # Column B: DoR (Yes/No)
    ws_summary.cell(row=row_idx, column=2).value = dor_value
    ws_summary.cell(row=row_idx, column=2).fill = pass_fill if dor_value == "Yes" else fail_fill
    ws_summary.cell(row=row_idx, column=2).border = thin_border
    ws_summary.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center', vertical='center')

    # Column C: Jira Tasks
    ws_summary.cell(row=row_idx, column=3).value = jira_count
    ws_summary.cell(row=row_idx, column=3).border = thin_border
    ws_summary.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center', vertical='center')

ws_summary.freeze_panes = 'A5'

# === SHEET 2: DoR Compliance ===
ws = wb.create_sheet(title="DoR Compliance")

# Columns: Team | Issue Key | Issue Type | URL | Title | Status | Assignee | DoR Compliance | Note
COMPLIANCE_COLUMNS = [
    ("Team", 15),
    ("Issue Key", 12),
    ("Issue Type", 10),
    ("URL", 50),
    ("Title", 40),
    ("Status", 12),
    ("Assignee", 15),
    ("DoR Compliance", 15),
    ("Note", 60)
]

# Write header row
for col_idx, (header, width) in enumerate(COMPLIANCE_COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.freeze_panes = 'A2'

# Write data rows
row_idx = 2
for issue in all_issues:
    team_name = issue.get("team_name", "")
    issue_key = issue.get("key", "")
    issue_type = issue.get("type", "")
    url = issue.get("url", f"https://adgear.atlassian.net/browse/{issue_key}")
    title = issue.get("summary", "")
    status = issue.get("status", "")
    assignee = issue.get("assignee", "Unassigned")

    # Get compliance from map
    comp_data = compliance_map.get(issue_key, {"passes": True, "note": ""})
    dor_compliance = "Pass" if comp_data["passes"] else "Fail"
    note = comp_data["note"] if not comp_data["passes"] else ""

    # Column A: Team
    ws.cell(row=row_idx, column=1).value = team_name
    # Column B: Issue Key
    ws.cell(row=row_idx, column=2).value = issue_key
    # Column C: Issue Type
    ws.cell(row=row_idx, column=3).value = issue_type
    # Column D: URL (hyperlink)
    ws.cell(row=row_idx, column=4).value = url
    ws.cell(row=row_idx, column=4).hyperlink = url
    ws.cell(row=row_idx, column=4).style = "Hyperlink"
    # Column E: Title
    ws.cell(row=row_idx, column=5).value = title
    # Column F: Status
    ws.cell(row=row_idx, column=6).value = status
    # Column G: Assignee
    ws.cell(row=row_idx, column=7).value = assignee
    # Column H: DoR Compliance (Pass/Fail)
    ws.cell(row=row_idx, column=8).value = dor_compliance
    ws.cell(row=row_idx, column=8).fill = pass_fill if dor_compliance == "Pass" else fail_fill
    # Column I: Note
    ws.cell(row=row_idx, column=9).value = note

    # Apply formatting to all cells
    for col in range(1, 10):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=(col in [5, 9]))

    row_idx += 1

# Save
output_path = os.path.join(OUTPUT_DIR, "Report.xlsx")
wb.save(output_path)

# Validation
print(f"[SUCCESS] Report.xlsx regenerated")
print(f"\nSCHEMA VALIDATION:")
print(f"  Sheets: {wb.sheetnames}")
print(f"  Summary: 3 columns (Team, DoR, Jira Tasks)")
print(f"  Summary KPIs: Teams with DoR={teams_with_dor_pct}%, Issues fitting DoR={issues_compliant_pct}%")
print(f"  DoR Compliance: 9 columns (Team, Issue Key, Issue Type, URL, Title, Status, Assignee, DoR Compliance, Note)")
print(f"  DoR Compliance values: Pass={compliant_issues}, Fail={total_analyzed_issues - compliant_issues}")
print(f"  Total data rows: {row_idx - 2}")
