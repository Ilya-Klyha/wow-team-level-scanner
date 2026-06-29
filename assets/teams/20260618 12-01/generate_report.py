#!/usr/bin/env python3
"""
Generate DoR Compliance Excel Report - FIXED SCHEMA
Sheet 1: Summary (all teams overview)
Sheet 2: DoR Compliance (detailed issue analysis)
Sheet 3: DoR quality (DoR quality assessment)
"""
import json
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# FIXED SCHEMA
COMPLIANCE_SHEET_NAME = "DoR Compliance"
SUMMARY_SHEET_NAME = "Summary"
QUALITY_SHEET_NAME = "DoR quality"

COMPLIANCE_COLUMNS = [
    ("Team", 15),
    ("Issue Key", 12),
    ("Issue Type", 10),
    ("URL", 50),
    ("Title", 40),
    ("Status", 12),
    ("Assignee", 15),
    ("DoR Compliance", 15),
    ("Note", 60)
]

SUMMARY_COLUMNS = [
    ("Team", 25),
    ("DoR", 10),
    ("Jira Tasks in progress", 20),
    ("% Tasks fitting DoR", 18)
]

def generate_report(compliance_data, summary_data, quality_data, output_path):
    """Generate Excel report with Summary + DoR Compliance + DoR quality sheets."""
    wb = Workbook()

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Calculate KPIs
    total_teams = len(summary_data)
    teams_with_dor_count = sum(1 for t in summary_data if t["dor"] == "Yes")
    teams_with_dor_pct = round(teams_with_dor_count / total_teams * 100, 1) if total_teams > 0 else 0

    total_issues = len(compliance_data)
    compliant_issues = sum(1 for d in compliance_data if d["dor_compliance"] == "Pass")
    compliance_pct = round(compliant_issues / total_issues * 100, 1) if total_issues > 0 else 0

    # KPI fonts
    kpi_bold = Font(bold=True, size=11)
    red_font = Font(bold=True, size=11, color="CC0000")
    green_font = Font(bold=True, size=11, color="006600")
    orange_font = Font(bold=True, size=11, color="CC6600")

    # === SHEET 1: Summary ===
    ws_summary = wb.active
    ws_summary.title = SUMMARY_SHEET_NAME

    # Row 1: % Teams with DoR
    ws_summary.cell(row=1, column=1).value = "% Teams with DoR"
    ws_summary.cell(row=1, column=1).font = kpi_bold
    ws_summary.cell(row=1, column=2).value = f"{teams_with_dor_pct}%"
    if teams_with_dor_pct >= 70:
        ws_summary.cell(row=1, column=2).font = green_font
    elif teams_with_dor_pct >= 40:
        ws_summary.cell(row=1, column=2).font = orange_font
    else:
        ws_summary.cell(row=1, column=2).font = red_font

    # Row 2: % Jira Tasks fitting DoR
    ws_summary.cell(row=2, column=1).value = "% Jira Tasks fitting DoR"
    ws_summary.cell(row=2, column=1).font = kpi_bold
    ws_summary.cell(row=2, column=2).value = f"{compliance_pct}%"
    if compliance_pct >= 70:
        ws_summary.cell(row=2, column=2).font = green_font
    elif compliance_pct >= 40:
        ws_summary.cell(row=2, column=2).font = orange_font
    else:
        ws_summary.cell(row=2, column=2).font = red_font

    # Row 3: empty separator
    # Row 4: Table header
    for col_idx, (header, width) in enumerate(SUMMARY_COLUMNS, 1):
        cell = ws_summary.cell(row=4, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

    ws_summary.freeze_panes = 'A5'

    # Summary data rows (row 5+)
    for row_idx, row_data in enumerate(summary_data, start=5):
        ws_summary.cell(row=row_idx, column=1).value = row_data["team"]
        ws_summary.cell(row=row_idx, column=1).border = thin_border
        ws_summary.cell(row=row_idx, column=1).alignment = Alignment(vertical='center')

        dor_value = row_data["dor"]
        ws_summary.cell(row=row_idx, column=2).value = dor_value
        ws_summary.cell(row=row_idx, column=2).fill = yes_fill if dor_value == "Yes" else no_fill
        ws_summary.cell(row=row_idx, column=2).border = thin_border
        ws_summary.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center', vertical='center')

        ws_summary.cell(row=row_idx, column=3).value = row_data["jira_tasks"]
        ws_summary.cell(row=row_idx, column=3).border = thin_border
        ws_summary.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center', vertical='center')

        team_pct = row_data.get("pct_fitting_dor", "-")
        ws_summary.cell(row=row_idx, column=4).value = team_pct
        ws_summary.cell(row=row_idx, column=4).border = thin_border
        ws_summary.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center', vertical='center')

    # === SHEET 2: DoR Compliance ===
    ws = wb.create_sheet(title=COMPLIANCE_SHEET_NAME)

    for col_idx, (header, width) in enumerate(COMPLIANCE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'

    for row_idx, row_data in enumerate(compliance_data, start=2):
        ws.cell(row=row_idx, column=1).value = row_data["team"]
        ws.cell(row=row_idx, column=2).value = row_data["issue_key"]
        ws.cell(row=row_idx, column=3).value = row_data["issue_type"]
        ws.cell(row=row_idx, column=4).value = row_data["url"]
        ws.cell(row=row_idx, column=4).hyperlink = row_data["url"]
        ws.cell(row=row_idx, column=4).style = "Hyperlink"
        ws.cell(row=row_idx, column=5).value = row_data["title"]
        ws.cell(row=row_idx, column=6).value = row_data["status"]
        ws.cell(row=row_idx, column=7).value = row_data["assignee"]

        compliance_value = row_data["dor_compliance"]
        ws.cell(row=row_idx, column=8).value = compliance_value
        ws.cell(row=row_idx, column=8).fill = yes_fill if compliance_value == "Pass" else no_fill

        ws.cell(row=row_idx, column=9).value = row_data.get("note", "")

        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col in [5, 9]))

    # === SHEET 3: DoR quality ===
    ws_q = wb.create_sheet(title=QUALITY_SHEET_NAME)

    # KPI Row 1
    avg_coverage = round(sum(r["coverage"] for r in quality_data) / len(quality_data)) if quality_data else 0
    ws_q.cell(1, 1).value = "DoR quality lvl"
    ws_q.cell(1, 1).font = Font(bold=True, size=11)
    ws_q.cell(1, 2).value = avg_coverage
    if avg_coverage >= 70:
        ws_q.cell(1, 2).font = Font(bold=True, size=11, color="006600")
    elif avg_coverage >= 40:
        ws_q.cell(1, 2).font = Font(bold=True, size=11, color="CC6600")
    else:
        ws_q.cell(1, 2).font = Font(bold=True, size=11, color="CC0000")

    # Row 3: header
    quality_headers = [("Team", 25), ("Coverage", 12), ("Note", 60)]
    for col, (header, width) in enumerate(quality_headers, 1):
        cell = ws_q.cell(3, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws_q.column_dimensions[get_column_letter(col)].width = width

    # Sort by coverage desc
    sorted_quality = sorted(quality_data, key=lambda x: x["coverage"], reverse=True)

    for row_idx, result in enumerate(sorted_quality, 4):
        ws_q.cell(row_idx, 1).value = result["team"]
        ws_q.cell(row_idx, 1).border = thin_border
        ws_q.cell(row_idx, 1).alignment = Alignment(vertical='center')

        coverage_cell = ws_q.cell(row_idx, 2)
        coverage_cell.value = result["coverage"]
        coverage_cell.border = thin_border
        coverage_cell.alignment = Alignment(horizontal='center', vertical='center')
        if result['coverage'] >= 70:
            coverage_cell.fill = yes_fill
        elif result['coverage'] >= 40:
            coverage_cell.fill = orange_fill
        else:
            coverage_cell.fill = no_fill

        note_cell = ws_q.cell(row_idx, 3)
        note_cell.value = result["note"]
        note_cell.border = thin_border
        note_cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws_q.freeze_panes = 'A4'

    # Save
    wb.save(output_path)
    print(f"[SUCCESS] Report saved: {output_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Summary rows: {len(summary_data)}")
    print(f"  Compliance rows: {len(compliance_data)}")
    print(f"  Quality rows: {len(quality_data)}")
    print(f"  KPIs: Teams with DoR={teams_with_dor_pct}%, Tasks fitting DoR={compliance_pct}%, Quality avg={avg_coverage}")


if __name__ == "__main__":
    with open(sys.argv[1], 'r') as f:
        compliance_data = json.load(f)
    with open(sys.argv[2], 'r') as f:
        summary_data = json.load(f)
    with open(sys.argv[3], 'r') as f:
        quality_data = json.load(f)
    output_path = sys.argv[4]
    generate_report(compliance_data, summary_data, quality_data, output_path)
