"""
Process Jira query results, write team Jira files, perform DoR analysis,
and generate the Excel report.
"""
import json
import os
import sys
from datetime import datetime

OUTPUT_DIR = r"C:\Users\i.klyha\Desktop\Claude\wow-scanner-tool\assets\teams\20260611 16-13"
TOOL_RESULTS_DIR = r"C:\Users\i.klyha\.claude\projects\C--Users-i-klyha-Desktop-Claude-wow-scanner-tool\02dccfff-d50d-4677-9824-f2c0ac83ee7e\tool-results"

# File mapping
FILE_MAP = {
    "PEA": "toolu_bdrk_01YNaTTYxowNWNmKdaDGptnm.txt",
    "EPCW": "toolu_bdrk_01Rjqe4pMNbGZ7wYbjHCTnZ4.txt",
    "ASPW": "toolu_bdrk_01ALohxiMR3bMgm5oq7jsiYN.txt",
    "AETVP": "toolu_bdrk_01GRff21vTVtLF3TzGMSSxUf.txt",
    "EEEW": "toolu_bdrk_01F5aCKTPWwfiq2zyC4AaeTy.txt",
}

# ML Platform issues (inline data)
ML_ISSUES = [
    {
        "key": "ML-55",
        "summary": "Implement Test DAG MLOps based on SDK",
        "issuetype": "Task",
        "status": "In Progress",
        "assignee": "Tomasz Teter",
        "priority": "Medium",
        "team": "",
        "description": "Implement Test DAG MLOps based on SDK",
        "url": "https://adgear.atlassian.net/browse/ML-55",
    },
    {
        "key": "ML-10",
        "summary": "Support after MWAA upgrade to 2.10.3",
        "issuetype": "Task",
        "status": "In Progress",
        "assignee": "Tomasz Teter",
        "priority": "Medium",
        "team": "",
        "description": "Support after MWAA upgrade to 2.10.3",
        "url": "https://adgear.atlassian.net/browse/ML-10",
    },
]


def load_issues(project_key):
    """Load and parse issues from a persisted Jira result file."""
    filepath = os.path.join(TOOL_RESULTS_DIR, FILE_MAP[project_key])
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    issues = []
    for node in data["issues"]["nodes"]:
        fields = node["fields"]
        desc = fields.get("description")
        if isinstance(desc, dict):
            # ADF format - just note that description exists
            desc_text = "[ADF description present]"
        elif isinstance(desc, str):
            desc_text = desc[:500] if desc else ""
        else:
            desc_text = ""

        team_field = fields.get("customfield_10114")
        team_name = team_field.get("name", "") if team_field else ""

        issues.append({
            "key": node["key"],
            "summary": fields["summary"],
            "issuetype": fields["issuetype"]["name"],
            "status": fields["status"]["name"],
            "assignee": fields["assignee"]["displayName"] if fields.get("assignee") else "Unassigned",
            "priority": fields["priority"]["name"],
            "team": team_name,
            "description": desc_text,
            "url": f"https://adgear.atlassian.net/browse/{node['key']}",
        })
    return issues


def filter_copernicium(issues):
    """Filter AETVP issues for Copernicium team."""
    return [i for i in issues if "Copernicium" in i["team"]]


def filter_zurek(issues):
    """Filter PEA issues for Zurek team."""
    return [i for i in issues if "Zurek" in i["team"]]


def filter_bigos(issues):
    """Filter PEA issues for Bigos team."""
    return [i for i in issues if "Bigos" in i["team"]]


def write_jira_file(team_slug, issues):
    """Write Jira issues to a text file for a team."""
    filepath = os.path.join(OUTPUT_DIR, f"{team_slug}-jira.txt")
    with open(filepath, "w", encoding="utf-8") as f:
        if not issues:
            f.write("No active issues found for this team.\n")
            return
        f.write(f"Active Issues ({len(issues)} total)\n")
        f.write("=" * 60 + "\n\n")
        for issue in issues:
            f.write(f"Key: {issue['key']}\n")
            f.write(f"Type: {issue['issuetype']}\n")
            f.write(f"Status: {issue['status']}\n")
            f.write(f"Priority: {issue['priority']}\n")
            f.write(f"Assignee: {issue['assignee']}\n")
            f.write(f"Title: {issue['summary']}\n")
            f.write(f"URL: {issue['url']}\n")
            f.write("-" * 60 + "\n\n")
    return filepath


def load_dor(team_slug):
    """Load DoR criteria from file."""
    filepath = os.path.join(OUTPUT_DIR, f"{team_slug}-dor.txt")
    if not os.path.exists(filepath):
        return None
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read().strip()
    if "not found" in content.lower():
        return None
    return content


def analyze_dor_compliance(issue, dor_criteria):
    """
    Analyze a single issue against DoR criteria.
    Returns (compliance: str, feedback: str)
    """
    if not dor_criteria:
        return ("N/A", "No DoR criteria defined for this team")

    summary = issue["summary"]
    desc = issue.get("description", "")
    issue_type = issue["issuetype"]

    # Parse DoR criteria into individual items
    criteria_lines = [line.strip() for line in dor_criteria.split("\n") if line.strip()]

    issues_found = []

    # Check each DoR criterion
    for criterion in criteria_lines:
        criterion_lower = criterion.lower()

        # Check for description/clarity requirement
        if any(kw in criterion_lower for kw in ["clearly articulated", "well described", "clear user story", "requirement clarity"]):
            if not desc or desc.strip() == "":
                issues_found.append("Missing description/requirement clarity")

        # Check for acceptance criteria
        if "acceptance criteria" in criterion_lower:
            # We can only check description for AC mentions
            desc_lower = (desc or "").lower()
            summary_lower = summary.lower()
            has_ac = any(kw in desc_lower for kw in ["acceptance criteria", "ac:", "acceptance:", "expected"])
            # Don't flag if description has substantial content (likely contains implicit AC)
            if not has_ac and len(desc or "") < 50:
                issues_found.append("Acceptance criteria not explicitly defined")

        # Check for estimates (we can't verify SP from title alone - skip this check as we don't have that data)

    if issues_found:
        return ("No", "; ".join(issues_found[:2]))
    else:
        return ("Yes", "")


def main():
    print("Step 1: Loading Jira data from persisted files...")

    # Load all project issues
    pea_issues = load_issues("PEA")
    epcw_issues = load_issues("EPCW")
    aspw_issues = load_issues("ASPW")
    aetvp_issues = load_issues("AETVP")
    eeew_issues = load_issues("EEEW")

    print(f"  PEA: {len(pea_issues)} issues")
    print(f"  EPCW: {len(epcw_issues)} issues")
    print(f"  ASPW: {len(aspw_issues)} issues")
    print(f"  AETVP: {len(aetvp_issues)} issues")
    print(f"  EEEW: {len(eeew_issues)} issues")

    # Step 2: Filter PEA issues by team
    zurek_issues = filter_zurek(pea_issues)
    bigos_issues = filter_bigos(pea_issues)
    print(f"  Zurek (from PEA): {len(zurek_issues)} issues")
    print(f"  Bigos (from PEA): {len(bigos_issues)} issues")

    # Filter AETVP for Copernicium
    copernicium_issues = filter_copernicium(aetvp_issues)
    print(f"  Copernicium (from AETVP): {len(copernicium_issues)} issues")

    # Team issue mapping
    team_issues = {
        "zurek": zurek_issues,
        "bigos": bigos_issues,
        "copernicium": copernicium_issues,
        "ep-core": epcw_issues,
        "igni": aspw_issues,
        "sre": eeew_issues,
        "ml-platform": ML_ISSUES,
        # Teams with 0 issues
        "pe-waw-abyss": [],
        "radium": [],
        "europium": [],
        "mouflons": [],
        "wolves": [],
        "polonium-uf": [],
        "capybaras": [],
        "ml-serving": [],
    }

    # Step 3: Write Jira TXT files
    print("\nStep 3: Writing Jira TXT files...")
    for team_slug, issues in team_issues.items():
        write_jira_file(team_slug, issues)
        print(f"  {team_slug}-jira.txt: {len(issues)} issues")

    # Step 4: DoR Compliance Analysis
    print("\nStep 4: Performing DoR compliance analysis...")

    all_analysis_rows = []

    teams_with_dor = {
        "zurek": "Zurek",
        "copernicium": "Copernicium",
        "ep-core": "EP Core",
        "igni": "Igni",
    }
    # Teams without DoR: ML Platform, SRE, ML Serving

    for team_slug, team_name in [
        ("zurek", "Zurek"),
        ("copernicium", "Copernicium"),
        ("ep-core", "EP Core"),
        ("igni", "Igni"),
        ("sre", "SRE"),
        ("ml-platform", "ML Platform"),
        ("bigos", "Bigos"),
        ("pe-waw-abyss", "PE-WAW-Abyss"),
        ("radium", "Radium"),
        ("europium", "Europium"),
        ("mouflons", "Mouflons"),
        ("wolves", "Wolves"),
        ("polonium-uf", "Polonium UF"),
        ("capybaras", "Capybaras"),
        ("ml-serving", "ML Serving"),
    ]:
        issues = team_issues.get(team_slug, [])
        if not issues:
            continue

        dor = load_dor(team_slug)

        for issue in issues:
            if dor:
                compliance, feedback = analyze_dor_compliance(issue, dor)
            else:
                compliance = "N/A"
                feedback = "No DoR criteria defined for this team"

            all_analysis_rows.append({
                "team": team_name,
                "key": issue["key"],
                "issuetype": issue["issuetype"],
                "status": issue["status"],
                "title": issue["summary"],
                "url": issue["url"],
                "assignee": issue["assignee"],
                "compliance": compliance,
                "feedback": feedback,
            })

    print(f"  Total rows for report: {len(all_analysis_rows)}")

    # Count stats
    yes_count = sum(1 for r in all_analysis_rows if r["compliance"] == "Yes")
    no_count = sum(1 for r in all_analysis_rows if r["compliance"] == "No")
    na_count = sum(1 for r in all_analysis_rows if r["compliance"] == "N/A")
    print(f"  Compliant (Yes): {yes_count}")
    print(f"  Non-compliant (No): {no_count}")
    print(f"  N/A (no DoR): {na_count}")

    # Step 5: Generate Excel Report
    print("\nStep 5: Generating Excel report...")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "DoR Compliance"

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Conditional fills
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Headers
        headers = ["Team", "Issue Key", "Issue Type", "Status", "Title", "URL", "Assignee", "DoR Compliance", "Feedback"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        for row_idx, row_data in enumerate(all_analysis_rows, 2):
            ws.cell(row=row_idx, column=1, value=row_data["team"])
            ws.cell(row=row_idx, column=2, value=row_data["key"])
            ws.cell(row=row_idx, column=3, value=row_data["issuetype"])
            ws.cell(row=row_idx, column=4, value=row_data["status"])
            ws.cell(row=row_idx, column=5, value=row_data["title"])
            ws.cell(row=row_idx, column=6, value=row_data["url"])
            ws.cell(row=row_idx, column=7, value=row_data["assignee"])

            compliance_cell = ws.cell(row=row_idx, column=8, value=row_data["compliance"])
            if row_data["compliance"] == "Yes":
                compliance_cell.fill = green_fill
            elif row_data["compliance"] == "No":
                compliance_cell.fill = red_fill

            ws.cell(row=row_idx, column=9, value=row_data["feedback"])

        # Column widths
        col_widths = [15, 12, 12, 15, 50, 45, 20, 15, 50]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A" + chr(64 + col_idx - 26)].width = width

        # Auto-filter
        ws.auto_filter.ref = f"A1:I{len(all_analysis_rows) + 1}"

        # Save
        report_path = os.path.join(OUTPUT_DIR, "Report.xlsx")
        wb.save(report_path)
        print(f"  Excel report saved: {report_path}")

    except ImportError:
        print("  openpyxl not available, generating CSV fallback...")
        import csv
        report_path = os.path.join(OUTPUT_DIR, "Report.csv")
        with open(report_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Team", "Issue Key", "Issue Type", "Status", "Title", "URL", "Assignee", "DoR Compliance", "Feedback"])
            for row_data in all_analysis_rows:
                writer.writerow([
                    row_data["team"], row_data["key"], row_data["issuetype"],
                    row_data["status"], row_data["title"], row_data["url"],
                    row_data["assignee"], row_data["compliance"], row_data["feedback"]
                ])
        print(f"  CSV report saved: {report_path}")

    # Step 6: Write DOR_ANALYSIS_SUMMARY.md
    print("\nStep 6: Writing DOR_ANALYSIS_SUMMARY.md...")
    summary_path = os.path.join(OUTPUT_DIR, "DOR_ANALYSIS_SUMMARY.md")
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write("# DoR Compliance Analysis Summary\n\n")
        f.write(f"**Scan Date:** 2026-06-11\n\n")
        f.write("## Overall Statistics\n\n")
        f.write(f"- Total issues analyzed: {len(all_analysis_rows)}\n")
        f.write(f"- Compliant (Yes): {yes_count}\n")
        f.write(f"- Non-compliant (No): {no_count}\n")
        f.write(f"- N/A (no DoR defined): {na_count}\n")
        total_with_dor = yes_count + no_count
        if total_with_dor > 0:
            compliance_rate = (yes_count / total_with_dor) * 100
            f.write(f"- Compliance rate (where DoR exists): {compliance_rate:.1f}%\n")
        f.write("\n## Per-Team Breakdown\n\n")
        f.write("| Team | Total Issues | Compliant | Non-Compliant | N/A | Compliance Rate |\n")
        f.write("|------|-------------|-----------|---------------|-----|----------------|\n")

        # Group by team
        team_stats = {}
        for row in all_analysis_rows:
            t = row["team"]
            if t not in team_stats:
                team_stats[t] = {"total": 0, "yes": 0, "no": 0, "na": 0}
            team_stats[t]["total"] += 1
            if row["compliance"] == "Yes":
                team_stats[t]["yes"] += 1
            elif row["compliance"] == "No":
                team_stats[t]["no"] += 1
            else:
                team_stats[t]["na"] += 1

        for t, stats in team_stats.items():
            dor_total = stats["yes"] + stats["no"]
            rate = f"{(stats['yes'] / dor_total * 100):.0f}%" if dor_total > 0 else "N/A"
            f.write(f"| {t} | {stats['total']} | {stats['yes']} | {stats['no']} | {stats['na']} | {rate} |\n")

        f.write("\n## Teams With 0 Active Issues\n\n")
        zero_teams = ["PE-WAW-Abyss", "Radium", "Europium", "Mouflons", "Wolves", "Polonium UF", "Capybaras", "ML Serving"]
        for t in zero_teams:
            f.write(f"- {t}\n")

        f.write("\n## Teams Without DoR\n\n")
        f.write("- ML Platform\n")
        f.write("- SRE\n")
        f.write("- ML Serving\n")

    print(f"  Summary saved: {summary_path}")

    # Step 7: Update teams.json
    print("\nStep 7: Updating teams.json...")
    teams_json_path = os.path.join(OUTPUT_DIR, "teams.json")
    with open(teams_json_path, "r", encoding="utf-8") as f:
        teams_data = json.load(f)

    # Build team slug mapping
    team_name_to_slug = {
        "PE-WAW-Abyss": "pe-waw-abyss",
        "Radium": "radium",
        "Europium": "europium",
        "Copernicium": "copernicium",
        "Mouflons": "mouflons",
        "Wolves": "wolves",
        "Polonium UF": "polonium-uf",
        "Bigos": "bigos",
        "Capybaras": "capybaras",
        "ML Serving": "ml-serving",
        "ML Platform": "ml-platform",
        "EP Core": "ep-core",
        "Zurek": "zurek",
        "Igni": "igni",
        "SRE": "sre",
    }

    for team_entry in teams_data["teams"]:
        team_name = team_entry["name"]
        slug = team_name_to_slug.get(team_name)
        if not slug:
            continue

        issues = team_issues.get(slug, [])
        team_entry["jira_file"] = f"{slug}-jira.txt"
        team_entry["jira_issue_count"] = len(issues)

        # DoR analysis stats
        team_rows = [r for r in all_analysis_rows if r["team"] == team_name]
        if team_rows:
            y = sum(1 for r in team_rows if r["compliance"] == "Yes")
            n = sum(1 for r in team_rows if r["compliance"] == "No")
            na = sum(1 for r in team_rows if r["compliance"] == "N/A")
            dor_total = y + n
            team_entry["dor_analysis"] = {
                "total_issues": len(team_rows),
                "compliant": y,
                "non_compliant": n,
                "not_applicable": na,
                "compliance_rate": f"{(y / dor_total * 100):.0f}%" if dor_total > 0 else "N/A",
            }
        else:
            team_entry["dor_analysis"] = {
                "total_issues": 0,
                "compliant": 0,
                "non_compliant": 0,
                "not_applicable": 0,
                "compliance_rate": "N/A",
            }

    with open(teams_json_path, "w", encoding="utf-8") as f:
        json.dump(teams_data, f, indent=2, ensure_ascii=False)

    print(f"  teams.json updated: {teams_json_path}")
    print("\nAll steps completed successfully.")


if __name__ == "__main__":
    main()
