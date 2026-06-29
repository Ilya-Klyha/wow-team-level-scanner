"""Generate Report-DoD.xlsx with Summary sheet for DoD Scanner."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_PATH = os.path.join(OUTPUT_DIR, "Report-DoD.xlsx")

# Team data: (name, has_dod)
teams = [
    ("Abyss", True),
    ("Radium", True),
    ("Europium", True),
    ("Copernicium", True),
    ("Mouflons", True),
    ("Wolves", True),
    ("Polonium UF", True),
    ("Bigos", True),
    ("Capybaras", True),
    ("ML Serving Sturgeons", False),
    ("ML Platform Pandas", True),
    ("EP Core", True),
    ("Zurek", True),
    ("Igni", True),
    ("SRE", False),
]

# Calculate KPI
teams_with_dod = sum(1 for _, has in teams if has)
total_teams = len(teams)
pct = round((teams_with_dod / total_teams) * 100, 1)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Summary"

# Styles
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
kpi_font = Font(name="Calibri", bold=True, size=14)
yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yes_font = Font(name="Calibri", color="006100")
no_font = Font(name="Calibri", color="9C0006")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# KPI row
ws.merge_cells("A1:B1")
ws["A1"] = "% Teams with DoD"
ws["A1"].font = Font(name="Calibri", bold=True, size=11)
ws["C1"] = f"{pct}%"
ws["C1"].font = kpi_font
if pct >= 80:
    ws["C1"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
elif pct >= 50:
    ws["C1"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
else:
    ws["C1"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Blank row
ws.append([])

# Header row
headers = ["Team", "DoD"]
ws.append(headers)
header_row = 3
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Data rows
for row_idx, (team_name, has_dod) in enumerate(teams, header_row + 1):
    ws.cell(row=row_idx, column=1, value=team_name).border = thin_border
    dod_cell = ws.cell(row=row_idx, column=2, value="Yes" if has_dod else "No")
    dod_cell.border = thin_border
    dod_cell.alignment = Alignment(horizontal="center")
    if has_dod:
        dod_cell.fill = yes_fill
        dod_cell.font = yes_font
    else:
        dod_cell.fill = no_fill
        dod_cell.font = no_font

# Column widths
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12

# Save
wb.save(REPORT_PATH)
print(f"Report saved to: {REPORT_PATH}")
print(f"KPI: {teams_with_dod}/{total_teams} teams ({pct}%) have DoD defined")
