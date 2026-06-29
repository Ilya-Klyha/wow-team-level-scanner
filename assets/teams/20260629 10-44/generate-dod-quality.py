"""Add DoD Quality sheet to Report-DoD.xlsx based on industry and company standards."""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_PATH = os.path.join(OUTPUT_DIR, "Report-DoD.xlsx")

# 7 dimensions with weights from dod-standard.txt
DIMENSIONS = [
    ("Coverage", 25),
    ("Clarity", 20),
    ("Measurability", 15),
    ("Company Alignment", 15),
    ("Industry", 10),
    ("Actionability", 10),
    ("Evidence", 5),
]

# Industry standard: 10 essential DoD areas
# 1. Code Review, 2. Unit/Integration Tests, 3. Acceptance Criteria, 4. CI/CD Pipeline,
# 5. Deployed to Staging/Production, 6. Documentation Updated, 7. No Known Defects,
# 8. Performance/Security, 9. Monitoring/Alerting, 10. PO/Stakeholder Acceptance

# Company standard (page 21735212005) key principles:
# - Team-owned policy
# - Clear, observable criteria (checklist form)
# - Covers task/story/feature/release levels
# - Visible, achievable, evolutionary
# - Applied consistently when closing work
# - Reviewed in retrospectives

# Assessment per team (scores 0-100 per dimension)
# Based on analysis of each team's DoD content against standards
assessments = {
    "Abyss": {
        "Coverage": 70,  # 4 criteria covering tests, CI, AC, deployment - missing code review, docs standalone, monitoring, PO acceptance, perf/security, no defects
        "Clarity": 75,  # Good descriptions, specific about what each criterion means
        "Measurability": 70,  # Most criteria pass/fail checkable, some vague ("measurement ready")
        "Company Alignment": 75,  # Checklist form, team-owned, covers story level
        "Industry": 65,  # Covers 4/10 essential areas well
        "Actionability": 60,  # Working agreement adds UX subtask flow, but no clear verification methods
        "Evidence": 50,  # No specific evidence requirements stated
    },
    "Radium": {
        "Coverage": 70,  # 7 criteria - tests, CI, local test, AC, deploy, UX, external review
        "Clarity": 50,  # Very terse descriptions ("CI pass"), some unclear ("Automated Pipeline ??")
        "Measurability": 55,  # Some clear pass/fail, but "Automated Pipeline ??" shows uncertainty
        "Company Alignment": 60,  # Checklist form, but seems incomplete/draft-like
        "Industry": 55,  # Covers testing, CI, deploy, AC but lacks docs, monitoring, PO acceptance
        "Actionability": 50,  # "Reviewers are notified" is actionable, but overall weak guidance
        "Evidence": 40,  # No evidence requirements
    },
    "Europium": {
        "Coverage": 85,  # Engineering, Quality, UX, Operational categories - comprehensive
        "Clarity": 85,  # Well-structured with clear categories and specific items
        "Measurability": 80,  # Clear pass/fail for most items, "Not done if" section adds clarity
        "Company Alignment": 85,  # Multi-level (engineering/quality/UX/operational), evolutionary
        "Industry": 80,  # Covers code review, tests, AC, CI, deploy, docs, monitoring, UX validation
        "Actionability": 80,  # Clear workflow integration (tag UX, Jira comments, pre-prod validation)
        "Evidence": 70,  # Specifies where to tag, what to inform, production-like environment
    },
    "Copernicium": {
        "Coverage": 70,  # 5 criteria - tests, UX QA, AC+monitoring+docs, deploy, dependencies
        "Clarity": 65,  # Decent descriptions but some bundled (AC + monitoring + docs in one)
        "Measurability": 65,  # Most verifiable, but bundled criteria make it harder
        "Company Alignment": 70,  # Table format, covers story level, team-owned
        "Industry": 60,  # Covers tests, AC, deploy, docs (bundled), UX - missing code review, monitoring standalone
        "Actionability": 55,  # "Review must be accepted" is actionable, but limited workflow detail
        "Evidence": 40,  # No specific evidence requirements
    },
    "Mouflons": {
        "Coverage": 55,  # 2 principles: team acceptance + developer quality. Limited scope
        "Clarity": 70,  # Well-written principles with clear expectations
        "Measurability": 60,  # "SonarQube quality gate passes" is measurable, PO verification clear
        "Company Alignment": 70,  # Principle-based, team-owned, includes Jira documentation
        "Industry": 50,  # Covers AC, code review, testing, documentation - but missing deploy, CI, monitoring
        "Actionability": 75,  # Good Jira workflow (comment before closing, status states), SonarQube integration
        "Evidence": 65,  # Jira comments required, SonarQube gate, ticket state tracking
    },
    "Wolves": {
        "Coverage": 55,  # Same structure as Mouflons - 2 principles
        "Clarity": 70,  # Well-written, clear expectations
        "Measurability": 60,  # SonarQube measurable, PO verification clear
        "Company Alignment": 70,  # Principle-based, team-owned
        "Industry": 50,  # Same coverage gaps as Mouflons
        "Actionability": 75,  # Good Jira workflow, SonarQube integration
        "Evidence": 65,  # Jira comments, SonarQube, status tracking
    },
    "Polonium UF": {
        "Coverage": 55,  # 6 items - AC, tests, review, merge/release, PR description, docs
        "Clarity": 60,  # Concise but somewhat generic items
        "Measurability": 65,  # Most are pass/fail checkable
        "Company Alignment": 60,  # Simple checklist, but covers basics
        "Industry": 55,  # Covers AC, tests, code review, deploy, docs - missing CI explicit, monitoring, PO formal
        "Actionability": 50,  # "nice-to-have" qualifier weakens review requirement
        "Evidence": 45,  # PR description mentioned but no other evidence spec
    },
    "Bigos": {
        "Coverage": 45,  # 5 items but high-level - AC, no defects, approval, release, no negative impact
        "Clarity": 55,  # Some items clear, but "not have negative impact" is vague
        "Measurability": 50,  # "No known open serious defects" subjective, "negative impact" unmeasurable
        "Company Alignment": 60,  # Team-owned, includes PO/tech lead approval
        "Industry": 45,  # Covers AC, release, PO acceptance - missing tests, CI, docs, monitoring
        "Actionability": 55,  # Clear approval chain (PO or tech lead), release requirement
        "Evidence": 40,  # No specific evidence requirements
    },
    "Capybaras": {
        "Coverage": 60,  # 4 criteria - result summary, tests (TBD), AC, deployment pipeline
        "Clarity": 65,  # Deployment criteria well detailed, but tests marked "TBD"
        "Measurability": 65,  # Deployment steps are very specific and measurable
        "Company Alignment": 65,  # Table format, team-owned, includes verification steps
        "Industry": 55,  # Covers AC, tests (TBD), deploy pipeline - missing code review, docs, monitoring
        "Actionability": 70,  # Very specific deployment verification steps
        "Evidence": 60,  # "Positive merge of PR", "Positive deploy", "Result summary in Jira"
    },
    "ML Platform Pandas": {
        "Coverage": 65,  # 6 criteria - github checks, tests, AC+docs, AI context, dependencies, demo
        "Clarity": 70,  # Clear individual criteria
        "Measurability": 70,  # Github checks pass/fail, demo requirement, AC verification
        "Company Alignment": 75,  # Includes demo requirement (review), documentation, team-owned
        "Industry": 60,  # Covers tests, CI (github checks), AC, docs, code review (dependencies) - missing deploy, monitoring
        "Actionability": 70,  # Demo before sprint close, AI context update - specific actions
        "Evidence": 60,  # "links to artifacts will be provided", demo requirement
    },
    "EP Core": {
        "Coverage": 45,  # 4 items - AC, approval, docs/runbook, demo
        "Clarity": 60,  # Clear but very brief
        "Measurability": 60,  # Demo and approval are verifiable, docs requirement clear
        "Company Alignment": 65,  # Includes component lead approval, demo requirement
        "Industry": 40,  # Missing tests, CI, deploy, monitoring, code review
        "Actionability": 60,  # Clear approval chain, demo before close
        "Evidence": 55,  # Demo serves as evidence, runbook requirement
    },
    "Zurek": {
        "Coverage": 60,  # 7 items - tested, branch updated, feature branch test, AC, rollback, docs, merge
        "Clarity": 55,  # Some unclear phrasing ("described in task if tests are not needed")
        "Measurability": 60,  # Most items checkable, but "should be tested" is vague
        "Company Alignment": 60,  # Simple list, team-owned, covers deploy/merge
        "Industry": 55,  # Covers testing, deploy (merge to main), docs, rollback - missing CI, monitoring, PO
        "Actionability": 55,  # Rollback plan requirement is good, but no clear workflow
        "Evidence": 45,  # Merge to main/release is traceable, but limited otherwise
    },
    "Igni": {
        "Coverage": 75,  # 7 criteria - tests, CI, AC, docs (TBD), deploy (2 types), comment description
        "Clarity": 75,  # Good descriptions, differentiated deploy criteria for different services
        "Measurability": 70,  # CI pipeline pass/fail, AC verification, deploy confirmation
        "Company Alignment": 75,  # Table format, covers multiple work types (Data Activation vs Consumerd)
        "Industry": 70,  # Covers tests, CI, AC, docs (TBD), deploy, documentation trail
        "Actionability": 70,  # Specific deploy rules per service type, comment requirement
        "Evidence": 65,  # "Short description in comment what was done" - explicit evidence trail
    },
}

# Calculate weighted scores
def calc_weighted_score(scores):
    total = 0
    for dim, weight in DIMENSIONS:
        total += scores[dim] * weight / 100
    return round(total, 1)


# Styles
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def get_score_fill(score):
    if score >= 75:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif score >= 50:
        return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def get_score_font(score):
    if score >= 75:
        return Font(name="Calibri", color="006100")
    elif score >= 50:
        return Font(name="Calibri", color="9C5700")
    else:
        return Font(name="Calibri", color="9C0006")


# Load existing workbook and add quality sheet
wb = load_workbook(REPORT_PATH)
if "DoD quality" in wb.sheetnames:
    del wb["DoD quality"]
ws = wb.create_sheet("DoD quality")

# Headers
headers = ["Team"] + [f"{dim} ({w}%)" for dim, w in DIMENSIONS] + ["Weighted Score"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

# Data rows
row_idx = 2
for team_name, scores in assessments.items():
    ws.cell(row=row_idx, column=1, value=team_name).border = thin_border
    for col_idx, (dim, _) in enumerate(DIMENSIONS, 2):
        score = scores[dim]
        cell = ws.cell(row=row_idx, column=col_idx, value=score)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        cell.fill = get_score_fill(score)
        cell.font = get_score_font(score)
    # Weighted score
    weighted = calc_weighted_score(scores)
    cell = ws.cell(row=row_idx, column=len(DIMENSIONS) + 2, value=weighted)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")
    cell.font = Font(name="Calibri", bold=True)
    cell.fill = get_score_fill(weighted)
    row_idx += 1

# Teams without DoD
no_dod_teams = ["ML Serving Sturgeons", "SRE"]
for team_name in no_dod_teams:
    ws.cell(row=row_idx, column=1, value=team_name).border = thin_border
    for col_idx in range(2, len(DIMENSIONS) + 3):
        cell = ws.cell(row=row_idx, column=col_idx, value="N/A")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(name="Calibri", italic=True, color="808080")
    row_idx += 1

# Average row
ws.cell(row=row_idx + 1, column=1, value="Average (teams with DoD)").font = Font(bold=True)
ws.cell(row=row_idx + 1, column=1).border = thin_border
all_weighted = [calc_weighted_score(s) for s in assessments.values()]
avg_weighted = round(sum(all_weighted) / len(all_weighted), 1)
for col_idx, (dim, _) in enumerate(DIMENSIONS, 2):
    avg_dim = round(sum(assessments[t][dim] for t in assessments) / len(assessments), 1)
    cell = ws.cell(row=row_idx + 1, column=col_idx, value=avg_dim)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")
    cell.font = Font(bold=True)
avg_cell = ws.cell(row=row_idx + 1, column=len(DIMENSIONS) + 2, value=avg_weighted)
avg_cell.border = thin_border
avg_cell.alignment = Alignment(horizontal="center")
avg_cell.font = Font(name="Calibri", bold=True, size=12)
avg_cell.fill = get_score_fill(avg_weighted)

# Column widths
ws.column_dimensions["A"].width = 24
for col_idx in range(2, len(DIMENSIONS) + 3):
    ws.column_dimensions[chr(64 + col_idx)].width = 16

# Save
wb.save(REPORT_PATH)
print(f"DoD Quality sheet added to: {REPORT_PATH}")
print(f"\nWeighted Scores by Team:")
for team, scores in assessments.items():
    print(f"  {team}: {calc_weighted_score(scores)}")
print(f"\nAverage weighted score: {avg_weighted}")
print(f"Teams without DoD: {', '.join(no_dod_teams)}")
