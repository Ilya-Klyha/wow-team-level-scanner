#!/usr/bin/env python3
"""Process Jira data, filter by team, analyze DoR compliance, generate Report.xlsx"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOOL_DIR = "C:/Users/i.klyha/.claude/projects/C--Users-i-klyha-Desktop-Claude-wow-scanner-tool/781d9587-74ad-45a3-934f-7fffed969ea7/tool-results"
OUTPUT_DIR = "C:/Users/i.klyha/Desktop/Claude/wow-scanner-tool/assets/teams/20260612 14-22"

TEAMS = [
    {"name": "Abyss", "pattern": "PE - WAW - Abyss", "project": "MAW", "board_id": "9980", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/MAW/boards/9980", "page_id": "22648881240", "page_link": "https://adgear.atlassian.net/wiki/x/WID6RQU", "dor_status": "success", "dor_source": "direct"},
    {"name": "Radium", "pattern": "AE - WAW - Radium", "project": "AENW", "board_id": "8976", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/AENW/boards/8976", "page_id": "20720615642", "page_link": "https://adgear.atlassian.net/wiki/spaces/ENG/pages/20720615642", "dor_status": "success", "dor_source": "direct"},
    {"name": "Europium", "pattern": "AP - WAW - Europium", "project": "AENW", "board_id": "8979", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/AENW/boards/8979", "page_id": "22431629392", "page_link": "https://adgear.atlassian.net/wiki/spaces/AGILE/pages/22431629392", "dor_status": "success", "dor_source": "direct"},
    {"name": "Copernicium", "pattern": "AE - WAW - Copernicium", "project": "AETVP", "board_id": "9246", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/AETVP/boards/9246", "page_id": "20734247032", "page_link": "https://adgear.atlassian.net/wiki/spaces/ENG/pages/20734247032", "dor_status": "success", "dor_source": "direct"},
    {"name": "Mouflons", "pattern": "AS - WAW - Mouflons", "project": "PEPI", "board_id": "4503", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/PEPI/boards/4503", "page_id": "22381756417", "page_link": "https://adgear.atlassian.net/wiki/spaces/ENG/pages/22381756417", "dor_status": "success", "dor_source": "direct"},
    {"name": "Wolves", "pattern": "AS - WAW - Wolves", "project": "PEPI", "board_id": "4504", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/PEPI/boards/4504", "page_id": "22374940673", "page_link": "https://adgear.atlassian.net/wiki/spaces/ENG/pages/22374940673", "dor_status": "success", "dor_source": "direct"},
    {"name": "Polonium UF", "pattern": "AS - WAW - Polonium UF", "project": "RSW", "board_id": "10403", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/RSW/boards/10403", "page_id": "22606053416", "page_link": "https://adgear.atlassian.net/wiki/spaces/AGILE/pages/22606053416", "dor_status": "success", "dor_source": "linked_page"},
    {"name": "Bigos", "pattern": "AS - WAW - Bigos", "project": "MAW", "board_id": "11439", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/MAW/boards/11439", "page_id": "22695936064", "page_link": "https://adgear.atlassian.net/wiki/spaces/AGILE/pages/22695936064", "dor_status": "success", "dor_source": "linked_page"},
    {"name": "Capybaras", "pattern": "AS - WAW - Capybaras", "project": "RSW", "board_id": "10156", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/RSW/boards/10156", "page_id": "22696132609", "page_link": "https://adgear.atlassian.net/wiki/spaces/AGILE/pages/22696132609", "dor_status": "success", "dor_source": "direct"},
    {"name": "ML Serving Sturgeons", "pattern": "T - WAW - ML Sturgeons", "project": "PEPI", "board_id": "4090", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/PEPI/boards/4090", "page_id": "21732425749", "page_link": "https://adgear.atlassian.net/wiki/spaces/ENG/pages/21732425749", "dor_status": "not_found", "dor_source": None},
    {"name": "ML Platform Pandas", "pattern": "T - WAW - ML Pandas", "project": "ML", "board_id": "10470", "board_url": "https://adgear.atlassian.net/jira/software/projects/ML/boards/10470", "page_id": "21732360213", "page_link": "https://adgear.atlassian.net/wiki/spaces/ENG/pages/21732360213", "dor_status": "success", "dor_source": "linked_page"},
    {"name": "EP Core", "pattern": "T - WAW - EP Core", "project": "EPCW", "board_id": "10972", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/EPCW/boards/10972", "page_id": "22817472513", "page_link": "https://adgear.atlassian.net/wiki/x/AQAHUAU", "dor_status": "success", "dor_source": "direct"},
    {"name": "Zurek", "pattern": "Zurek", "project": "PEA", "board_id": "2881", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/PEA/boards/2881", "page_id": "21748023571", "page_link": "https://adgear.atlassian.net/wiki/spaces/ENG/pages/21748023571", "dor_status": "success", "dor_source": "direct"},
    {"name": "Igni", "pattern": "AP - WAW - Igni", "project": "ASPW", "board_id": "9477", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/ASPW/boards/9477", "page_id": "22435922026", "page_link": "https://adgear.atlassian.net/wiki/spaces/ENG/pages/22435922026", "dor_status": "success", "dor_source": "direct"},
    {"name": "SRE", "pattern": "T - WAW - Embedded SREs SRPOL", "project": "EEEW", "board_id": "10332", "board_url": "https://adgear.atlassian.net/jira/software/c/projects/EEEW/boards/10332", "page_id": "22719529363", "page_link": "https://adgear.atlassian.net/wiki/spaces/AGILE/pages/22719529363", "dor_status": "not_found", "dor_source": None},
]

# --- Load all Jira data ---
all_project_issues = {}
for f in os.listdir(TOOL_DIR):
    if not f.endswith('.txt'):
        continue
    try:
        with open(os.path.join(TOOL_DIR, f), 'r', encoding='utf-8') as fh:
            data = json.load(fh)
        if 'issues' in data and 'nodes' in data['issues']:
            for node in data['issues']['nodes']:
                pk = node.get('fields', {}).get('project', {}).get('key', '')
                if pk:
                    all_project_issues.setdefault(pk, []).append(node)
    except:
        continue

# Add ML issues (inline, customfield_10114 was null)
all_project_issues.setdefault("ML", []).extend([
    {"key": "ML-55", "fields": {"summary": "Implement Test DAG MLOps based on SDK", "issuetype": {"name": "Task"}, "status": {"name": "In Progress"}, "assignee": {"displayName": "Tomasz Teter"}, "priority": {"name": "Medium"}, "created": "2026-05-07", "updated": "2026-05-20", "customfield_10114": None, "description": "Acceptance criteria:\n1. Test DAG fully written using SDK\n2. Test DAG successfully running E2E 2.10.3"}, "webUrl": "https://adgear.atlassian.net/browse/ML-55"},
    {"key": "ML-10", "fields": {"summary": "Support after MWAA upgrade to 2.10.3", "issuetype": {"name": "Task"}, "status": {"name": "In Progress"}, "assignee": {"displayName": "Tomasz Teter"}, "priority": {"name": "Medium"}, "created": "2026-04-20", "updated": "2026-04-23", "customfield_10114": None, "description": "Provide support and handle issues arising from the MWAA upgrade. Acceptance criteria: TBD"}, "webUrl": "https://adgear.atlassian.net/browse/ML-10"},
])

# --- Filter issues by team ---
def get_team_field(issue):
    cf = issue.get('fields', {}).get('customfield_10114')
    if cf and isinstance(cf, dict):
        return cf.get('name', '')
    return None

team_issues_map = {t['name']: [] for t in TEAMS}
for project, issues in all_project_issues.items():
    for issue in issues:
        tf = get_team_field(issue)
        fields = issue.get('fields', {})
        matched = False
        for team in TEAMS:
            if team['project'] == project and tf and tf == team['pattern']:
                team_issues_map[team['name']].append(issue)
                matched = True
                break
        # ML project has null team field - assign to ML Platform Pandas
        if not matched and project == 'ML':
            team_issues_map['ML Platform Pandas'].append(issue)

# --- DoR Compliance Analysis ---
def has_description(issue):
    desc = issue.get('fields', {}).get('description', '')
    if desc is None:
        return False
    if isinstance(desc, dict):  # ADF format
        content = desc.get('content', [])
        return len(content) > 0
    return len(str(desc).strip()) > 10

def has_acceptance_criteria(issue):
    desc = issue.get('fields', {}).get('description', '')
    if desc is None:
        return False
    if isinstance(desc, dict):
        desc = json.dumps(desc)
    desc_lower = str(desc).lower()
    ac_keywords = ['acceptance criteria', 'ac:', 'done when', 'success criteria', 'expected result', 'expected outcome']
    return any(kw in desc_lower for kw in ac_keywords)

def analyze_issue(issue, team_name):
    """Simple DoR check - description + AC presence"""
    has_desc = has_description(issue)
    has_ac = has_acceptance_criteria(issue)

    missing = []
    if not has_desc:
        missing.append("No description provided")
    if not has_ac:
        missing.append("Acceptance criteria not found in description")

    passes = len(missing) == 0
    note = "; ".join(missing) if missing else ""
    return passes, note

# --- Build report data ---
report_data = []
for team in TEAMS:
    issues = team_issues_map[team['name']]
    has_dor = team['dor_status'] == 'success'

    for issue in issues:
        fields = issue.get('fields', {})
        key = issue.get('key', '')
        issue_type = fields.get('issuetype', {}).get('name', '')
        summary = fields.get('summary', '')
        status = fields.get('status', {}).get('name', '')
        assignee = fields.get('assignee', {})
        assignee_name = assignee.get('displayName', 'Unassigned') if assignee else 'Unassigned'
        url = issue.get('webUrl', f"https://adgear.atlassian.net/browse/{key}")

        if has_dor:
            passes, note = analyze_issue(issue, team['name'])
        else:
            passes, note = True, ""  # Skip analysis for teams without DoR

        report_data.append({
            "team": team['name'],
            "issue_key": key,
            "issue_type": issue_type,
            "url": url,
            "title": summary,
            "status": status,
            "assignee": assignee_name,
            "dor_compliance": "Pass" if passes else "Fail",
            "note": note
        })

# --- Save per-team Jira files ---
for team in TEAMS:
    issues = team_issues_map[team['name']]
    kebab = team['name'].lower().replace(' ', '-')

    # JSON
    jira_data = {
        "team": team['name'],
        "boardUrl": team['board_url'],
        "boardId": team['board_id'],
        "projectKey": team['project'],
        "extractedAt": "2026-06-12T14:22:47+02:00",
        "issues": [{
            "key": i.get('key', ''),
            "type": i.get('fields', {}).get('issuetype', {}).get('name', ''),
            "summary": i.get('fields', {}).get('summary', ''),
            "status": i.get('fields', {}).get('status', {}).get('name', ''),
            "assignee": (i.get('fields', {}).get('assignee') or {}).get('displayName', 'Unassigned'),
            "priority": (i.get('fields', {}).get('priority') or {}).get('name', 'None'),
            "url": i.get('webUrl', f"https://adgear.atlassian.net/browse/{i.get('key', '')}"),
            "description": str(i.get('fields', {}).get('description', ''))[:200]
        } for i in issues]
    }
    with open(os.path.join(OUTPUT_DIR, f"{kebab}-jira.json"), 'w') as f:
        json.dump(jira_data, f, indent=2)

    # TXT
    txt_lines = [f"Team: {team['name']}", f"Board: {team['board_url']}", f"Project: {team['project']}", f"Total issues: {len(issues)}", ""]
    for i in issues:
        fields = i.get('fields', {})
        txt_lines.append(f"[{i.get('key','')}] {fields.get('issuetype',{}).get('name','')}: {fields.get('summary','')}")
        txt_lines.append(f"  Status: {fields.get('status',{}).get('name','')} | Assignee: {(fields.get('assignee') or {}).get('displayName','Unassigned')}")
        txt_lines.append(f"  {i.get('webUrl','')}")
        txt_lines.append("")
    with open(os.path.join(OUTPUT_DIR, f"{kebab}-jira.txt"), 'w') as f:
        f.write("\n".join(txt_lines))

# --- Save teams.json ---
teams_json = {
    "metadata": {
        "scan_date": "2026-06-12T14:22:47+02:00",
        "scan_timestamp_cet": "20260612 14-22",
        "source_page": "https://adgear.atlassian.net/wiki/spaces/ENG/pages/19470090380/SRPOL+Teams",
        "team_count": 15,
        "cloudId": "adgear.atlassian.net",
        "scanner_version": "2.0",
        "total_active_issues": len(report_data)
    },
    "teams": [{
        "name": t['name'],
        "page_link": t['page_link'],
        "sprint_board_link": t['board_url'],
        "dor_file": f"{t['name'].lower().replace(' ', '-')}-dor.txt",
        "page_id": t['page_id'],
        "extraction_status": t['dor_status'],
        "extraction_error": None,
        "dor_source": t['dor_source'],
        "jira_issues_count": len(team_issues_map[t['name']]),
        "jira_file": f"{t['name'].lower().replace(' ', '-')}-jira.json"
    } for t in TEAMS]
}
with open(os.path.join(OUTPUT_DIR, "teams.json"), 'w') as f:
    json.dump(teams_json, f, indent=2)

# --- Generate Report.xlsx ---
wb = Workbook()
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
kpi_bold = Font(bold=True, size=11)
green_font = Font(bold=True, size=11, color="006600")
orange_font = Font(bold=True, size=11, color="CC6600")
red_font = Font(bold=True, size=11, color="CC0000")

# KPI calculations
total_teams = len(TEAMS)
teams_with_dor = sum(1 for t in TEAMS if t['dor_status'] == 'success')
teams_with_dor_pct = round(teams_with_dor / total_teams * 100, 1)
total_issues_analyzed = len(report_data)
compliant = sum(1 for r in report_data if r['dor_compliance'] == 'Pass')
compliance_pct = round(compliant / total_issues_analyzed * 100, 1) if total_issues_analyzed > 0 else 0

# Sheet 1: Summary
ws = wb.active
ws.title = "Summary"
ws.cell(1, 1).value = "Teams with DoR:"
ws.cell(1, 1).font = kpi_bold
ws.cell(1, 2).value = f"{teams_with_dor_pct}%"
ws.cell(1, 2).font = green_font if teams_with_dor_pct >= 70 else (orange_font if teams_with_dor_pct >= 40 else red_font)
ws.cell(2, 1).value = "Issues fitting DoR:"
ws.cell(2, 1).font = kpi_bold
ws.cell(2, 2).value = f"{compliance_pct}%"
ws.cell(2, 2).font = green_font if compliance_pct >= 70 else (orange_font if compliance_pct >= 40 else red_font)

# Header row 4
for ci, (h, w) in enumerate([("Team", 25), ("DoR", 10), ("Jira Tasks", 12)], 1):
    c = ws.cell(4, ci)
    c.value = h
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border
    ws.column_dimensions[get_column_letter(ci)].width = w

for ri, t in enumerate(TEAMS, 5):
    has_dor = "Yes" if t['dor_status'] == 'success' else "No"
    ws.cell(ri, 1).value = t['name']
    ws.cell(ri, 1).border = thin_border
    ws.cell(ri, 2).value = has_dor
    ws.cell(ri, 2).fill = pass_fill if has_dor == "Yes" else fail_fill
    ws.cell(ri, 2).border = thin_border
    ws.cell(ri, 2).alignment = Alignment(horizontal='center')
    ws.cell(ri, 3).value = len(team_issues_map[t['name']])
    ws.cell(ri, 3).border = thin_border
    ws.cell(ri, 3).alignment = Alignment(horizontal='center')

ws.freeze_panes = 'A5'

# Sheet 2: DoR Compliance
ws2 = wb.create_sheet("DoR Compliance")
cols = [("Team", 15), ("Issue Key", 12), ("Issue Type", 10), ("URL", 50), ("Title", 40), ("Status", 12), ("Assignee", 15), ("DoR Compliance", 15), ("Note", 60)]
for ci, (h, w) in enumerate(cols, 1):
    c = ws2.cell(1, ci)
    c.value = h
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border
    ws2.column_dimensions[get_column_letter(ci)].width = w

for ri, row in enumerate(report_data, 2):
    ws2.cell(ri, 1).value = row['team']
    ws2.cell(ri, 2).value = row['issue_key']
    ws2.cell(ri, 3).value = row['issue_type']
    ws2.cell(ri, 4).value = row['url']
    ws2.cell(ri, 4).hyperlink = row['url']
    ws2.cell(ri, 4).style = "Hyperlink"
    ws2.cell(ri, 5).value = row['title']
    ws2.cell(ri, 6).value = row['status']
    ws2.cell(ri, 7).value = row['assignee']
    ws2.cell(ri, 8).value = row['dor_compliance']
    ws2.cell(ri, 8).fill = pass_fill if row['dor_compliance'] == 'Pass' else fail_fill
    ws2.cell(ri, 9).value = row['note'] if row['note'] else None
    for col in range(1, 10):
        ws2.cell(ri, col).border = thin_border
        ws2.cell(ri, col).alignment = Alignment(vertical='top', wrap_text=(col in [5, 9]))

ws2.freeze_panes = 'A2'
wb.save(os.path.join(OUTPUT_DIR, "Report.xlsx"))

# --- Summary ---
print(f"[SUCCESS] Report.xlsx generated")
print(f"Teams with DoR: {teams_with_dor_pct}%")
print(f"Issues fitting DoR: {compliance_pct}%")
print(f"Total issues: {total_issues_analyzed}")
print(f"Pass: {compliant}, Fail: {total_issues_analyzed - compliant}")
print(f"\nPer-team issue counts:")
for t in TEAMS:
    print(f"  {t['name']}: {len(team_issues_map[t['name']])}")
