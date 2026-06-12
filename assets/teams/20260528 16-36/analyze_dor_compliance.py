#!/usr/bin/env python3
"""
DoR Compliance Analysis - Step 11
Analyzes Jira issues against team DoR criteria using LLM-based semantic analysis.
Generates Excel report with exactly 9 columns as specified.
"""

import json
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARNING: openpyxl not installed. Will generate CSV instead.")

WORK_DIR = r"C:\Users\i.klyha\Desktop\Claude\wow-scanner-tool\assets\teams\20260528 16-36"

# Analyzable teams (extraction_status="success" AND jiraIssueCount>0)
ANALYZABLE_TEAMS = [
    ("PE-WAW-Abyss", "pe-waw-abyss"),
    ("Radium", "radium"),
    ("Europium", "europium"),
    ("Copernicium", "copernicium"),
    ("Capybaras", "capybaras"),
    ("EP Core", "ep-core"),
    ("Igni", "igni"),
]

def load_dor(team_slug):
    """Load DoR criteria from file."""
    dor_path = os.path.join(WORK_DIR, f'{team_slug}-dor.txt')
    if not os.path.exists(dor_path):
        return None

    with open(dor_path, 'r', encoding='utf-8') as f:
        return f.read()

def load_jira_issues(team_slug):
    """Load Jira issues from JSON file."""
    jira_path = os.path.join(WORK_DIR, f'{team_slug}-jira.json')
    if not os.path.exists(jira_path):
        return []

    with open(jira_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    return data.get('issues', [])

def extract_issue_data(issue):
    """Extract relevant fields from Jira issue object."""
    fields = issue.get('fields', {})

    return {
        'key': issue.get('key', 'N/A'),
        'type': fields.get('issuetype', {}).get('name', 'N/A'),
        'status': fields.get('status', {}).get('name', 'N/A'),
        'summary': fields.get('summary', 'N/A'),
        'description': fields.get('description', ''),
        'assignee': fields.get('assignee', {}).get('displayName', 'Unassigned') if fields.get('assignee') else 'Unassigned',
        'priority': fields.get('priority', {}).get('name', 'N/A'),
        'url': issue.get('webUrl', f"https://adgear.atlassian.net/browse/{issue.get('key', '')}")
    }

def prepare_analysis_data():
    """Prepare all data for LLM analysis."""
    analysis_data = {}

    for team_name, team_slug in ANALYZABLE_TEAMS:
        print(f"Loading {team_name}...")

        dor_content = load_dor(team_slug)
        if not dor_content:
            print(f"  WARNING: No DoR found for {team_name}")
            continue

        issues = load_jira_issues(team_slug)
        if not issues:
            print(f"  WARNING: No issues found for {team_name}")
            continue

        # Extract issue data
        issues_data = [extract_issue_data(issue) for issue in issues]

        analysis_data[team_name] = {
            'slug': team_slug,
            'dor': dor_content,
            'issues': issues_data,
            'issue_count': len(issues_data)
        }

        print(f"  Loaded {len(issues_data)} issues")

    return analysis_data

def save_analysis_input(analysis_data):
    """Save prepared data for LLM analysis (for debugging)."""
    output_path = os.path.join(WORK_DIR, 'dor_analysis_input.json')

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(analysis_data, f, indent=2, ensure_ascii=False)

    print(f"\nAnalysis input data saved to: dor_analysis_input.json")

    # Also create a summary for LLM prompt
    summary = {
        'teams_count': len(analysis_data),
        'total_issues': sum(team['issue_count'] for team in analysis_data.values()),
        'teams': [
            {
                'name': team_name,
                'issue_count': team_data['issue_count']
            }
            for team_name, team_data in analysis_data.items()
        ]
    }

    return summary

def generate_csv_report(report_data):
    """Generate CSV report if openpyxl not available."""
    csv_path = os.path.join(WORK_DIR, 'Report.csv')

    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        # Header
        f.write('Team,Issue Key,Issue Type,Status,Title,URL,Assignee,DoR Compliance,Feedback\n')

        # Data rows
        for row in report_data:
            # Escape commas and quotes in text fields
            def escape_csv(text):
                if ',' in text or '"' in text or '\n' in text:
                    return '"' + text.replace('"', '""') + '"'
                return text

            f.write(','.join([
                escape_csv(row['team']),
                escape_csv(row['issue_key']),
                escape_csv(row['issue_type']),
                escape_csv(row['status']),
                escape_csv(row['title']),
                escape_csv(row['url']),
                escape_csv(row['assignee']),
                escape_csv(row['dor_compliance']),
                escape_csv(row['feedback'])
            ]) + '\n')

    print(f"\nCSV report generated: Report.csv")
    return csv_path

def generate_excel_report(report_data):
    """Generate Excel report with EXACT schema as specified."""

    if not OPENPYXL_AVAILABLE:
        return generate_csv_report(report_data)

    print("\nGenerating Excel report...")

    wb = Workbook()
    ws = wb.active
    ws.title = "DoR Compliance"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    compliant_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    non_compliant_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # MANDATORY SCHEMA - EXACTLY 9 COLUMNS
    headers = [
        "Team",
        "Issue Key",
        "Issue Type",
        "Status",
        "Title",
        "URL",
        "Assignee",
        "DoR Compliance",
        "Feedback"
    ]

    # Write header
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Set column widths (A:15, B:12, C:10, D:12, E:40, F:50, G:15, H:15, I:60)
    column_widths = [15, 12, 10, 12, 40, 50, 15, 15, 60]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze row 1
    ws.freeze_panes = 'A2'

    # Write data rows
    for row_idx, row_data in enumerate(report_data, 2):
        # Column 1: Team
        ws.cell(row=row_idx, column=1).value = row_data['team']

        # Column 2: Issue Key
        ws.cell(row=row_idx, column=2).value = row_data['issue_key']

        # Column 3: Issue Type
        ws.cell(row=row_idx, column=3).value = row_data['issue_type']

        # Column 4: Status
        ws.cell(row=row_idx, column=4).value = row_data['status']

        # Column 5: Title (wrap text)
        cell = ws.cell(row=row_idx, column=5)
        cell.value = row_data['title']
        cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Column 6: URL (hyperlink)
        cell = ws.cell(row=row_idx, column=6)
        cell.value = row_data['url']
        cell.hyperlink = row_data['url']
        cell.style = "Hyperlink"

        # Column 7: Assignee
        ws.cell(row=row_idx, column=7).value = row_data['assignee']

        # Column 8: DoR Compliance ("Yes" or "No" only)
        cell = ws.cell(row=row_idx, column=8)
        cell.value = row_data['dor_compliance']

        # Apply conditional fill
        if row_data['dor_compliance'] == "Yes":
            cell.fill = compliant_fill
        else:
            cell.fill = non_compliant_fill

        # Column 9: Feedback (wrap text)
        cell = ws.cell(row=row_idx, column=9)
        cell.value = row_data['feedback']
        cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Apply borders and vertical top alignment to all cells
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if col not in [5, 9]:  # Title and Feedback already have alignment set
                cell.alignment = Alignment(vertical='top')

    # Save workbook
    output_path = os.path.join(WORK_DIR, 'Report.xlsx')
    wb.save(output_path)

    print(f"Excel report generated: Report.xlsx")
    return output_path

def generate_summary(analysis_data, report_data):
    """Generate analysis summary markdown."""

    total_issues = len(report_data)
    compliant_count = sum(1 for row in report_data if row['dor_compliance'] == 'Yes')
    non_compliant_count = total_issues - compliant_count

    compliance_rate = (compliant_count / total_issues * 100) if total_issues > 0 else 0

    # Team breakdown
    team_stats = {}
    for row in report_data:
        team = row['team']
        if team not in team_stats:
            team_stats[team] = {'total': 0, 'compliant': 0}
        team_stats[team]['total'] += 1
        if row['dor_compliance'] == 'Yes':
            team_stats[team]['compliant'] += 1

    # Most common DoR gaps
    gap_counts = {}
    for row in report_data:
        if row['dor_compliance'] == 'No' and row['feedback']:
            # Count each gap mentioned
            gaps = [g.strip() for g in row['feedback'].split(',') if g.strip()]
            for gap in gaps[:3]:  # Top 3 per issue
                gap_short = gap[:50]
                gap_counts[gap_short] = gap_counts.get(gap_short, 0) + 1

    top_gaps = sorted(gap_counts.items(), key=lambda x: x[1], reverse=True)[:10]

    # Generate markdown
    summary_lines = [
        "# DoR Compliance Analysis Summary",
        "",
        f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## Overall Statistics",
        "",
        f"- **Total Issues Analyzed:** {total_issues}",
        f"- **Compliant Issues:** {compliant_count} ({compliance_rate:.1f}%)",
        f"- **Non-Compliant Issues:** {non_compliant_count} ({100-compliance_rate:.1f}%)",
        f"- **Teams Analyzed:** {len(team_stats)}",
        "",
        "## Team Breakdown",
        "",
        "| Team | Total Issues | Compliant | Non-Compliant | Compliance Rate |",
        "|------|--------------|-----------|---------------|-----------------|"
    ]

    for team in sorted(team_stats.keys()):
        stats = team_stats[team]
        rate = (stats['compliant'] / stats['total'] * 100) if stats['total'] > 0 else 0
        summary_lines.append(
            f"| {team} | {stats['total']} | {stats['compliant']} | "
            f"{stats['total'] - stats['compliant']} | {rate:.1f}% |"
        )

    summary_lines.extend([
        "",
        "## Most Common DoR Gaps",
        ""
    ])

    if top_gaps:
        for idx, (gap, count) in enumerate(top_gaps, 1):
            summary_lines.append(f"{idx}. {gap} ({count} occurrences)")
    else:
        summary_lines.append("No specific gaps identified.")

    summary_lines.extend([
        "",
        "## Teams Skipped",
        ""
    ])

    all_teams = [name for name, _ in ANALYZABLE_TEAMS]
    analyzed_teams = list(team_stats.keys())
    skipped_teams = [t for t in all_teams if t not in analyzed_teams]

    if skipped_teams:
        for team in skipped_teams:
            summary_lines.append(f"- {team} (No DoR or no issues)")
    else:
        summary_lines.append("All teams with DoR and active issues were analyzed.")

    summary_lines.extend([
        "",
        "## Analysis Method",
        "",
        "- DoR criteria extracted from team-specific DoR documentation",
        "- Each issue analyzed for compliance with team's DoR",
        "- Issues marked as 'Yes' (compliant) or 'No' (non-compliant)",
        "- Feedback provided for non-compliant issues detailing missing criteria",
        "",
        "## Output Files",
        "",
        "- `Report.xlsx` - Excel report with 9-column schema",
        "- `DOR_ANALYSIS_SUMMARY.md` - This summary document",
        "- `dor_analysis_input.json` - Raw analysis data (for debugging)",
        ""
    ])

    summary_text = '\n'.join(summary_lines)

    # Save summary
    summary_path = os.path.join(WORK_DIR, 'DOR_ANALYSIS_SUMMARY.md')
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write(summary_text)

    print(f"Summary generated: DOR_ANALYSIS_SUMMARY.md")

    return summary_text

def update_teams_json(analysis_summary):
    """Update teams.json with dor_analysis metadata."""
    teams_json_path = os.path.join(WORK_DIR, 'teams.json')

    with open(teams_json_path, 'r', encoding='utf-8') as f:
        teams_data = json.load(f)

    # Add dor_analysis metadata
    if 'dor_analysis' not in teams_data:
        teams_data['dor_analysis'] = {}

    teams_data['dor_analysis'] = {
        'performed': True,
        'timestamp': datetime.now().isoformat(),
        'teams_analyzed': analysis_summary['teams_analyzed'],
        'issues_analyzed': analysis_summary['total_issues'],
        'compliance_rate': analysis_summary['compliance_rate'],
        'compliant_count': analysis_summary['compliant_count'],
        'non_compliant_count': analysis_summary['non_compliant_count']
    }

    # Save updated teams.json
    with open(teams_json_path, 'w', encoding='utf-8') as f:
        json.dump(teams_data, f, indent=2, ensure_ascii=False)

    print(f"Updated teams.json with dor_analysis metadata")

def main():
    """Main execution flow."""
    print("=" * 60)
    print("DoR Compliance Analysis - Step 11")
    print("=" * 60)
    print()

    # Step 11.1: Identify teams to analyze
    print("Step 11.1: Identifying analyzable teams...")
    print(f"Found {len(ANALYZABLE_TEAMS)} teams to analyze:")
    for team_name, _ in ANALYZABLE_TEAMS:
        print(f"  - {team_name}")
    print()

    # Step 11.2: Prepare data (descriptions already in JSON)
    print("Step 11.2: Loading issue data from JSON files...")
    analysis_data = prepare_analysis_data()
    summary = save_analysis_input(analysis_data)
    print()

    # Step 11.3: LLM Analysis placeholder
    print("Step 11.3: Ready for LLM-based DoR compliance analysis")
    print(f"  Teams to analyze: {summary['teams_count']}")
    print(f"  Total issues: {summary['total_issues']}")
    print()
    print("NEXT STEP: Run LLM analysis on prepared data")
    print("  Input file: dor_analysis_input.json")
    print()
    print("This script has prepared all data.")
    print("You now need to:")
    print("  1. Load dor_analysis_input.json")
    print("  2. For each team, batch all issues into one LLM call")
    print("  3. Get DoR compliance results (JSON)")
    print("  4. Call generate_excel_report() with results")
    print("  5. Call generate_summary() with results")
    print("  6. Call update_teams_json() with summary stats")

if __name__ == '__main__':
    main()
