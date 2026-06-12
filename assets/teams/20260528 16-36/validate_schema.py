#!/usr/bin/env python3
"""
Schema Validator - Verifies Report.xlsx has correct 9-column schema
"""

import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed")
    print("Install with: py -m pip install openpyxl")
    sys.exit(1)

WORK_DIR = r"C:\Users\i.klyha\Desktop\Claude\wow-scanner-tool\assets\teams\20260528 16-36"
REPORT_PATH = os.path.join(WORK_DIR, 'Report.xlsx')

# Expected schema
EXPECTED_HEADERS = [
    "Team",
    "Issue Key",
    "Issue Type",
    "Status",
    "Title",
    "URL",
    "Assignee",
    "DoR Compliance",
    "Feedback"
]

def validate_schema():
    """Validate Report.xlsx schema."""
    if not os.path.exists(REPORT_PATH):
        print(f"ERROR: Report.xlsx not found at {REPORT_PATH}")
        return False

    print("Loading Report.xlsx...")
    wb = load_workbook(REPORT_PATH)
    ws = wb.active

    print(f"Sheet name: {ws.title}")

    # Check sheet name
    if ws.title != "DoR Compliance":
        print(f"  WARNING: Sheet name is '{ws.title}', expected 'DoR Compliance'")

    # Check headers
    print("\nValidating headers...")
    headers = []
    for col in range(1, 20):  # Check up to 20 columns
        value = ws.cell(row=1, column=col).value
        if value:
            headers.append(value)
        else:
            break

    print(f"  Found {len(headers)} columns")

    if len(headers) != 9:
        print(f"  ERROR: Expected 9 columns, found {len(headers)}")
        print(f"  Headers: {headers}")
        return False

    # Check each header
    errors = []
    for idx, (expected, actual) in enumerate(zip(EXPECTED_HEADERS, headers), 1):
        if expected != actual:
            errors.append(f"  Column {idx}: Expected '{expected}', got '{actual}'")
        else:
            print(f"  Column {idx}: {actual} ✓")

    if errors:
        print("\nERRORS:")
        for error in errors:
            print(error)
        return False

    # Check DoR Compliance values
    print("\nValidating DoR Compliance column (H)...")
    invalid_values = []
    row = 2
    while True:
        value = ws.cell(row=row, column=8).value
        if value is None:
            break

        if value not in ["Yes", "No"]:
            invalid_values.append(f"  Row {row}: '{value}' (should be 'Yes' or 'No')")

        row += 1

    if invalid_values:
        print(f"  ERROR: Found {len(invalid_values)} invalid values:")
        for inv in invalid_values[:10]:  # Show first 10
            print(inv)
        return False
    else:
        print(f"  Validated {row-2} rows ✓")

    # Check Feedback column
    print("\nValidating Feedback column (I)...")
    yes_with_feedback = 0
    no_without_feedback = 0

    for r in range(2, row):
        compliance = ws.cell(row=r, column=8).value
        feedback = ws.cell(row=r, column=9).value or ""

        if compliance == "Yes" and feedback.strip():
            yes_with_feedback += 1
        elif compliance == "No" and not feedback.strip():
            no_without_feedback += 1

    if yes_with_feedback > 0:
        print(f"  WARNING: {yes_with_feedback} 'Yes' rows have feedback (should be empty)")

    if no_without_feedback > 0:
        print(f"  WARNING: {no_without_feedback} 'No' rows lack feedback (should explain why)")

    if yes_with_feedback == 0 and no_without_feedback == 0:
        print(f"  Feedback consistency ✓")

    print("\n" + "=" * 60)
    print("SCHEMA VALIDATION PASSED ✓")
    print("=" * 60)
    return True

if __name__ == '__main__':
    try:
        success = validate_schema()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\nFATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
