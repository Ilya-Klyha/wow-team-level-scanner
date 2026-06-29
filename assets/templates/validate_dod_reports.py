#!/usr/bin/env python3
"""
DoD Scanner - Report Validation Script
Validates that generated Report-DoD.xlsx and Report-DoD.html conform to expected schema.

Usage:
    python3 validate_dod_reports.py <output_dir>
"""
import json
import os
import sys
from pathlib import Path

errors = []
warnings = []


def error(msg):
    errors.append(msg)
    print(f"  [ERROR] {msg}")


def warn(msg):
    warnings.append(msg)
    print(f"  [WARN] {msg}")


def ok(msg):
    print(f"  [OK] {msg}")


def validate_excel(output_dir):
    """Validate Report-DoD.xlsx structure."""
    print("\n=== Validating Report-DoD.xlsx ===")
    xlsx_path = os.path.join(output_dir, "Report-DoD.xlsx")

    if not os.path.exists(xlsx_path):
        error("Report-DoD.xlsx does not exist")
        return

    file_size = os.path.getsize(xlsx_path)
    if file_size < 3000:
        error(f"Report-DoD.xlsx too small ({file_size} bytes), likely corrupt")
        return

    try:
        from openpyxl import load_workbook
    except ImportError:
        warn("openpyxl not installed, cannot validate Excel structure")
        return

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        error(f"Cannot open Report-DoD.xlsx: {e}")
        return

    # Check sheet names
    expected_sheets = ["Summary", "DoD quality"]
    if wb.sheetnames != expected_sheets:
        error(f"Sheet names are {wb.sheetnames}, expected {expected_sheets}")
    else:
        ok(f"Sheets: {wb.sheetnames}")

    # Validate Summary sheet
    ws = wb["Summary"] if "Summary" in wb.sheetnames else wb[wb.sheetnames[0]]

    # Row 1: KPI
    if ws["A1"].value != "% Teams with DoD":
        error(f"A1 value is '{ws['A1'].value}', expected '% Teams with DoD'")
    else:
        ok("KPI label correct")

    kpi_value = ws["C1"].value
    if kpi_value is None or "%" not in str(kpi_value):
        error(f"C1 KPI value is '{kpi_value}', expected percentage string")
    else:
        ok(f"KPI value: {kpi_value}")

    # Row 3: Headers
    if ws.cell(3, 1).value != "Team":
        error(f"Header A3 is '{ws.cell(3, 1).value}', expected 'Team'")
    if ws.cell(3, 2).value != "DoD":
        error(f"Header B3 is '{ws.cell(3, 2).value}', expected 'DoD'")
    else:
        ok("Headers correct: Team | DoD")

    # Data rows (4-18 = 15 teams)
    data_rows = 0
    for row in range(4, ws.max_row + 1):
        if ws.cell(row, 1).value:
            data_rows += 1
            dod_val = ws.cell(row, 2).value
            if dod_val not in ("Yes", "No"):
                error(f"Row {row} DoD value is '{dod_val}', expected 'Yes' or 'No'")

    if data_rows != 15:
        error(f"Found {data_rows} data rows, expected 15")
    else:
        ok("15 team data rows present")

    # Check freeze panes
    if ws.freeze_panes != "A4":
        warn(f"Freeze panes is '{ws.freeze_panes}', expected 'A4'")
    else:
        ok("Freeze panes: A4")

    # Validate DoD quality sheet
    if "DoD quality" in wb.sheetnames:
        ws_q = wb["DoD quality"]

        if ws_q.cell(1, 1).value != "DoD quality lvl":
            error(f"Quality sheet A1 is '{ws_q.cell(1, 1).value}', expected 'DoD quality lvl'")
        else:
            ok("Quality KPI label correct")

        if ws_q.cell(3, 1).value != "Team":
            error(f"Quality header A3 is '{ws_q.cell(3, 1).value}', expected 'Team'")
        if ws_q.cell(3, 2).value != "Quality":
            error(f"Quality header B3 is '{ws_q.cell(3, 2).value}', expected 'Quality'")
        if ws_q.cell(3, 3).value != "Note":
            error(f"Quality header C3 is '{ws_q.cell(3, 3).value}', expected 'Note'")
        else:
            ok("Quality headers correct: Team | Quality | Note")

        if ws_q.freeze_panes != "A4":
            warn(f"Quality freeze panes is '{ws_q.freeze_panes}', expected 'A4'")


def validate_html(output_dir):
    """Validate Report-DoD.html structure."""
    print("\n=== Validating Report-DoD.html ===")
    html_path = os.path.join(output_dir, "Report-DoD.html")

    if not os.path.exists(html_path):
        error("Report-DoD.html does not exist")
        return

    file_size = os.path.getsize(html_path)
    if file_size < 5000:
        error(f"Report-DoD.html too small ({file_size} bytes)")
        return
    ok(f"File size: {file_size} bytes")

    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()

    # Check essential elements
    checks = [
        ("<!DOCTYPE html>", "HTML5 doctype"),
        ("DoD Scanner Report", "Page title"),
        ("linear-gradient(135deg, #1a1a2e", "Header gradient"),
        ("Teams with DoD Defined", "KPI 1 label"),
        ("DoD Quality Level", "KPI 2 label"),
        ("Team Overview", "Team Overview section"),
        ("DoD Quality Assessment", "Quality Assessment section"),
        (".kpi-card", "KPI card CSS"),
        (".badge-yes", "Badge yes CSS"),
        (".badge-no", "Badge no CSS"),
        (".bar-container", "Progress bar CSS"),
        ("@media (max-width: 768px)", "Responsive CSS"),
        ("<footer>", "Footer element"),
        ("wow-dod-scanner", "Scanner attribution"),
    ]

    for pattern, description in checks:
        if pattern in html:
            ok(description)
        else:
            error(f"Missing: {description} (pattern: '{pattern}')")

    # Check it does NOT have DoR-specific sections
    if "Failed Issues" in html:
        error("HTML contains 'Failed Issues' section (DoD should not have this)")

    # Check self-contained (no external dependencies)
    if '<link rel="stylesheet"' in html:
        error("HTML has external CSS link (should be self-contained)")
    if '<script src=' in html:
        error("HTML has external script (should be self-contained)")

    # Count table rows to verify team data
    team_rows = html.count('class="team-name"')
    if team_rows < 15:
        warn(f"Found {team_rows} team-name cells, expected at least 15")
    else:
        ok(f"Team rows: {team_rows}")


def validate_data_files(output_dir):
    """Validate input JSON data files."""
    print("\n=== Validating Data Files ===")

    # summary_data.json
    summary_path = os.path.join(output_dir, "summary_data.json")
    if not os.path.exists(summary_path):
        error("summary_data.json does not exist")
    else:
        with open(summary_path, 'r', encoding='utf-8') as f:
            summary = json.load(f)
        if len(summary) != 15:
            error(f"summary_data.json has {len(summary)} entries, expected 15")
        else:
            ok("summary_data.json: 15 entries")
        for t in summary:
            if "team" not in t or "dod" not in t:
                error(f"summary_data.json entry missing required fields: {t}")
                break

    # quality_data.json
    quality_path = os.path.join(output_dir, "quality_data.json")
    if not os.path.exists(quality_path):
        error("quality_data.json does not exist")
    else:
        with open(quality_path, 'r', encoding='utf-8') as f:
            quality = json.load(f)
        ok(f"quality_data.json: {len(quality)} entries")
        for q in quality:
            required = ["team", "overall", "note"]
            for field in required:
                if field not in q:
                    error(f"quality_data.json entry for '{q.get('team', '?')}' missing '{field}'")
                    break
            if not (0 <= q.get("overall", -1) <= 100):
                error(f"quality_data.json: '{q.get('team')}' overall={q.get('overall')} out of range")


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 validate_dod_reports.py <output_dir>")
        sys.exit(1)

    output_dir = sys.argv[1]
    print(f"Validating DoD reports in: {output_dir}")

    validate_data_files(output_dir)
    validate_excel(output_dir)
    validate_html(output_dir)

    print(f"\n=== Validation Summary ===")
    print(f"  Errors: {len(errors)}")
    print(f"  Warnings: {len(warnings)}")

    if errors:
        print(f"\n  VALIDATION FAILED ({len(errors)} errors)")
        sys.exit(1)
    else:
        print(f"\n  VALIDATION PASSED")
        sys.exit(0)


if __name__ == "__main__":
    main()
