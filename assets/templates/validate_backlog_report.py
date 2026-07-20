#!/usr/bin/env python3
"""
Backlog Readiness Scanner - Report Validator
Validates Report-backlog.xlsx, Report-backlog.html, and JSON data files.

Usage:
    python3 validate_backlog_report.py <output_dir>

This script checks schema compliance and cross-file consistency.
"""
import json
import sys
import os
import re

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def validate(output_dir):
    """Run all validation checks. Returns (errors, warnings) lists."""
    errors = []
    warnings = []

    # ============================================================
    # 1. Required files exist
    # ============================================================

    required_files = [
        "Report-backlog.xlsx",
        "Report-backlog.html",
        "backlog_data.json",
        "summary_data.json",
    ]

    for fname in required_files:
        fpath = os.path.join(output_dir, fname)
        if not os.path.exists(fpath):
            errors.append(f"MISSING FILE: {fname}")

    if errors:
        print("[FATAL] Required files missing. Cannot continue validation.")
        return errors, warnings

    # ============================================================
    # 2. backlog_data.json schema
    # ============================================================

    with open(os.path.join(output_dir, "backlog_data.json"), 'r', encoding='utf-8') as f:
        backlog_data = json.load(f)

    if not isinstance(backlog_data, list):
        errors.append("backlog_data.json: root is not an array")
    else:
        required_keys = {"team", "issue_key", "issue_type", "url", "title", "status", "assignee", "story_points"}
        valid_types = {"Story", "Bug", "Task"}
        issue_key_pattern = re.compile(r'^[A-Z]+-\d+$')
        url_prefix = "https://adgear.atlassian.net/browse/"
        seen_keys = set()

        for idx, entry in enumerate(backlog_data):
            if not isinstance(entry, dict):
                errors.append(f"backlog_data.json[{idx}]: entry is not an object")
                continue

            entry_keys = set(entry.keys())
            if entry_keys != required_keys:
                missing = required_keys - entry_keys
                extra = entry_keys - required_keys
                if missing:
                    errors.append(f"backlog_data.json[{idx}]: missing keys: {missing}")
                if extra:
                    warnings.append(f"backlog_data.json[{idx}]: extra keys: {extra}")

            # issue_key format
            ik = entry.get("issue_key", "")
            if not issue_key_pattern.match(str(ik)):
                errors.append(f"backlog_data.json[{idx}]: invalid issue_key format: '{ik}'")

            # duplicate check
            if ik in seen_keys:
                errors.append(f"backlog_data.json[{idx}]: duplicate issue_key: '{ik}'")
            seen_keys.add(ik)

            # url format
            url = entry.get("url", "")
            if not str(url).startswith(url_prefix):
                errors.append(f"backlog_data.json[{idx}]: url does not start with '{url_prefix}': '{url}'")

            # issue_type
            it = entry.get("issue_type", "")
            if it not in valid_types:
                errors.append(f"backlog_data.json[{idx}]: invalid issue_type: '{it}' (expected: {valid_types})")

            # story_points
            sp = entry.get("story_points")
            if sp is not None:
                if not isinstance(sp, (int, float)):
                    errors.append(f"backlog_data.json[{idx}]: story_points must be number or null, got: {type(sp).__name__}")
                elif sp < 0:
                    errors.append(f"backlog_data.json[{idx}]: story_points is negative: {sp}")

            # assignee
            assignee = entry.get("assignee")
            if not assignee or not isinstance(assignee, str) or assignee.strip() == "":
                errors.append(f"backlog_data.json[{idx}]: assignee is empty or not a string")

    # ============================================================
    # 3. summary_data.json schema
    # ============================================================

    with open(os.path.join(output_dir, "summary_data.json"), 'r', encoding='utf-8') as f:
        summary_data = json.load(f)

    if not isinstance(summary_data, list):
        errors.append("summary_data.json: root is not an array")
    else:
        if len(summary_data) < 15:
            errors.append(f"summary_data.json: expected at least 15 entries, got {len(summary_data)}")

        rate_pattern = re.compile(r'^\d+%$')

        for idx, entry in enumerate(summary_data):
            if not isinstance(entry, dict):
                errors.append(f"summary_data.json[{idx}]: entry is not an object")
                continue

            for key in ["team", "backlog_total", "estimated", "unestimated", "estimation_rate"]:
                if key not in entry:
                    errors.append(f"summary_data.json[{idx}]: missing key: '{key}'")

            bt = entry.get("backlog_total", 0)
            est = entry.get("estimated", 0)
            unest = entry.get("unestimated", 0)

            if isinstance(bt, int) and isinstance(est, int) and isinstance(unest, int):
                if bt != est + unest:
                    errors.append(
                        f"summary_data.json[{idx}] ({entry.get('team', '?')}): "
                        f"backlog_total ({bt}) != estimated ({est}) + unestimated ({unest})"
                    )

            rate = entry.get("estimation_rate", "")
            if rate != "-" and not rate_pattern.match(str(rate)):
                errors.append(f"summary_data.json[{idx}]: invalid estimation_rate: '{rate}' (expected 'X%' or '-')")

    # ============================================================
    # 4. Excel structure
    # ============================================================

    if HAS_OPENPYXL:
        xlsx_path = os.path.join(output_dir, "Report-backlog.xlsx")
        wb = load_workbook(xlsx_path, read_only=True)

        allowed_sheets = {"Backlog Readiness", "Velocity & Runway"}
        if not set(wb.sheetnames).issubset(allowed_sheets):
            unexpected = set(wb.sheetnames) - allowed_sheets
            errors.append(f"Excel: unexpected sheets: {unexpected}")
        if "Backlog Readiness" not in wb.sheetnames:
            errors.append(f"Excel: sheet 'Backlog Readiness' not found. Sheets: {wb.sheetnames}")
        else:
            ws = wb["Backlog Readiness"]
            expected_headers = ["Team", "Issue Key", "Issue Type", "URL", "Title", "Status", "Assignee", "Story Points"]
            actual_headers = [ws.cell(row=4, column=c).value for c in range(1, 9)]

            if actual_headers != expected_headers:
                errors.append(f"Excel row 4 headers mismatch.\n  Expected: {expected_headers}\n  Actual:   {actual_headers}")

            # Check column count
            max_col = ws.max_column
            if max_col != 8:
                warnings.append(f"Excel: expected 8 columns, detected max_column={max_col}")

            # Spot-check Story Points column (column H) for text values
            for row in range(5, min(ws.max_row + 1, 25)):
                val = ws.cell(row=row, column=8).value
                if val is not None and isinstance(val, str):
                    errors.append(f"Excel row {row} col H: Story Points contains text '{val}' (expected number or empty)")

        if "Velocity & Runway" in wb.sheetnames:
            ws2 = wb["Velocity & Runway"]
            expected_runway_headers = ["Team", "Sprint -3 (SP)", "Sprint -2 (SP)", "Sprint -1 (SP)", "Avg Velocity", "Ready SP (DoR>=75%)", "Sprint Runway", "Status"]
            actual_runway_headers = [ws2.cell(row=4, column=c).value for c in range(1, 9)]
            if actual_runway_headers != expected_runway_headers:
                errors.append(f"Excel 'Velocity & Runway' row 4 headers mismatch.\n  Expected: {expected_runway_headers}\n  Actual:   {actual_runway_headers}")

            # Verify Status column values
            allowed_statuses = {"Healthy", "Attention", "Critical", "No Data", "No DoR", None}
            for row in range(5, min(ws2.max_row + 1, 25)):
                status_val = ws2.cell(row=row, column=8).value
                if status_val not in allowed_statuses:
                    errors.append(f"Excel 'Velocity & Runway' row {row}: invalid Status '{status_val}'")

        wb.close()
    else:
        warnings.append("openpyxl not installed - skipping Excel structure validation")

    # ============================================================
    # 5. HTML structure
    # ============================================================

    html_path = os.path.join(output_dir, "Report-backlog.html")
    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()

    if "<title>" not in html or "Backlog" not in html:
        errors.append("HTML: missing <title> or 'Backlog' not in title")

    kpi_count = html.count("kpi-card")
    if kpi_count < 3:
        errors.append(f"HTML: expected 3 KPI cards, found {kpi_count}")

    section_count = html.count('<div class="section">')
    if section_count < 2:
        errors.append(f"HTML: expected at least 2 sections, found {section_count}")

    # Count team rows in summary table (15 expected)
    team_row_count = html.count('class="team-name"')
    # Each team appears in summary + possibly in unestimated table
    # At minimum, summary should have 15 team-name cells
    if team_row_count < 15:
        warnings.append(f"HTML: expected at least 15 team-name cells, found {team_row_count}")

    # ============================================================
    # 6. Cross-file consistency
    # ============================================================

    if isinstance(backlog_data, list) and isinstance(summary_data, list):
        # Count from backlog_data should match sum of summary backlog_total
        backlog_count = len(backlog_data)
        summary_total = sum(t.get("backlog_total", 0) for t in summary_data)

        if backlog_count != summary_total:
            errors.append(
                f"Cross-file: backlog_data.json has {backlog_count} entries, "
                f"but summary_data.json sums to {summary_total}"
            )

        # Teams in backlog_data must be subset of teams in summary_data
        summary_teams = set(t.get("team", "") for t in summary_data)
        backlog_teams = set(d.get("team", "") for d in backlog_data)
        extra_teams = backlog_teams - summary_teams

        if extra_teams:
            errors.append(f"Cross-file: teams in backlog_data not in summary_data: {extra_teams}")

    # ============================================================
    # 7. Velocity & Runway consistency (conditional)
    # ============================================================

    if isinstance(summary_data, list):
        has_runway = any(t.get("sprint_runway") is not None for t in summary_data)
        if has_runway:
            allowed_runway_statuses = {"Healthy", "Attention", "Critical", "No Data", "No DoR"}
            for idx, entry in enumerate(summary_data):
                runway = entry.get("sprint_runway")
                status = entry.get("runway_status")

                if status and status not in allowed_runway_statuses:
                    errors.append(f"summary_data.json[{idx}]: invalid runway_status '{status}'")

                # Check runway formula consistency
                if isinstance(runway, (int, float)):
                    vel = entry.get("velocity_avg")
                    ready_sp = entry.get("dor_ready_sp")
                    if vel and ready_sp is not None and vel > 0:
                        expected = round(ready_sp / vel, 1)
                        if abs(runway - expected) > 0.15:
                            warnings.append(f"summary_data.json[{idx}] ({entry.get('team')}): runway {runway} != ready_sp({ready_sp})/velocity({vel})={expected}")

    return errors, warnings


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 validate_backlog_report.py <output_dir>")
        sys.exit(1)

    output_dir = sys.argv[1]
    errors, warnings = validate(output_dir)

    print(f"\n{'='*60}")
    print(f"BACKLOG REPORT VALIDATION RESULTS")
    print(f"{'='*60}")

    if warnings:
        print(f"\nWARNINGS ({len(warnings)}):")
        for w in warnings:
            print(f"  [WARN] {w}")

    if errors:
        print(f"\nERRORS ({len(errors)}):")
        for e in errors:
            print(f"  [ERROR] {e}")
        print(f"\nVALIDATION: FAILED ({len(errors)} errors, {len(warnings)} warnings)")
        sys.exit(1)
    else:
        print(f"\nVALIDATION: PASSED (0 errors, {len(warnings)} warnings)")
        sys.exit(0)


if __name__ == "__main__":
    main()
