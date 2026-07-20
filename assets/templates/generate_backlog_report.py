#!/usr/bin/env python3
"""
Backlog Readiness Scanner - Canonical Report Generator
Produces Report-backlog.xlsx and Report-backlog.html from JSON data files.

Usage:
    python3 generate_backlog_report.py <backlog_data.json> <summary_data.json> <output_dir> <scan_date>

This script is the SINGLE SOURCE OF TRUTH for report formatting.
It must NEVER be modified by the LLM during skill execution.
"""
import json
import sys
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl not installed. Excel report will not be generated.")


# ============================================================
# FIXED SCHEMA CONSTANTS
# ============================================================

SHEET_NAME = "Backlog Readiness"

COLUMNS = [
    ("Team", 20),
    ("Issue Key", 14),
    ("Issue Type", 12),
    ("URL", 50),
    ("Title", 45),
    ("Status", 15),
    ("Assignee", 18),
    ("Story Points", 14),
]


# ============================================================
# EXCEL REPORT GENERATION
# ============================================================

def generate_excel(backlog_data, summary_data, output_path):
    """Generate Report-backlog.xlsx with fixed schema."""
    if not HAS_OPENPYXL:
        print("[ERROR] Cannot generate Excel report without openpyxl.")
        return False

    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # KPI calculations
    total_items = sum(t["backlog_total"] for t in summary_data)
    total_estimated = sum(t["estimated"] for t in summary_data)
    estimation_pct = round(total_estimated / total_items * 100, 1) if total_items > 0 else 0

    # KPI fonts
    kpi_bold = Font(bold=True, size=11)
    red_font = Font(bold=True, size=11, color="CC0000")
    green_font = Font(bold=True, size=11, color="006600")
    orange_font = Font(bold=True, size=11, color="CC6600")

    # === Sheet: Backlog Readiness ===
    ws = wb.active
    ws.title = SHEET_NAME

    # Row 1: Total Backlog Items
    ws.cell(row=1, column=1).value = "Total Backlog Items"
    ws.cell(row=1, column=1).font = kpi_bold
    ws.cell(row=1, column=2).value = total_items
    ws.cell(row=1, column=2).font = kpi_bold

    # Row 2: % Estimated
    ws.cell(row=2, column=1).value = "% Estimated (Story Points)"
    ws.cell(row=2, column=1).font = kpi_bold
    ws.cell(row=2, column=2).value = f"{estimation_pct}%"
    if estimation_pct >= 70:
        ws.cell(row=2, column=2).font = green_font
    elif estimation_pct >= 40:
        ws.cell(row=2, column=2).font = orange_font
    else:
        ws.cell(row=2, column=2).font = red_font

    # Row 3: empty separator

    # Row 4: Table header
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes below KPI + header
    ws.freeze_panes = 'A5'

    # Data rows (row 5+)
    for row_idx, row_data in enumerate(backlog_data, start=5):
        # Column A: Team
        ws.cell(row=row_idx, column=1).value = row_data["team"]

        # Column B: Issue Key
        ws.cell(row=row_idx, column=2).value = row_data["issue_key"]

        # Column C: Issue Type
        ws.cell(row=row_idx, column=3).value = row_data["issue_type"]

        # Column D: URL (hyperlink)
        ws.cell(row=row_idx, column=4).value = row_data["url"]
        ws.cell(row=row_idx, column=4).hyperlink = row_data["url"]
        ws.cell(row=row_idx, column=4).style = "Hyperlink"

        # Column E: Title
        ws.cell(row=row_idx, column=5).value = row_data["title"]

        # Column F: Status
        ws.cell(row=row_idx, column=6).value = row_data["status"]

        # Column G: Assignee
        ws.cell(row=row_idx, column=7).value = row_data["assignee"]

        # Column H: Story Points
        sp_value = row_data["story_points"]
        if sp_value is not None:
            ws.cell(row=row_idx, column=8).value = sp_value
        else:
            ws.cell(row=row_idx, column=8).value = None
            ws.cell(row=row_idx, column=8).fill = no_fill

        # Apply formatting to all cells in row
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col == 5))

    # Conditionally add Velocity & Runway sheet
    has_runway_data = any(t.get("sprint_runway") is not None for t in summary_data)
    if has_runway_data:
        generate_runway_sheet(wb, summary_data)

    # Save
    wb.save(output_path)
    return True


# ============================================================
# VELOCITY & RUNWAY EXCEL SHEET
# ============================================================

def generate_runway_sheet(wb, summary_data):
    """Create 'Velocity & Runway' sheet in the workbook."""
    ws = wb.create_sheet("Velocity & Runway")

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    kpi_bold = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Calculate KPI values
    runway_values = [
        t["sprint_runway"] for t in summary_data
        if t.get("sprint_runway") is not None and isinstance(t.get("sprint_runway"), (int, float))
    ]
    avg_runway = round(sum(runway_values) / len(runway_values), 2) if runway_values else 0
    teams_low_runway = sum(1 for v in runway_values if v < 1.0)

    # Row 1: Average Sprint Runway
    ws.cell(row=1, column=1).value = "Average Sprint Runway"
    ws.cell(row=1, column=1).font = kpi_bold
    ws.cell(row=1, column=2).value = avg_runway
    ws.cell(row=1, column=2).font = kpi_bold

    # Row 2: Teams with Runway < 1.0 Sprint
    ws.cell(row=2, column=1).value = "Teams with Runway < 1.0 Sprint"
    ws.cell(row=2, column=1).font = kpi_bold
    ws.cell(row=2, column=2).value = teams_low_runway
    ws.cell(row=2, column=2).font = kpi_bold

    # Row 3: separator (empty)

    # Row 4: Headers
    runway_columns = [
        ("Team", 22),
        ("Sprint -3 (SP)", 14),
        ("Sprint -2 (SP)", 14),
        ("Sprint -1 (SP)", 14),
        ("Avg Velocity", 14),
        ("Ready SP (DoR>=75%)", 20),
        ("Sprint Runway", 16),
        ("Status", 14),
    ]
    for col_idx, (header, width) in enumerate(runway_columns, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes below header
    ws.freeze_panes = 'A5'

    # Data rows (row 5+), sorted alphabetically by team name
    sorted_teams = sorted(summary_data, key=lambda t: t.get("team", "").lower())
    for row_idx, team in enumerate(sorted_teams, start=5):
        # Column A: Team
        ws.cell(row=row_idx, column=1).value = team.get("team", "")

        # Columns B-D: Sprint -3, -2, -1 (last 3 sprints)
        velocity_sprints = team.get("velocity_sprints", [])
        last_3 = velocity_sprints[-3:] if velocity_sprints else []
        for i in range(3):
            col = 2 + i
            if i < len(last_3):
                ws.cell(row=row_idx, column=col).value = last_3[i].get("completed_sp")
            else:
                ws.cell(row=row_idx, column=col).value = "-"

        # Column E: Avg Velocity
        ws.cell(row=row_idx, column=5).value = team.get("velocity_avg", "-")

        # Column F: Ready SP (DoR>=75%)
        ws.cell(row=row_idx, column=6).value = team.get("dor_ready_sp", "-")

        # Column G: Sprint Runway
        ws.cell(row=row_idx, column=7).value = team.get("sprint_runway", "-")

        # Column H: Status
        status = team.get("runway_status", "No Data")
        ws.cell(row=row_idx, column=8).value = status

        # Apply status fill color
        status_lower = status.lower() if status else ""
        if status_lower == "healthy":
            ws.cell(row=row_idx, column=8).fill = green_fill
        elif status_lower == "attention":
            ws.cell(row=row_idx, column=8).fill = yellow_fill
        elif status_lower == "critical":
            ws.cell(row=row_idx, column=8).fill = red_fill
        else:
            ws.cell(row=row_idx, column=8).fill = grey_fill

        # Apply borders to all cells in row
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', horizontal='center' if col > 1 else 'left')


# ============================================================
# HTML REPORT GENERATION
# ============================================================

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Backlog Readiness Report - SRPOL Teams</title>
  <style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
      background: #f0f2f5;
      color: #1a1a2e;
      line-height: 1.5;
      padding: 24px;
    }}
    .container {{ max-width: 1400px; margin: 0 auto; }}
    header {{
      background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
      color: #fff;
      padding: 32px 40px;
      border-radius: 12px;
      margin-bottom: 24px;
      box-shadow: 0 4px 20px rgba(0,0,0,0.15);
    }}
    header h1 {{ font-size: 28px; font-weight: 700; margin-bottom: 4px; }}
    header .subtitle {{ font-size: 14px; opacity: 0.8; }}
    .kpi-grid {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 20px;
      margin-bottom: 32px;
    }}
    .kpi-card {{
      background: #fff;
      border-radius: 10px;
      padding: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.06);
      border-top: 4px solid #ccc;
    }}
    .kpi-card.good {{ border-top-color: #27ae60; }}
    .kpi-card.warn {{ border-top-color: #f39c12; }}
    .kpi-card.bad {{ border-top-color: #e74c3c; }}
    .kpi-card .kpi-label {{
      font-size: 13px;
      text-transform: uppercase;
      letter-spacing: 0.5px;
      color: #666;
      margin-bottom: 8px;
      font-weight: 600;
    }}
    .kpi-card .kpi-value {{
      font-size: 36px;
      font-weight: 800;
    }}
    .kpi-card .kpi-detail {{
      font-size: 12px;
      color: #888;
      margin-top: 4px;
    }}
    .kpi-value.good {{ color: #27ae60; }}
    .kpi-value.warn {{ color: #f39c12; }}
    .kpi-value.bad {{ color: #e74c3c; }}
    .section {{
      background: #fff;
      border-radius: 10px;
      padding: 24px;
      margin-bottom: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.06);
    }}
    .section h2 {{
      font-size: 18px;
      font-weight: 700;
      margin-bottom: 16px;
      padding-bottom: 12px;
      border-bottom: 2px solid #f0f2f5;
      color: #1a1a2e;
    }}
    .section h2 .count {{
      font-size: 14px;
      font-weight: 400;
      color: #888;
      margin-left: 8px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }}
    thead th {{
      background: #1a1a2e;
      color: #fff;
      padding: 10px 12px;
      text-align: left;
      font-weight: 600;
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: 0.3px;
      white-space: nowrap;
    }}
    thead th.center {{ text-align: center; }}
    tbody td {{
      padding: 10px 12px;
      border-bottom: 1px solid #eee;
      vertical-align: middle;
    }}
    tbody tr:hover {{ background: #f8f9fc; }}
    tbody tr:last-child td {{ border-bottom: none; }}
    .center {{ text-align: center; }}
    .team-name {{ font-weight: 600; white-space: nowrap; }}
    .title-col {{ max-width: 350px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
    .bar-container {{
      width: 100%;
      height: 20px;
      background: #eee;
      border-radius: 4px;
      overflow: hidden;
    }}
    .bar {{
      height: 100%;
      border-radius: 4px;
    }}
    .bar.good {{ background: linear-gradient(90deg, #27ae60, #2ecc71); }}
    .bar.warn {{ background: linear-gradient(90deg, #f39c12, #f1c40f); }}
    .bar.bad {{ background: linear-gradient(90deg, #e74c3c, #c0392b); }}
    .score {{ font-weight: 700; }}
    .score.good {{ color: #27ae60; }}
    .score.warn {{ color: #f39c12; }}
    .score.bad {{ color: #e74c3c; }}
    .score.na {{ color: #aaa; }}
    .badge {{
      display: inline-block;
      padding: 3px 10px;
      border-radius: 12px;
      font-size: 11px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.3px;
    }}
    .badge.good {{ background: #d4edda; color: #155724; }}
    .badge.warn {{ background: #fff3cd; color: #856404; }}
    .badge.bad {{ background: #f8d7da; color: #721c24; }}
    .badge.na {{ background: #e2e3e5; color: #6c757d; }}
    a {{ color: #2980b9; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    footer {{
      text-align: center;
      padding: 16px;
      font-size: 12px;
      color: #999;
      margin-top: 8px;
    }}
    @media (max-width: 768px) {{
      .kpi-grid {{ grid-template-columns: 1fr; }}
      body {{ padding: 12px; }}
      header {{ padding: 20px; }}
      .section {{ padding: 16px; }}
      table {{ font-size: 11px; }}
    }}
  </style>
</head>
<body>
  <div class="container">
    <header>
      <h1>Backlog Readiness Report</h1>
      <div class="subtitle">SRPOL Teams | Scan Date: {scan_date} | Teams: {team_count}</div>
    </header>

    <div class="kpi-grid">
      <div class="kpi-card {kpi1_card_class}">
        <div class="kpi-label">Total Backlog Items</div>
        <div class="kpi-value">{kpi1_value}</div>
        <div class="kpi-detail">{kpi1_detail}</div>
      </div>
      <div class="kpi-card {kpi2_card_class}">
        <div class="kpi-label">% Estimated (Story Points)</div>
        <div class="kpi-value {kpi2_value_class}">{kpi2_value}</div>
        <div class="kpi-detail">{kpi2_detail}</div>
      </div>
      <div class="kpi-card {kpi3_card_class}">
        <div class="kpi-label">Teams &lt; 50% Estimated</div>
        <div class="kpi-value {kpi3_value_class}">{kpi3_value}</div>
        <div class="kpi-detail">{kpi3_detail}</div>
      </div>
    </div>

    <div class="section">
      <h2>Team Summary</h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th class="center">Backlog Items</th>
            <th class="center">Estimated</th>
            <th class="center">Unestimated</th>
            <th class="center">Rate</th>
            <th>Progress</th>
          </tr>
        </thead>
        <tbody>
{team_summary_rows}
        </tbody>
      </table>
    </div>

{runway_section}
    <div class="section">
      <h2>Unestimated Issues<span class="count">({unestimated_count} items)</span></h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th>Issue Key</th>
            <th>Issue Type</th>
            <th>Title</th>
            <th>Status</th>
            <th>Assignee</th>
          </tr>
        </thead>
        <tbody>
{unestimated_rows}
        </tbody>
      </table>
    </div>

    <footer>
      Backlog Readiness Scanner v2.0 | Generated by wow-team-backlog-readiness skill | Data source: SRPOL Teams Confluence
    </footer>
  </div>
</body>
</html>"""


def _color_class(value):
    """Return CSS class based on threshold."""
    if value >= 70:
        return "good"
    elif value >= 40:
        return "warn"
    else:
        return "bad"


def _escape_html(text):
    """Escape HTML special characters."""
    if not text:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _runway_status_class(status):
    """Return CSS badge class for runway status."""
    if not status:
        return "na"
    status_lower = status.lower()
    if status_lower == "healthy":
        return "good"
    elif status_lower == "attention":
        return "warn"
    elif status_lower == "critical":
        return "bad"
    else:
        return "na"


def generate_html(backlog_data, summary_data, scan_date, output_path):
    """Generate Report-backlog.html using the fixed template."""
    team_count = len(summary_data)

    # KPI 1: Total Backlog Items
    total_items = sum(t["backlog_total"] for t in summary_data)
    teams_with_items = sum(1 for t in summary_data if t["backlog_total"] > 0)
    kpi1_card_class = ""
    kpi1_value = str(total_items)
    kpi1_detail = f"{total_items} across {teams_with_items} teams"

    # KPI 2: % Estimated
    total_estimated = sum(t["estimated"] for t in summary_data)
    if total_items > 0:
        estimation_pct = round(total_estimated / total_items * 100, 1)
        kpi2_card_class = _color_class(estimation_pct)
        kpi2_value_class = _color_class(estimation_pct)
        kpi2_value = f"{estimation_pct}%"
        kpi2_detail = f"{total_estimated} of {total_items} items"
    else:
        kpi2_card_class = ""
        kpi2_value_class = ""
        kpi2_value = "-"
        kpi2_detail = "No backlog items found"

    # KPI 3: Teams < 50% estimated
    teams_below_50 = [
        t for t in summary_data
        if t["backlog_total"] > 0 and (t["estimated"] / t["backlog_total"] * 100) < 50
    ]
    count_below = len(teams_below_50)
    if count_below > 0:
        kpi3_card_class = "bad"
        kpi3_value_class = "bad"
    else:
        kpi3_card_class = "good"
        kpi3_value_class = "good"
    kpi3_value = str(count_below)
    kpi3_detail = ", ".join(t["team"] for t in teams_below_50[:5]) if teams_below_50 else "All teams above 50%"

    # Team Summary rows
    summary_lines = []
    for team in summary_data:
        bt = team["backlog_total"]
        est = team["estimated"]
        unest = team["unestimated"]
        rate_str = team["estimation_rate"]

        if bt > 0:
            rate_val = round(est / bt * 100)
            bar_class = _color_class(rate_val)
            score_cls = _color_class(rate_val)
        else:
            rate_val = 0
            bar_class = ""
            score_cls = "na"
            rate_str = "-"

        summary_lines.append(
            f'        <tr>\n'
            f'          <td class="team-name">{_escape_html(team["team"])}</td>\n'
            f'          <td class="center">{bt}</td>\n'
            f'          <td class="center">{est}</td>\n'
            f'          <td class="center">{unest}</td>\n'
            f'          <td class="center"><span class="score {score_cls}">{rate_str}</span></td>\n'
            f'          <td>\n'
            f'            <div class="bar-container">\n'
            f'              <div class="bar {bar_class}" style="width: {rate_val}%"></div>\n'
            f'            </div>\n'
            f'          </td>\n'
            f'        </tr>'
        )
    team_summary_rows = "\n".join(summary_lines)

    # Runway section (conditional)
    has_runway_data = any(t.get("sprint_runway") is not None for t in summary_data)
    if has_runway_data:
        runway_lines = []
        sorted_teams = sorted(summary_data, key=lambda t: t.get("team", "").lower())
        for team in sorted_teams:
            velocity_avg = team.get("velocity_avg", "-")
            dor_ready_sp = team.get("dor_ready_sp", "-")
            sprint_runway = team.get("sprint_runway", "-")
            runway_status = team.get("runway_status", "No Data")
            badge_cls = _runway_status_class(runway_status)

            runway_lines.append(
                f'        <tr>\n'
                f'          <td class="team-name">{_escape_html(team.get("team", ""))}</td>\n'
                f'          <td class="center">{velocity_avg}</td>\n'
                f'          <td class="center">{dor_ready_sp}</td>\n'
                f'          <td class="center">{sprint_runway}</td>\n'
                f'          <td class="center"><span class="badge {badge_cls}">{_escape_html(runway_status)}</span></td>\n'
                f'        </tr>'
            )
        runway_rows = "\n".join(runway_lines)
        runway_section = (
            '    <div class="section">\n'
            '      <h2>Velocity &amp; Runway</h2>\n'
            '      <table>\n'
            '        <thead>\n'
            '          <tr>\n'
            '            <th>Team</th>\n'
            '            <th class="center">Avg Velocity</th>\n'
            '            <th class="center">Ready SP (DoR&gt;=75%)</th>\n'
            '            <th class="center">Sprint Runway</th>\n'
            '            <th class="center">Status</th>\n'
            '          </tr>\n'
            '        </thead>\n'
            '        <tbody>\n'
            f'{runway_rows}\n'
            '        </tbody>\n'
            '      </table>\n'
            '    </div>\n'
        )
    else:
        runway_section = ""

    # Unestimated Issues rows
    unestimated_items = [d for d in backlog_data if d["story_points"] is None]
    unestimated_count = len(unestimated_items)
    unest_lines = []
    for item in unestimated_items:
        unest_lines.append(
            f'        <tr>\n'
            f'          <td class="team-name">{_escape_html(item["team"])}</td>\n'
            f'          <td><a href="{_escape_html(item["url"])}" target="_blank">{_escape_html(item["issue_key"])}</a></td>\n'
            f'          <td>{_escape_html(item["issue_type"])}</td>\n'
            f'          <td class="title-col">{_escape_html(item["title"])}</td>\n'
            f'          <td>{_escape_html(item["status"])}</td>\n'
            f'          <td>{_escape_html(item["assignee"])}</td>\n'
            f'        </tr>'
        )
    unestimated_rows = "\n".join(unest_lines)

    # Render template
    html = HTML_TEMPLATE.format(
        scan_date=scan_date,
        team_count=team_count,
        kpi1_card_class=kpi1_card_class,
        kpi1_value=kpi1_value,
        kpi1_detail=kpi1_detail,
        kpi2_card_class=kpi2_card_class,
        kpi2_value_class=kpi2_value_class,
        kpi2_value=kpi2_value,
        kpi2_detail=kpi2_detail,
        kpi3_card_class=kpi3_card_class,
        kpi3_value_class=kpi3_value_class,
        kpi3_value=kpi3_value,
        kpi3_detail=kpi3_detail,
        team_summary_rows=team_summary_rows,
        runway_section=runway_section,
        unestimated_count=unestimated_count,
        unestimated_rows=unestimated_rows,
    )

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return True


# ============================================================
# MAIN ENTRY POINT
# ============================================================

def main():
    if len(sys.argv) < 5:
        print("Usage: python3 generate_backlog_report.py <backlog_data.json> <summary_data.json> <output_dir> <scan_date>")
        print("Example: python3 generate_backlog_report.py backlog_data.json summary_data.json . 2026-06-29")
        sys.exit(1)

    backlog_path = sys.argv[1]
    summary_path = sys.argv[2]
    output_dir = sys.argv[3]
    scan_date = sys.argv[4]

    # Load data
    with open(backlog_path, 'r', encoding='utf-8') as f:
        backlog_data = json.load(f)
    with open(summary_path, 'r', encoding='utf-8') as f:
        summary_data = json.load(f)

    print(f"[INFO] Loaded data: {len(backlog_data)} backlog items, {len(summary_data)} teams")

    # Generate Excel
    xlsx_path = os.path.join(output_dir, "Report-backlog.xlsx")
    if generate_excel(backlog_data, summary_data, xlsx_path):
        print(f"[SUCCESS] Excel report: {xlsx_path}")
        print(f"  Sheet: [{SHEET_NAME}]")
        has_runway_data = any(t.get("sprint_runway") is not None for t in summary_data)
        if has_runway_data:
            print(f"  Sheet: [Velocity & Runway]")
    else:
        print("[WARNING] Excel report not generated (openpyxl missing)")

    # Generate HTML
    html_path = os.path.join(output_dir, "Report-backlog.html")
    if generate_html(backlog_data, summary_data, scan_date, html_path):
        print(f"[SUCCESS] HTML report: {html_path}")
    else:
        print("[ERROR] HTML report generation failed")

    # Summary
    total_items = len(backlog_data)
    estimated = sum(1 for d in backlog_data if d["story_points"] is not None)
    unestimated = total_items - estimated
    estimation_pct = round(estimated / total_items * 100, 1) if total_items > 0 else 0

    print(f"\n=== Report Generation Complete ===")
    print(f"  Teams: {len(summary_data)}")
    print(f"  Backlog items: {total_items} (Estimated: {estimated}, Unestimated: {unestimated})")
    print(f"  Estimation rate: {estimation_pct}%")


if __name__ == "__main__":
    main()
