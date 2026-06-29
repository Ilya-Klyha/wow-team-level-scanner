#!/usr/bin/env python3
"""
DoR Scanner - Report Validation Script
Verifies that generated reports conform to the fixed schema.

Usage:
    python3 validate_reports.py <output_dir>

Exit code 0 = all checks pass, 1 = validation failures found.
"""
import json
import sys
import os
import re
from pathlib import Path


def validate(output_dir):
    """Run all validation checks. Returns list of error strings."""
    errors = []
    output_path = Path(output_dir)

    # ================================================================
    # CHECK: Required files exist
    # ================================================================
    required_files = [
        "Report-DoR.xlsx",
        "Report-DoR.html",
        "summary_data.json",
        "compliance_data.json",
        "quality_data.json",
    ]
    for f in required_files:
        if not (output_path / f).exists():
            errors.append(f"MISSING FILE: {f}")

    # If critical files missing, return early
    if any("summary_data.json" in e for e in errors):
        return errors
    if any("compliance_data.json" in e for e in errors):
        return errors
    if any("quality_data.json" in e for e in errors):
        return errors

    # ================================================================
    # VALIDATE: summary_data.json
    # ================================================================
    try:
        with open(output_path / "summary_data.json", 'r', encoding='utf-8') as f:
            summary = json.load(f)

        if not isinstance(summary, list):
            errors.append("summary_data.json: root is not an array")
        elif len(summary) != 15:
            errors.append(f"summary_data.json: has {len(summary)} entries, expected 15")
        else:
            expected_keys = {"team", "dor", "jira_tasks", "pct_fitting_dor"}
            for i, entry in enumerate(summary):
                if set(entry.keys()) != expected_keys:
                    errors.append(f"summary_data.json[{i}] ({entry.get('team', '?')}): keys are {set(entry.keys())}, expected {expected_keys}")
                    break
                if entry["dor"] not in ("Yes", "No"):
                    errors.append(f"summary_data.json[{i}]: dor='{entry['dor']}', expected 'Yes' or 'No'")
                if not isinstance(entry["jira_tasks"], int):
                    errors.append(f"summary_data.json[{i}]: jira_tasks is not integer")
                pct = entry["pct_fitting_dor"]
                if not re.match(r"^\d+%$|^-$", str(pct)):
                    errors.append(f"summary_data.json[{i}]: pct_fitting_dor='{pct}' doesn't match 'X%' or '-'")
    except Exception as e:
        errors.append(f"summary_data.json: parse error: {e}")

    # ================================================================
    # VALIDATE: compliance_data.json
    # ================================================================
    try:
        with open(output_path / "compliance_data.json", 'r', encoding='utf-8') as f:
            compliance = json.load(f)

        if not isinstance(compliance, list):
            errors.append("compliance_data.json: root is not an array")
        else:
            expected_keys = {"team", "issue_key", "issue_type", "url", "title", "status", "assignee", "dor_compliance", "note"}
            for i, entry in enumerate(compliance):
                if set(entry.keys()) != expected_keys:
                    errors.append(f"compliance_data.json[{i}] ({entry.get('issue_key', '?')}): keys mismatch")
                    break
                if entry["dor_compliance"] not in ("Pass", "Fail"):
                    errors.append(f"compliance_data.json[{i}]: dor_compliance='{entry['dor_compliance']}', expected 'Pass' or 'Fail'")
                if entry["dor_compliance"] == "Pass" and entry["note"] != "":
                    errors.append(f"compliance_data.json[{i}] ({entry['issue_key']}): Pass but note is not empty")
                if entry["dor_compliance"] == "Fail" and entry["note"] == "":
                    errors.append(f"compliance_data.json[{i}] ({entry['issue_key']}): Fail but note is empty")
    except Exception as e:
        errors.append(f"compliance_data.json: parse error: {e}")

    # ================================================================
    # VALIDATE: quality_data.json
    # ================================================================
    try:
        with open(output_path / "quality_data.json", 'r', encoding='utf-8') as f:
            quality = json.load(f)

        if not isinstance(quality, list):
            errors.append("quality_data.json: root is not an array")
        else:
            for i, entry in enumerate(quality):
                if "team" not in entry or "coverage" not in entry or "note" not in entry:
                    errors.append(f"quality_data.json[{i}]: missing required keys (team, coverage, note)")
                    break
                cov = entry["coverage"]
                if not isinstance(cov, int):
                    errors.append(f"quality_data.json[{i}] ({entry['team']}): coverage is not integer")
                elif cov < 0 or cov > 100:
                    errors.append(f"quality_data.json[{i}] ({entry['team']}): coverage={cov} out of range [0,100]")
                elif cov % 10 != 0:
                    errors.append(f"quality_data.json[{i}] ({entry['team']}): coverage={cov} not multiple of 10")
                note = entry.get("note", "")
                if note != "All standard criteria covered" and not note.startswith("Missing:"):
                    errors.append(f"quality_data.json[{i}] ({entry['team']}): note format invalid: '{note[:50]}'")
    except Exception as e:
        errors.append(f"quality_data.json: parse error: {e}")

    # ================================================================
    # VALIDATE: Report-DoR.html
    # ================================================================
    html_path = output_path / "Report-DoR.html"
    if html_path.exists():
        try:
            html = html_path.read_text(encoding="utf-8")

            # Title check
            if "<title>DoR Scanner Report - SRPOL Teams</title>" not in html:
                errors.append("HTML: title tag doesn't match expected value")

            # KPI card count (3 cards)
            kpi_card_count = html.count('class="kpi-card')
            if kpi_card_count != 3:
                errors.append(f"HTML: found {kpi_card_count} kpi-card elements, expected 3")

            # Forbidden classes
            if "info-box" in html:
                errors.append("HTML: contains forbidden 'info-box' class")
            if 'kpi-card na' in html:
                errors.append("HTML: contains forbidden 'kpi-card na' class")

            # Footer check
            expected_footer = "DoR Scanner v2.0 | Generated by wow-dor-scanner skill | Data source: SRPOL Teams Confluence"
            if expected_footer not in html:
                errors.append("HTML: footer text doesn't match expected value")

            # Bar container height
            if "height: 20px" not in html and "height:20px" not in html:
                errors.append("HTML: bar-container height is not 20px")

            # Table header counts
            # Extract all thead sections
            thead_sections = re.findall(r'<thead>(.*?)</thead>', html, re.DOTALL)
            if len(thead_sections) >= 1:
                # Team Overview: 4 columns
                th_count = len(re.findall(r'<th', thead_sections[0]))
                if th_count != 4:
                    errors.append(f"HTML: Team Overview table has {th_count} <th> elements, expected 4")
            if len(thead_sections) >= 2:
                # Failed Issues: 7 columns
                th_count = len(re.findall(r'<th', thead_sections[1]))
                if th_count != 7:
                    errors.append(f"HTML: Failed Issues table has {th_count} <th> elements, expected 7")
            if len(thead_sections) >= 3:
                # Quality Assessment: 4 columns
                th_count = len(re.findall(r'<th', thead_sections[2]))
                if th_count != 4:
                    errors.append(f"HTML: Quality Assessment table has {th_count} <th> elements, expected 4")
            if len(thead_sections) < 3:
                errors.append(f"HTML: found {len(thead_sections)} <thead> sections, expected 3")

            # Team Overview row count (15 teams)
            tbody_sections = re.findall(r'<tbody>(.*?)</tbody>', html, re.DOTALL)
            if len(tbody_sections) >= 1:
                tr_count = len(re.findall(r'<tr>', tbody_sections[0]))
                if tr_count != 15:
                    errors.append(f"HTML: Team Overview has {tr_count} rows, expected 15")

            # No extra columns
            if ">DoR Source<" in html:
                errors.append("HTML: contains forbidden 'DoR Source' column")

        except Exception as e:
            errors.append(f"HTML: read/parse error: {e}")

    # ================================================================
    # VALIDATE: Report-DoR.xlsx
    # ================================================================
    xlsx_path = output_path / "Report-DoR.xlsx"
    if xlsx_path.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True)

            # Sheet names
            expected_sheets = ["Summary", "DoR Compliance", "DoR quality"]
            if wb.sheetnames != expected_sheets:
                errors.append(f"Excel: sheets are {wb.sheetnames}, expected {expected_sheets}")

            # Summary sheet structure
            ws_s = wb["Summary"]
            if ws_s.cell(1, 1).value != "% Teams with DoR":
                errors.append(f"Excel Summary: A1='{ws_s.cell(1, 1).value}', expected '% Teams with DoR'")
            if ws_s.cell(2, 1).value != "% Jira Tasks fitting DoR":
                errors.append(f"Excel Summary: A2='{ws_s.cell(2, 1).value}', expected '% Jira Tasks fitting DoR'")
            if ws_s.cell(4, 1).value != "Team":
                errors.append(f"Excel Summary: row 4 col A='{ws_s.cell(4, 1).value}', expected 'Team'")
            # Check 4 header columns
            header_count = sum(1 for col in range(1, 10) if ws_s.cell(4, col).value is not None)
            if header_count != 4:
                errors.append(f"Excel Summary: header row has {header_count} columns, expected 4")
            # Check data rows (should be 15)
            data_rows = sum(1 for row in range(5, 25) if ws_s.cell(row, 1).value is not None)
            if data_rows != 15:
                errors.append(f"Excel Summary: has {data_rows} data rows, expected 15")

            # DoR Compliance sheet structure
            ws_c = wb["DoR Compliance"]
            if ws_c.cell(1, 9).value != "Note":
                errors.append(f"Excel Compliance: col 9 header='{ws_c.cell(1, 9).value}', expected 'Note'")
            if ws_c.cell(1, 1).value != "Team":
                errors.append(f"Excel Compliance: col 1 header='{ws_c.cell(1, 1).value}', expected 'Team'")
            compliance_header_count = sum(1 for col in range(1, 15) if ws_c.cell(1, col).value is not None)
            if compliance_header_count != 9:
                errors.append(f"Excel Compliance: header has {compliance_header_count} columns, expected 9")

            # DoR quality sheet structure
            ws_q = wb["DoR quality"]
            if ws_q.cell(1, 1).value != "DoR quality lvl":
                errors.append(f"Excel Quality: A1='{ws_q.cell(1, 1).value}', expected 'DoR quality lvl'")
            if ws_q.cell(3, 1).value != "Team":
                errors.append(f"Excel Quality: row 3 col A='{ws_q.cell(3, 1).value}', expected 'Team'")
            quality_header_count = sum(1 for col in range(1, 10) if ws_q.cell(3, col).value is not None)
            if quality_header_count != 3:
                errors.append(f"Excel Quality: header has {quality_header_count} columns, expected 3")

            wb.close()
        except ImportError:
            errors.append("Excel: openpyxl not available, skipping Excel validation")
        except Exception as e:
            errors.append(f"Excel: validation error: {e}")

    return errors


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 validate_reports.py <output_dir>")
        sys.exit(1)

    output_dir = sys.argv[1]
    if not os.path.isdir(output_dir):
        print(f"ERROR: Directory not found: {output_dir}")
        sys.exit(1)

    print(f"Validating reports in: {output_dir}")
    print("=" * 60)

    errors = validate(output_dir)

    if errors:
        print(f"\nVALIDATION FAILED ({len(errors)} errors):")
        for e in errors:
            print(f"  [FAIL] {e}")
        sys.exit(1)
    else:
        print("\nVALIDATION PASSED - All checks OK")
        sys.exit(0)


if __name__ == "__main__":
    main()
