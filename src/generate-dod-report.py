import json, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = 'C:/Users/i.klyha/Desktop/Claude/wow-scanner-tool/assets/teams/20260612 17-00'
TOOL_DIR = 'C:/Users/i.klyha/.claude/projects/C--Users-i-klyha-Desktop-Claude-wow-scanner-tool/124276f0-644d-4500-9eb6-c31940028801/tool-results'

os.makedirs(OUTPUT_DIR, exist_ok=True)

DOD_STATUS = {
    'Abyss': 'found', 'Radium': 'found', 'Europium': 'found',
    'Copernicium': 'found', 'Mouflons': 'found', 'Wolves': 'found',
    'Polonium LF': 'found', 'Bigos': 'found', 'Capybaras': 'found',
    'ML Serving Sturgeons': 'not_found', 'ML Platform Pandas': 'found',
    'EP Core': 'found', 'Zurek': 'found', 'Igni': 'found', 'SRE': 'not_found'
}

TEAM_PATTERNS = {
    'Abyss': 'PE - WAW - Abyss', 'Radium': 'AE - WAW - Radium',
    'Europium': 'AP - WAW - Europium', 'Copernicium': 'AE - WAW - Copernicium',
    'Mouflons': 'AS - WAW - Mouflons', 'Wolves': 'AS - WAW - Wolves',
    'Polonium LF': 'Polonium', 'Bigos': 'AS - WAW - Bigos',
    'Capybaras': 'AS - WAW - Capybaras', 'ML Serving Sturgeons': 'T - WAW - ML Sturgeons',
    'ML Platform Pandas': 'T - WAW - ML Pandas', 'EP Core': 'T - WAW - EP Core',
    'Zurek': 'Zurek', 'Igni': 'AP - WAW - Igni', 'SRE': 'T - WAW - Embedded SREs SRPOL'
}

PROJECT_FILES = {
    'MAW': 'toolu_bdrk_019cEnoSjqkm5La4G4NpP2tK.txt',
    'AENW': 'toolu_bdrk_01CajqoPtDv4Cy8rznK2tWAM.txt',
    'AETVP': 'toolu_bdrk_015efTeMVxCVRYRN5XF8kaEN.txt',
    'PEPI': 'toolu_bdrk_018HpirJWW8TJW6KLaL7iHde.txt',
    'RSW': 'toolu_bdrk_01HB2cTcLZL1TeXFmqGe5Rz2.txt',
    'EPCW': 'toolu_bdrk_01RSFzSftLyviMfphMRFUZP4.txt',
    'PEA': 'toolu_bdrk_019weE5qRvvq4oJjdZmGt5DY.txt',
    'ASPW': 'toolu_bdrk_01U5AJ6JfmtcDf5q2VYpjA1W.txt',
    'EEEW': 'toolu_bdrk_011Tixssrt8gPYrnF2ymDSod.txt',
}

TEAM_PROJECTS = {
    'Abyss': 'MAW', 'Radium': 'AENW', 'Europium': 'AENW',
    'Copernicium': 'AETVP', 'Mouflons': 'PEPI', 'Wolves': 'PEPI',
    'Polonium LF': 'RSW', 'Bigos': 'MAW', 'Capybaras': 'RSW',
    'ML Serving Sturgeons': 'PEPI', 'ML Platform Pandas': 'ML',
    'EP Core': 'EPCW', 'Zurek': 'PEA', 'Igni': 'ASPW', 'SRE': 'EEEW'
}


def to_kebab(name):
    return re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')


def get_team_value(fields):
    tf = fields.get('customfield_10114')
    if not tf:
        return ''
    if isinstance(tf, dict):
        return tf.get('name', '') or tf.get('value', '')
    if isinstance(tf, list):
        return tf[0].get('name', '') if tf else ''
    return str(tf)


# Load project data
project_issues = {}
for proj, fname in PROJECT_FILES.items():
    fpath = os.path.join(TOOL_DIR, fname)
    if os.path.exists(fpath):
        with open(fpath, 'r', encoding='utf-8') as f:
            data = json.load(f)
            project_issues[proj] = data.get('issues', {}).get('nodes', [])
    else:
        project_issues[proj] = []

# Filter issues by team
team_issues = {}
for team_name, pattern in TEAM_PATTERNS.items():
    proj = TEAM_PROJECTS[team_name]
    issues = project_issues.get(proj, [])
    matched = []
    for issue in issues:
        fields = issue.get('fields', {})
        team_val = get_team_value(fields)
        if pattern.lower() in team_val.lower() or team_val.lower() in pattern.lower():
            matched.append(issue)
    team_issues[team_name] = matched

# Load DoD content
team_dods = {}
for team in DOD_STATUS:
    if DOD_STATUS[team] == 'found':
        dod_file = os.path.join(OUTPUT_DIR, f'{to_kebab(team)}-dod.txt')
        if os.path.exists(dod_file):
            with open(dod_file, 'r', encoding='utf-8') as f:
                team_dods[team] = f.read()


def analyze_dod_compliance(team_name, dod_text, issue):
    """Analyze DoD compliance based on description content."""
    desc = (issue.get('description') or '').lower()
    summary_lower = (issue.get('summary') or '').lower()
    is_investigation = any(kw in summary_lower for kw in [
        'investigate', 'spike', 'research', 'poc', 'exploration',
        'analysis', 'investigation', 'explore', 'study'
    ])

    dod_lower = dod_text.lower()
    missing = []

    # 1. Tests evidence
    has_test_criterion = any(kw in dod_lower for kw in [
        'unit and integration test', 'tests written', 'all new logic tested',
        'sonarqube', 'github checks', 'automated pipeline'
    ])
    if has_test_criterion and not is_investigation:
        test_evidence = any(kw in desc for kw in [
            'test', 'tested', 'ci pass', 'pipeline', 'sonarqube',
            'coverage', 'unit test', 'integration test', 'checks pass',
            'all tests', 'green', 'ci/', 'github actions'
        ])
        if not test_evidence:
            if 'unit and integration' in dod_lower:
                missing.append("DoD criterion 'Unit and Integration Tests': no test evidence in description")
            elif 'sonarqube' in dod_lower:
                missing.append("DoD criterion 'SonarQube quality gate': no quality gate evidence")
            elif 'github checks' in dod_lower:
                missing.append("DoD criterion 'Github checks': no CI evidence in description")
            elif 'automated pipeline' in dod_lower:
                missing.append("DoD criterion 'Automated Pipeline Passed': no CI/pipeline evidence")
            else:
                missing.append("DoD criterion 'Tests': no test evidence in description")

    # 2. Code review evidence
    has_review_criterion = any(kw in dod_lower for kw in [
        'code review', 'review is completed', 'review must be accepted',
        'sent for approval', 'self-reviewed'
    ])
    if has_review_criterion and not is_investigation:
        review_evidence = any(kw in desc for kw in [
            'review', 'pr', 'pull request', 'approved', 'lgtm',
            'merged', 'github.', 'code review', '/pull/'
        ])
        if not review_evidence:
            missing.append("DoD criterion 'Code Review': no review evidence in description")

    # 3. Acceptance criteria
    has_ac_criterion = 'acceptance criteria' in dod_lower
    if has_ac_criterion:
        ac_evidence = any(kw in desc for kw in [
            'acceptance criteria', 'ac:', 'ac ', 'criteria met',
            'verified', 'validated', 'approved by po'
        ])
        # If description is very short and has no AC info
        if not ac_evidence and len(desc.strip()) < 100:
            missing.append("DoD criterion 'Acceptance Criteria Met': no AC evidence in description")

    # 4. Deployment
    has_deploy_criterion = any(kw in dod_lower for kw in [
        'deployed to', 'released to production', 'deploy readiness',
        'merged to main', 'released on production', 'feature flag'
    ])
    if has_deploy_criterion and not is_investigation:
        deploy_evidence = any(kw in desc for kw in [
            'deploy', 'production', 'staging', 'released', 'merged',
            'live', 'rollout', 'feature flag', 'prod', 'pre-prod'
        ])
        if not deploy_evidence:
            missing.append("DoD criterion 'Deployment': no deployment evidence in description")

    # 5. Documentation
    has_doc_criterion = any(kw in dod_lower for kw in [
        'documentation updated', 'documentation should be updated',
        'runbook', 'docs updated', 'required documentation'
    ])
    if has_doc_criterion and not is_investigation:
        doc_evidence = any(kw in desc for kw in [
            'doc', 'documentation', 'readme', 'confluence', 'runbook',
            'wiki', 'updated doc'
        ])
        if not doc_evidence and len(desc.strip()) > 200:
            missing.append("DoD criterion 'Documentation': no documentation evidence in description")

    # 6. Jira comment/summary requirement
    has_jira_doc = any(kw in dod_lower for kw in [
        'jira comment', 'result summary', 'short description in the comment',
        'outputs documented in jira', 'at least 1 comment'
    ])
    if has_jira_doc:
        # Cannot verify comments from search data
        missing.append("DoD criterion 'Jira Documentation': comments not verifiable from search data")

    # 7. Demo
    has_demo = 'demo' in dod_lower and 'demo before' in dod_lower
    if has_demo:
        demo_evidence = any(kw in desc for kw in ['demo', 'demonstrated', 'sprint review'])
        if not demo_evidence:
            missing.append("DoD criterion 'Demo before close': no demo evidence in description")

    return len(missing) == 0, missing[:5]


# Simplify issues for processing
def simplify_issue(issue):
    f = issue.get('fields', {})
    assignee = f.get('assignee')
    return {
        'key': issue.get('key'),
        'summary': f.get('summary', ''),
        'issuetype': f.get('issuetype', {}).get('name', ''),
        'status': 'Done',
        'assignee': assignee.get('displayName', 'Unassigned') if assignee else 'Unassigned',
        'resolved': f.get('resolved', ''),
        'description': f.get('description', '') or ''
    }


# Perform DoD analysis
report_data = []
team_stats = {}

for team in DOD_STATUS:
    raw_issues = team_issues.get(team, [])
    issues = [simplify_issue(i) for i in raw_issues]
    has_dod = DOD_STATUS[team] == 'found'
    dod_text = team_dods.get(team, '')

    compliant_count = 0
    total_analyzed = 0

    for issue in issues:
        total_analyzed += 1
        if has_dod and dod_text:
            meets_dod, missing = analyze_dod_compliance(team, dod_text, issue)
            if meets_dod:
                compliant_count += 1
            report_data.append({
                'team': team,
                'issue_key': issue['key'],
                'issue_type': issue.get('issuetype', ''),
                'url': f"https://adgear.atlassian.net/browse/{issue['key']}",
                'title': issue.get('summary', ''),
                'status': 'Done',
                'assignee': issue.get('assignee', 'Unassigned'),
                'dod_compliance': 'Pass' if meets_dod else 'Fail',
                'note': '; '.join(missing) if not meets_dod else ''
            })
        else:
            report_data.append({
                'team': team,
                'issue_key': issue['key'],
                'issue_type': issue.get('issuetype', ''),
                'url': f"https://adgear.atlassian.net/browse/{issue['key']}",
                'title': issue.get('summary', ''),
                'status': 'Done',
                'assignee': issue.get('assignee', 'Unassigned'),
                'dod_compliance': 'Fail',
                'note': 'No DoD defined for team'
            })

    team_stats[team] = {
        'has_dod': has_dod,
        'issue_count': len(issues),
        'compliant': compliant_count,
        'total_analyzed': total_analyzed
    }

# Generate Excel Report
wb = Workbook()

# --- Sheet 1: Summary ---
ws_summary = wb.active
ws_summary.title = 'Summary'

header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# KPI Section
teams_with_dod = sum(1 for v in DOD_STATUS.values() if v == 'found')
total_teams = len(DOD_STATUS)
pct_dod = round(teams_with_dod / total_teams * 100)

total_compliant = sum(s['compliant'] for s in team_stats.values())
total_analyzed = sum(s['total_analyzed'] for s in team_stats.values())
pct_fitting = round(total_compliant / total_analyzed * 100) if total_analyzed > 0 else 0

ws_summary['A1'] = '% Teams with DoD'
ws_summary['B1'] = f'{pct_dod}%'
ws_summary['A2'] = '% Jira Tasks fitting DoD'
ws_summary['B2'] = f'{pct_fitting}%'

for row in [1, 2]:
    ws_summary.cell(row=row, column=1).font = Font(bold=True)
    pct_val = pct_dod if row == 1 else pct_fitting
    color = '006100' if pct_val >= 70 else ('9C5700' if pct_val >= 40 else 'C00000')
    ws_summary.cell(row=row, column=2).font = Font(bold=True, color=color)

# Table headers (row 4)
headers = ['Team', 'DoD', 'Jira Tasks Done (14d)', '% Tasks fitting DoD']
widths = [25, 10, 20, 18]
for col, (header, width) in enumerate(zip(headers, widths), 1):
    cell = ws_summary.cell(row=4, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    ws_summary.column_dimensions[get_column_letter(col)].width = width

# Data rows
row_num = 5
for team in DOD_STATUS:
    stats = team_stats[team]
    ws_summary.cell(row=row_num, column=1, value=team).border = thin_border

    dod_cell = ws_summary.cell(row=row_num, column=2, value='Yes' if stats['has_dod'] else 'No')
    dod_cell.fill = green_fill if stats['has_dod'] else red_fill
    dod_cell.border = thin_border
    dod_cell.alignment = Alignment(horizontal='center')

    ws_summary.cell(row=row_num, column=3, value=stats['issue_count']).border = thin_border
    ws_summary.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')

    if stats['total_analyzed'] > 0 and stats['has_dod']:
        pct = round(stats['compliant'] / stats['total_analyzed'] * 100)
        ws_summary.cell(row=row_num, column=4, value=f'{pct}%').border = thin_border
    else:
        ws_summary.cell(row=row_num, column=4, value='-').border = thin_border
    ws_summary.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')

    row_num += 1

ws_summary.freeze_panes = 'A5'

# --- Sheet 2: DoD Compliance ---
ws_dod = wb.create_sheet('DoD Compliance')

dod_headers = ['Team', 'Issue Key', 'Issue Type', 'URL', 'Title', 'Status', 'Assignee', 'DoD Compliance', 'Note']
dod_widths = [15, 12, 10, 50, 40, 12, 15, 15, 60]

for col, (header, width) in enumerate(zip(dod_headers, dod_widths), 1):
    cell = ws_dod.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    ws_dod.column_dimensions[get_column_letter(col)].width = width

# Data rows
for row_idx, item in enumerate(report_data, 2):
    ws_dod.cell(row=row_idx, column=1, value=item['team']).border = thin_border
    ws_dod.cell(row=row_idx, column=2, value=item['issue_key']).border = thin_border
    ws_dod.cell(row=row_idx, column=3, value=item['issue_type']).border = thin_border
    ws_dod.cell(row=row_idx, column=4, value=item['url']).border = thin_border

    title_cell = ws_dod.cell(row=row_idx, column=5, value=item['title'])
    title_cell.border = thin_border
    title_cell.alignment = Alignment(wrap_text=True)

    ws_dod.cell(row=row_idx, column=6, value=item['status']).border = thin_border
    ws_dod.cell(row=row_idx, column=7, value=item['assignee']).border = thin_border

    compliance_cell = ws_dod.cell(row=row_idx, column=8, value=item['dod_compliance'])
    compliance_cell.border = thin_border
    compliance_cell.alignment = Alignment(horizontal='center')
    compliance_cell.fill = green_fill if item['dod_compliance'] == 'Pass' else red_fill

    note_cell = ws_dod.cell(row=row_idx, column=9, value=item['note'])
    note_cell.border = thin_border
    note_cell.alignment = Alignment(wrap_text=True)

ws_dod.freeze_panes = 'A2'

# Save Excel
excel_path = os.path.join(OUTPUT_DIR, 'Report-DoD.xlsx')
wb.save(excel_path)
print(f'Report-DoD.xlsx saved to: {excel_path}')

# Generate DOD_ANALYSIS_SUMMARY.md
summary_lines = [
    '# DoD Compliance Analysis Summary',
    '',
    '**Date:** 2026-06-12 17:00 CET',
    '**Source:** SRPOL Teams Confluence Page',
    '**Period:** Last 14 days (resolved issues)',
    '',
    '## KPIs',
    '',
    f'- **% Teams with DoD:** {pct_dod}% ({teams_with_dod}/{total_teams} teams)',
    f'- **% Jira Tasks fitting DoD:** {pct_fitting}% ({total_compliant}/{total_analyzed} issues)',
    '',
    '## Team Summary',
    '',
    '| Team | DoD | Issues Done | % Fitting DoD |',
    '|------|-----|-------------|---------------|',
]

for team in DOD_STATUS:
    stats = team_stats[team]
    dod_str = 'Yes' if stats['has_dod'] else 'No'
    if stats['total_analyzed'] > 0 and stats['has_dod']:
        pct_str = f"{round(stats['compliant']/stats['total_analyzed']*100)}%"
    else:
        pct_str = '-'
    summary_lines.append(f"| {team} | {dod_str} | {stats['issue_count']} | {pct_str} |")

summary_lines.extend([
    '',
    '## Analysis Notes',
    '',
    f'- Total teams scanned: {total_teams}',
    f'- Teams with DoD defined: {teams_with_dod}',
    f'- Teams without DoD: {total_teams - teams_with_dod} (ML Serving Sturgeons, SRE)',
    f'- Total issues analyzed: {total_analyzed}',
    f'- Compliant issues: {total_compliant}',
    f'- Non-compliant issues: {total_analyzed - total_compliant}',
    '',
    '## Known Limitations',
    '',
    '1. Analysis is limited to evidence documented in Jira (description only). Comments were not individually fetched due to API call volume constraints.',
    '2. Teams using external tools (GitHub PRs, CI dashboards) without linking to Jira may show false negatives.',
    '3. Does not check issue status transition history.',
    '4. "Fail" means no documented evidence in Jira description, not necessarily that work was not done.',
    '5. Heuristic-based analysis checks for keywords indicating compliance evidence.',
])

summary_path = os.path.join(OUTPUT_DIR, 'DOD_ANALYSIS_SUMMARY.md')
with open(summary_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(summary_lines))

print(f'DOD_ANALYSIS_SUMMARY.md saved.')
print(f'\nFinal Stats:')
print(f'  Teams with DoD: {teams_with_dod}/{total_teams} ({pct_dod}%)')
print(f'  Issues fitting DoD: {total_compliant}/{total_analyzed} ({pct_fitting}%)')
