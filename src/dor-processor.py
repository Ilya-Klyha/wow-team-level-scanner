"""
DoR Scanner - Jira Data Processing, Compliance Analysis, and Report Generation

DEPRECATED: This script uses the legacy 7-dimension quality model.
For report generation, use assets/templates/generate_reports.py instead.
This file is retained for reference only and should NOT be executed for production reports.
"""
import json
import os
import sys
from datetime import datetime

# ============================================================
# CONFIGURATION
# ============================================================
OUTPUT_DIR = r"C:\Users\i.klyha\Desktop\Claude\wow-scanner-tool\assets\teams\20260618 11-09"
JIRA_RESULTS_DIR = r"C:\Users\i.klyha\.claude\projects\C--Users-i-klyha-Desktop-Claude-wow-scanner-tool\56ffa65a-be32-46e7-9c43-5a4da39105c1\tool-results"

# Map project -> result file
PROJECT_FILES = {
    "MAW": "toolu_bdrk_01KTuUYxuMLxThjSQzYRBBms.txt",
    "AENW": "toolu_bdrk_015xKUDVhGdifpSnLqbRSSW6.txt",
    "AETVP": "toolu_bdrk_01DRE4nwoZN8KYbnrRG6FaTi.txt",
    "PEPI": "toolu_bdrk_01Ha9FNHs7MeVKA1EixotFRx.txt",
    "RSW": "toolu_bdrk_01CHcUWwfvjWrCn9bCZsY9Wc.txt",
    "EPCW": "toolu_bdrk_017z6SRjRrePc5tecKDszQ1p.txt",
    "PEA": "toolu_bdrk_01HwSZ6GxEkzuq6ERtNrx21G.txt",
    "ASPW": "toolu_bdrk_01Vbf18vhGNy6WtJgdVVvfBR.txt",
    "EEEW": "toolu_bdrk_01Vj3VcQKF4kYaCByZ8m4ozn.txt",
}

# Team mapping: team_name -> jira_team_field_value
TEAM_MAPPING = {
    "Abyss": "PE - WAW - Abyss",
    "Radium": "AE - WAW - Radium",
    "Europium": "AP - WAW - Europium",
    "Copernicium": "AE - WAW - Copernicium",
    "Mouflons": "AS - WAW - Mouflons",
    "Wolves": "AS - WAW - Wolves",
    "Polonium UF": "AS - WAW - Polonium UF",
    "Bigos": "AS - WAW - Bigos",
    "Capybaras": "AS - WAW - Capybaras",
    "ML Serving Sturgeons": "T - WAW - ML Sturgeons",
    "ML Platform Pandas": "T - WAW - ML Pandas",
    "EP Core": "T - WAW - EP Core",
    "Zurek": "Zurek",
    "Igni": "AP - WAW - Igni",
    "SRE": "T - WAW - Embedded SREs SRPOL",
}

# Team name -> kebab-case filename prefix
TEAM_KEBAB = {
    "Abyss": "abyss",
    "Radium": "radium",
    "Europium": "europium",
    "Copernicium": "copernicium",
    "Mouflons": "mouflons",
    "Wolves": "wolves",
    "Polonium UF": "polonium-uf",
    "Bigos": "bigos",
    "Capybaras": "capybaras",
    "ML Serving Sturgeons": "ml-serving-sturgeons",
    "ML Platform Pandas": "ml-platform-pandas",
    "EP Core": "ep-core",
    "Zurek": "zurek",
    "Igni": "igni",
    "SRE": "sre",
}

# Teams without DoR (skip analysis)
NO_DOR_TEAMS = ["ML Serving Sturgeons", "SRE"]


def load_jira_data():
    """Load all Jira results and filter by team."""
    # Reverse mapping: jira field value -> team name
    field_to_team = {}
    for team_name, field_value in TEAM_MAPPING.items():
        field_to_team[field_value] = team_name
    # Handle variant for Zurek
    field_to_team["T - WAW - Zurek"] = "Zurek"

    team_issues = {name: [] for name in TEAM_MAPPING.keys()}

    for project, filename in PROJECT_FILES.items():
        filepath = os.path.join(JIRA_RESULTS_DIR, filename)
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        nodes = data.get("issues", {}).get("nodes", [])
        for node in nodes:
            fields = node.get("fields", {})
            team_field = fields.get("customfield_10114")
            team_name_value = None
            if team_field and isinstance(team_field, dict):
                team_name_value = team_field.get("name", "")

            issue_info = {
                "key": node.get("key", ""),
                "summary": fields.get("summary", ""),
                "status": fields.get("status", {}).get("name", "") if fields.get("status") else "",
                "issuetype": fields.get("issuetype", {}).get("name", "") if fields.get("issuetype") else "",
                "assignee": fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned",
                "priority": fields.get("priority", {}).get("name", "") if fields.get("priority") else "",
                "description": fields.get("description", "") or "",
                "team_field": team_name_value,
                "url": node.get("webUrl", f"https://adgear.atlassian.net/browse/{node.get('key', '')}"),
            }

            # Match to team
            if team_name_value and team_name_value in field_to_team:
                matched_team = field_to_team[team_name_value]
                team_issues[matched_team].append(issue_info)

    return team_issues


def save_jira_files(team_issues):
    """Save jira JSON and TXT files per team."""
    for team_name, issues in team_issues.items():
        kebab = TEAM_KEBAB[team_name]

        # Save JSON
        json_path = os.path.join(OUTPUT_DIR, f"{kebab}-jira.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({
                "team": team_name,
                "query_date": "2026-06-18",
                "issue_count": len(issues),
                "issues": issues
            }, f, indent=2, ensure_ascii=False)

        # Save TXT
        txt_path = os.path.join(OUTPUT_DIR, f"{kebab}-jira.txt")
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(f"JIRA ACTIVE ISSUES - {team_name}\n")
            f.write(f"{'=' * 60}\n")
            f.write(f"Query Date: 2026-06-18\n")
            f.write(f"Total Issues: {len(issues)}\n")
            f.write(f"{'=' * 60}\n\n")
            if not issues:
                f.write("No active issues found for this team.\n")
            for issue in issues:
                f.write(f"[{issue['key']}] {issue['summary']}\n")
                f.write(f"  Type: {issue['issuetype']} | Status: {issue['status']} | Assignee: {issue['assignee']}\n")
                f.write(f"  URL: {issue['url']}\n\n")

    print(f"Saved Jira files for {len(team_issues)} teams")


def load_dor_files():
    """Load DoR text files for each team."""
    dor_content = {}
    for team_name in TEAM_MAPPING.keys():
        if team_name in NO_DOR_TEAMS:
            continue
        kebab = TEAM_KEBAB[team_name]
        dor_path = os.path.join(OUTPUT_DIR, f"{kebab}-dor.txt")
        if os.path.exists(dor_path):
            with open(dor_path, "r", encoding="utf-8") as f:
                dor_content[team_name] = f.read()
        else:
            dor_content[team_name] = ""
    return dor_content


def assess_dor_compliance(issue, dor_content):
    """Assess a single issue against DoR criteria. Returns (pass/fail, note)."""
    desc = issue.get("description", "") or ""
    summary = issue.get("summary", "") or ""
    issue_type = issue.get("issuetype", "")
    desc_lower = desc.lower()

    # Check for key DoR elements
    has_description = len(desc.strip()) > 50
    has_ac = any(kw in desc_lower for kw in [
        "acceptance criteria", "ac:", "ac1:", "ac-1",
        "condition to complete", "expected", "given/when/then",
        "given", "criteria", "requirements"
    ])
    has_user_story = any(kw in desc_lower for kw in [
        "as a", "i want", "so that", "as an"
    ])
    has_scope = any(kw in desc_lower for kw in [
        "scope", "out of scope", "in scope", "not in scope"
    ])
    has_technical_context = any(kw in desc_lower for kw in [
        "context", "background", "technical", "architecture",
        "design", "approach", "solution"
    ])

    # For bugs specifically
    is_bug = issue_type == "Bug"
    has_repro_steps = any(kw in desc_lower for kw in [
        "steps to reproduce", "reproduction", "repro",
        "current state", "expected", "actual"
    ])

    # Scoring
    score = 0
    max_score = 3  # Minimum passing: description + AC + (story or context)

    if has_description:
        score += 1
    if has_ac:
        score += 1
    if is_bug:
        if has_repro_steps or has_description:
            score += 1
    else:
        if has_user_story or has_technical_context:
            score += 1

    # Determine pass/fail
    if score >= 2:
        return "Pass", ""
    else:
        notes = []
        if not has_description:
            notes.append("Missing description")
        if not has_ac:
            notes.append("No acceptance criteria")
        if not is_bug and not has_user_story and not has_technical_context:
            notes.append("No user story or context")
        if is_bug and not has_repro_steps and not has_description:
            notes.append("No reproduction steps")
        return "Fail", "; ".join(notes)


def run_compliance_analysis(team_issues, dor_content):
    """Run DoR compliance analysis for all teams with DoR + issues."""
    compliance_results = {}

    for team_name, issues in team_issues.items():
        if team_name in NO_DOR_TEAMS:
            continue
        if not issues:
            compliance_results[team_name] = []
            continue
        if team_name not in dor_content or not dor_content.get(team_name, "").strip():
            continue

        team_results = []
        for issue in issues:
            result, note = assess_dor_compliance(issue, dor_content.get(team_name, ""))
            team_results.append({
                "team": team_name,
                "key": issue["key"],
                "issuetype": issue["issuetype"],
                "url": issue["url"],
                "summary": issue["summary"],
                "status": issue["status"],
                "assignee": issue["assignee"],
                "compliance": result,
                "note": note,
            })
        compliance_results[team_name] = team_results

    return compliance_results


def assess_dor_quality(dor_content):
    """Score each team's DoR 0-100 across 7 dimensions."""
    quality_scores = {}

    for team_name, content in dor_content.items():
        if not content.strip():
            quality_scores[team_name] = {
                "coverage": 0, "clarity": 0, "measurability": 0,
                "company_alignment": 0, "industry_alignment": 0,
                "actionability": 0, "ac_guidance": 0, "total": 0
            }
            continue

        content_lower = content.lower()

        # 1. COVERAGE (25%) - check 10 essential areas
        coverage_areas = 0
        if any(kw in content_lower for kw in ["user story", "requirement", "description", "clearly defined"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["acceptance criteria", "ac ", "ac:", "criteria"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["estimat", "story point", "sizing", "size"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["dependenc", "blocked", "blocker"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["design", "mockup", "wireframe", "technical spec", "ux", "ui"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["scope", "sprint fit", "fits in sprint", "single sprint"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["risk", "blocker", "impediment"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["product owner", "po ", "stakeholder", "approval", "reviewed"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["feasib", "technical feasibility", "spike", "investigation"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["test", "testing", "qa", "quality"]):
            coverage_areas += 1
        coverage_score = int((coverage_areas / 10) * 100)

        # 2. CLARITY & SPECIFICITY (20%)
        specificity_indicators = 0
        total_checks = 5
        if any(kw in content_lower for kw in ["must", "shall", "required", "mandatory"]):
            specificity_indicators += 1
        if any(kw in content_lower for kw in ["defined", "documented", "written", "specified"]):
            specificity_indicators += 1
        if any(kw in content_lower for kw in ["verif", "check", "confirm", "validated"]):
            specificity_indicators += 1
        # Penalize vagueness
        vague_terms = sum(1 for kw in ["appropriate", "sufficient", "if needed", "as needed", "if applicable"]
                         if kw in content_lower)
        if vague_terms <= 1:
            specificity_indicators += 1
        if len(content) > 200:  # Substantive content
            specificity_indicators += 1
        clarity_score = int((specificity_indicators / total_checks) * 100)

        # 3. MEASURABILITY (15%)
        measurability_indicators = 0
        if any(kw in content_lower for kw in ["pass/fail", "yes/no", "checklist", "checkbox"]):
            measurability_indicators += 2
        if any(kw in content_lower for kw in ["verified", "confirmed", "approved", "signed off"]):
            measurability_indicators += 1
        if any(kw in content_lower for kw in ["all team", "team agrees", "consensus", "reviewed by"]):
            measurability_indicators += 1
        measurability_score = min(100, int((measurability_indicators / 3) * 100))

        # 4. COMPANY STANDARD ALIGNMENT (15%)
        # Company standard emphasizes: AC, dependencies, scope/priority
        company_indicators = 0
        if "acceptance criteria" in content_lower or "ac " in content_lower:
            company_indicators += 1
        if "dependenc" in content_lower:
            company_indicators += 1
        if "scope" in content_lower or "priority" in content_lower:
            company_indicators += 1
        if "confidence" in content_lower or "clarity" in content_lower:
            company_indicators += 1
        company_score = int((company_indicators / 4) * 100)

        # 5. INDUSTRY BEST PRACTICES ALIGNMENT (10%)
        # INVEST, Scrum Guide, SAFe
        industry_indicators = 0
        if any(kw in content_lower for kw in ["independent", "negotiable", "valuable", "estimable", "small", "testable", "invest"]):
            industry_indicators += 1
        if any(kw in content_lower for kw in ["sprint", "iteration", "increment"]):
            industry_indicators += 1
        if any(kw in content_lower for kw in ["refinement", "grooming", "backlog"]):
            industry_indicators += 1
        industry_score = int((industry_indicators / 3) * 100)

        # 6. ACTIONABILITY (10%)
        actionability_indicators = 0
        if any(kw in content_lower for kw in ["responsible", "owner", "who", "role"]):
            actionability_indicators += 1
        if any(kw in content_lower for kw in ["before", "prior to", "gate", "block"]):
            actionability_indicators += 1
        if any(kw in content_lower for kw in ["process", "workflow", "step", "procedure"]):
            actionability_indicators += 1
        actionability_score = int((actionability_indicators / 3) * 100)

        # 7. AC GUIDANCE (5%)
        ac_indicators = 0
        if any(kw in content_lower for kw in ["given/when/then", "given when then", "gherkin"]):
            ac_indicators += 2
        if any(kw in content_lower for kw in ["format", "structured", "template"]):
            ac_indicators += 1
        if "acceptance criteria" in content_lower and any(kw in content_lower for kw in ["specific", "measurable", "testable"]):
            ac_indicators += 1
        ac_score = min(100, int((ac_indicators / 3) * 100))

        # Calculate weighted total
        total = int(
            coverage_score * 0.25 +
            clarity_score * 0.20 +
            measurability_score * 0.15 +
            company_score * 0.15 +
            industry_score * 0.10 +
            actionability_score * 0.10 +
            ac_score * 0.05
        )

        quality_scores[team_name] = {
            "coverage": coverage_score,
            "clarity": clarity_score,
            "measurability": measurability_score,
            "company_alignment": company_score,
            "industry_alignment": industry_score,
            "actionability": actionability_score,
            "ac_guidance": ac_score,
            "total": total,
        }

    return quality_scores


def generate_excel_report(team_issues, compliance_results, quality_scores):
    """Generate Report-DoR.xlsx with 3 sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ============================================================
    # SHEET 1: Summary
    # ============================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Styles
    header_font = Font(bold=True, size=12)
    kpi_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Calculate KPIs
    total_teams = len(TEAM_MAPPING)
    teams_with_dor = total_teams - len(NO_DOR_TEAMS)
    pct_teams_with_dor = round((teams_with_dor / total_teams) * 100, 1)

    total_tasks = 0
    total_passing = 0
    for team_name, results in compliance_results.items():
        total_tasks += len(results)
        total_passing += sum(1 for r in results if r["compliance"] == "Pass")

    pct_tasks_fitting_dor = round((total_passing / total_tasks * 100), 1) if total_tasks > 0 else 0

    # KPI Section
    ws_summary["A1"] = "DoR Scanner Report - SRPOL Teams"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A2"] = f"Scan Date: 2026-06-18"
    ws_summary["A3"] = ""
    ws_summary["A4"] = "KEY PERFORMANCE INDICATORS"
    ws_summary["A4"].font = kpi_font
    ws_summary["A5"] = f"% Teams with DoR defined:"
    ws_summary["B5"] = f"{pct_teams_with_dor}%"
    ws_summary["B5"].font = Font(bold=True, size=12)
    ws_summary["A6"] = f"% Tasks fitting DoR:"
    ws_summary["B6"] = f"{pct_tasks_fitting_dor}%"
    ws_summary["B6"].font = Font(bold=True, size=12)
    ws_summary["A7"] = f"Total active tasks analyzed:"
    ws_summary["B7"] = total_tasks

    # Table header
    row = 9
    headers = ["Team", "DoR Defined", "Jira Tasks In Progress", "% Tasks Fitting DoR"]
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Table data
    row = 10
    for team_name in TEAM_MAPPING.keys():
        has_dor = "Yes" if team_name not in NO_DOR_TEAMS else "No"
        issues = team_issues.get(team_name, [])
        issue_count = len(issues)

        if team_name in compliance_results and compliance_results[team_name]:
            results = compliance_results[team_name]
            passing = sum(1 for r in results if r["compliance"] == "Pass")
            pct = round((passing / len(results) * 100), 1) if results else 0
            pct_str = f"{pct}%"
        elif team_name in NO_DOR_TEAMS:
            pct_str = "N/A (no DoR)"
        elif issue_count == 0:
            pct_str = "N/A (no tasks)"
        else:
            pct_str = "N/A"

        ws_summary.cell(row=row, column=1, value=team_name).border = thin_border
        ws_summary.cell(row=row, column=2, value=has_dor).border = thin_border
        ws_summary.cell(row=row, column=3, value=issue_count).border = thin_border
        ws_summary.cell(row=row, column=4, value=pct_str).border = thin_border
        row += 1

    # Column widths
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 18
    ws_summary.column_dimensions["C"].width = 22
    ws_summary.column_dimensions["D"].width = 22

    # ============================================================
    # SHEET 2: DoR Compliance
    # ============================================================
    ws_compliance = wb.create_sheet("DoR Compliance")

    headers = ["Team", "Issue Key", "Issue Type", "URL", "Title", "Status", "Assignee", "DoR Compliance", "Note"]
    for col, h in enumerate(headers, 1):
        cell = ws_compliance.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    row = 2
    for team_name in TEAM_MAPPING.keys():
        if team_name not in compliance_results:
            continue
        for result in compliance_results[team_name]:
            ws_compliance.cell(row=row, column=1, value=result["team"]).border = thin_border
            ws_compliance.cell(row=row, column=2, value=result["key"]).border = thin_border
            ws_compliance.cell(row=row, column=3, value=result["issuetype"]).border = thin_border
            ws_compliance.cell(row=row, column=4, value=result["url"]).border = thin_border
            ws_compliance.cell(row=row, column=5, value=result["summary"]).border = thin_border
            ws_compliance.cell(row=row, column=6, value=result["status"]).border = thin_border
            ws_compliance.cell(row=row, column=7, value=result["assignee"]).border = thin_border

            compliance_cell = ws_compliance.cell(row=row, column=8, value=result["compliance"])
            compliance_cell.border = thin_border
            if result["compliance"] == "Pass":
                compliance_cell.fill = pass_fill
            else:
                compliance_cell.fill = fail_fill

            ws_compliance.cell(row=row, column=9, value=result["note"]).border = thin_border
            row += 1

    # Column widths for compliance sheet
    ws_compliance.column_dimensions["A"].width = 20
    ws_compliance.column_dimensions["B"].width = 14
    ws_compliance.column_dimensions["C"].width = 12
    ws_compliance.column_dimensions["D"].width = 45
    ws_compliance.column_dimensions["E"].width = 50
    ws_compliance.column_dimensions["F"].width = 14
    ws_compliance.column_dimensions["G"].width = 22
    ws_compliance.column_dimensions["H"].width = 16
    ws_compliance.column_dimensions["I"].width = 40

    # ============================================================
    # SHEET 3: DoR Quality
    # ============================================================
    ws_quality = wb.create_sheet("DoR quality")

    quality_headers = [
        "Team", "Coverage (25%)", "Clarity & Specificity (20%)",
        "Measurability (15%)", "Company Standard Alignment (15%)",
        "Industry Best Practices (10%)", "Actionability (10%)",
        "AC Guidance (5%)", "Total Score"
    ]
    for col, h in enumerate(quality_headers, 1):
        cell = ws_quality.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    row = 2
    for team_name in TEAM_MAPPING.keys():
        if team_name in NO_DOR_TEAMS:
            continue
        if team_name not in quality_scores:
            continue
        scores = quality_scores[team_name]
        ws_quality.cell(row=row, column=1, value=team_name).border = thin_border
        ws_quality.cell(row=row, column=2, value=scores["coverage"]).border = thin_border
        ws_quality.cell(row=row, column=3, value=scores["clarity"]).border = thin_border
        ws_quality.cell(row=row, column=4, value=scores["measurability"]).border = thin_border
        ws_quality.cell(row=row, column=5, value=scores["company_alignment"]).border = thin_border
        ws_quality.cell(row=row, column=6, value=scores["industry_alignment"]).border = thin_border
        ws_quality.cell(row=row, column=7, value=scores["actionability"]).border = thin_border
        ws_quality.cell(row=row, column=8, value=scores["ac_guidance"]).border = thin_border

        total_cell = ws_quality.cell(row=row, column=9, value=scores["total"])
        total_cell.border = thin_border
        total_cell.font = Font(bold=True)

        # Color code total
        if scores["total"] >= 70:
            total_cell.fill = pass_fill
        elif scores["total"] >= 40:
            total_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            total_cell.fill = fail_fill

        row += 1

    # Column widths for quality sheet
    for col in range(1, 10):
        ws_quality.column_dimensions[get_column_letter(col)].width = 22

    # Save
    report_path = os.path.join(OUTPUT_DIR, "Report-DoR.xlsx")
    wb.save(report_path)
    print(f"Report saved: {report_path}")
    return report_path


def update_teams_json(team_issues, compliance_results):
    """Update teams.json with Jira and analysis metadata."""
    teams_json_path = os.path.join(OUTPUT_DIR, "teams.json")
    with open(teams_json_path, "r", encoding="utf-8") as f:
        teams_data = json.load(f)

    for team in teams_data["teams"]:
        team_name = team["name"]
        issues = team_issues.get(team_name, [])
        team["jira_issues_count"] = len(issues)
        team["jira_query_date"] = "2026-06-18"
        team["jira_file"] = f"{TEAM_KEBAB[team_name]}-jira.json" if issues else None

        if team_name in compliance_results and compliance_results[team_name]:
            results = compliance_results[team_name]
            passing = sum(1 for r in results if r["compliance"] == "Pass")
            team["dor_compliance_pct"] = round((passing / len(results) * 100), 1) if results else None
        elif team_name in NO_DOR_TEAMS:
            team["dor_compliance_pct"] = None
        else:
            team["dor_compliance_pct"] = None

    # Add metadata
    teams_data["metadata"]["jira_scan_date"] = "2026-06-18"
    teams_data["metadata"]["report_file"] = "Report-DoR.xlsx"

    with open(teams_json_path, "w", encoding="utf-8") as f:
        json.dump(teams_data, f, indent=2, ensure_ascii=False)
    print(f"Updated teams.json")


def generate_analysis_summary(team_issues, compliance_results, quality_scores):
    """Generate DOR_ANALYSIS_SUMMARY.md."""
    lines = []
    lines.append("# DoR Analysis Summary")
    lines.append(f"**Scan Date:** 2026-06-18")
    lines.append(f"**Teams Analyzed:** {len(TEAM_MAPPING)}")
    lines.append("")

    # KPIs
    total_teams = len(TEAM_MAPPING)
    teams_with_dor = total_teams - len(NO_DOR_TEAMS)
    total_tasks = sum(len(r) for r in compliance_results.values())
    total_passing = sum(sum(1 for i in r if i["compliance"] == "Pass") for r in compliance_results.values())
    pct_passing = round((total_passing / total_tasks * 100), 1) if total_tasks > 0 else 0

    lines.append("## Key Performance Indicators")
    lines.append(f"- Teams with DoR defined: {teams_with_dor}/{total_teams} ({round(teams_with_dor/total_teams*100, 1)}%)")
    lines.append(f"- Total active tasks analyzed: {total_tasks}")
    lines.append(f"- Tasks fitting DoR: {total_passing}/{total_tasks} ({pct_passing}%)")
    lines.append("")

    # Per-team summary
    lines.append("## Per-Team Results")
    lines.append("")
    lines.append("| Team | DoR | Active Tasks | DoR Compliance | DoR Quality Score |")
    lines.append("|------|-----|-------------|----------------|-------------------|")

    for team_name in TEAM_MAPPING.keys():
        has_dor = "Yes" if team_name not in NO_DOR_TEAMS else "No"
        issue_count = len(team_issues.get(team_name, []))

        if team_name in compliance_results and compliance_results[team_name]:
            results = compliance_results[team_name]
            passing = sum(1 for r in results if r["compliance"] == "Pass")
            pct = f"{round(passing / len(results) * 100, 1)}%"
        else:
            pct = "N/A"

        quality = quality_scores.get(team_name, {}).get("total", "N/A")
        if isinstance(quality, int):
            quality = f"{quality}/100"

        lines.append(f"| {team_name} | {has_dor} | {issue_count} | {pct} | {quality} |")

    lines.append("")
    lines.append("## DoR Quality Assessment")
    lines.append("")
    lines.append("Scoring dimensions (weighted):")
    lines.append("- Coverage (25%): How many of 10 essential DoR areas are addressed")
    lines.append("- Clarity & Specificity (20%): How concrete and unambiguous criteria are")
    lines.append("- Measurability (15%): Whether criteria have clear pass/fail checks")
    lines.append("- Company Standard Alignment (15%): Alignment with company DoR guidance")
    lines.append("- Industry Best Practices (10%): Adherence to INVEST, Scrum Guide, SAFe")
    lines.append("- Actionability (10%): Whether DoR drives specific behaviors")
    lines.append("- AC Guidance (5%): How well DoR addresses acceptance criteria quality")
    lines.append("")

    # Top and bottom performers
    sorted_quality = sorted(
        [(k, v["total"]) for k, v in quality_scores.items() if v["total"] > 0],
        key=lambda x: x[1], reverse=True
    )
    if sorted_quality:
        lines.append("### Top Performers")
        for name, score in sorted_quality[:3]:
            lines.append(f"- {name}: {score}/100")
        lines.append("")
        lines.append("### Needs Improvement")
        for name, score in sorted_quality[-3:]:
            lines.append(f"- {name}: {score}/100")

    summary_path = os.path.join(OUTPUT_DIR, "DOR_ANALYSIS_SUMMARY.md")
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"Generated: {summary_path}")


def generate_html_report(team_issues, compliance_results, quality_scores):
    """Generate a self-contained HTML dashboard page presenting all DoR data.

    Matches the Report-DoR.xlsx schema:
    - Summary: Team | DoR | Jira Tasks in progress | % Tasks fitting DoR
    - DoR Compliance: only Fail items (Team | Issue Key | Type | URL | Title | Status | Assignee | Note)
    - DoR quality: Team | Coverage (0-100) | Note (missing criteria)
    """
    scan_date = datetime.now().strftime("%Y-%m-%d")
    total_teams = len(TEAM_MAPPING)
    teams_with_dor = total_teams - len(NO_DOR_TEAMS)
    pct_teams_with_dor = round((teams_with_dor / total_teams) * 100, 1)

    total_tasks = sum(len(r) for r in compliance_results.values())
    total_passing = sum(
        sum(1 for i in r if i["compliance"] == "Pass")
        for r in compliance_results.values()
    )
    total_failing = total_tasks - total_passing
    pct_tasks_fitting = round((total_passing / total_tasks * 100), 1) if total_tasks > 0 else 0

    # Average DoR quality (coverage score from quality_data)
    quality_totals = [v["total"] for v in quality_scores.values() if v.get("total", 0) > 0]
    avg_quality = round(sum(quality_totals) / len(quality_totals)) if quality_totals else 0

    # Build team rows for summary table (matches xlsx Summary sheet)
    summary_rows = ""
    for team_name in TEAM_MAPPING.keys():
        has_dor = team_name not in NO_DOR_TEAMS
        issues = team_issues.get(team_name, [])
        issue_count = len(issues)

        if team_name in compliance_results and compliance_results[team_name]:
            results = compliance_results[team_name]
            passing = sum(1 for r in results if r["compliance"] == "Pass")
            pct = round((passing / len(results) * 100))
            pct_str = f"{pct}%"
            pct_class = "good" if pct >= 70 else ("warn" if pct >= 40 else "bad")
        elif not has_dor:
            pct_str = "-"
            pct_class = "na"
        elif issue_count == 0:
            pct_str = "-"
            pct_class = "na"
        else:
            pct_str = "-"
            pct_class = "na"

        dor_badge = '<span class="badge badge-yes">Yes</span>' if has_dor else '<span class="badge badge-no">No</span>'

        summary_rows += f"""        <tr>
          <td class="team-name">{team_name}</td>
          <td class="center">{dor_badge}</td>
          <td class="center">{issue_count}</td>
          <td class="center"><span class="score {pct_class}">{pct_str}</span></td>
        </tr>\n"""

    # Build compliance rows - ONLY Fail items
    compliance_rows = ""
    for team_name in TEAM_MAPPING.keys():
        if team_name not in compliance_results:
            continue
        for r in compliance_results[team_name]:
            if r["compliance"] == "Pass":
                continue
            compliance_rows += f"""        <tr>
          <td class="team-name">{r["team"]}</td>
          <td><a href="{r["url"]}" target="_blank">{r["key"]}</a></td>
          <td>{r["issuetype"]}</td>
          <td class="title-col">{r["summary"]}</td>
          <td>{r["status"]}</td>
          <td>{r["assignee"]}</td>
          <td class="note-col">{r["note"]}</td>
        </tr>\n"""

    # Build quality rows (matches xlsx DoR quality sheet: Team | Coverage | Note)
    sorted_quality = sorted(
        [(k, v) for k, v in quality_scores.items() if v.get("total", 0) > 0],
        key=lambda x: x[1]["total"], reverse=True
    )
    quality_rows = ""
    for team_name, scores in sorted_quality:
        total = scores["total"]
        bar_class = "good" if total >= 70 else ("warn" if total >= 40 else "bad")
        # Build note from missing areas
        missing = []
        if scores["coverage"] < 50:
            missing.append("Coverage")
        if scores["clarity"] < 50:
            missing.append("Clarity")
        if scores["measurability"] < 50:
            missing.append("Measurability")
        if scores["company_alignment"] < 50:
            missing.append("Company Alignment")
        if scores["industry_alignment"] < 50:
            missing.append("Industry Practices")
        if scores["actionability"] < 50:
            missing.append("Actionability")
        if scores["ac_guidance"] < 50:
            missing.append("AC Guidance")
        note = f"Missing: {', '.join(missing)}" if missing else "All standard criteria covered"

        quality_rows += f"""        <tr>
          <td class="team-name">{team_name}</td>
          <td class="center"><strong class="score {bar_class}">{total}</strong></td>
          <td>
            <div class="bar-container">
              <div class="bar {bar_class}" style="width: {total}%"></div>
            </div>
          </td>
          <td class="note-col">{note}</td>
        </tr>\n"""

    # KPI color classes
    kpi_teams_class = "good" if pct_teams_with_dor >= 70 else ("warn" if pct_teams_with_dor >= 40 else "bad")
    kpi_tasks_class = "good" if pct_tasks_fitting >= 70 else ("warn" if pct_tasks_fitting >= 40 else "bad")
    kpi_quality_class = "good" if avg_quality >= 70 else ("warn" if avg_quality >= 40 else "bad")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>DoR Scanner Report - SRPOL Teams</title>
  <style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
      background: #f0f2f5;
      color: #1a1a2e;
      line-height: 1.5;
      padding: 24px;
    }}
    .container {{ max-width: 1400px; margin: 0 auto; }}
    header {{
      background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
      color: #fff;
      padding: 32px 40px;
      border-radius: 12px;
      margin-bottom: 24px;
      box-shadow: 0 4px 20px rgba(0,0,0,0.15);
    }}
    header h1 {{ font-size: 28px; font-weight: 700; margin-bottom: 4px; }}
    header .subtitle {{ font-size: 14px; opacity: 0.8; }}
    .kpi-grid {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 20px;
      margin-bottom: 32px;
    }}
    .kpi-card {{
      background: #fff;
      border-radius: 10px;
      padding: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.06);
      border-top: 4px solid #ccc;
    }}
    .kpi-card.good {{ border-top-color: #27ae60; }}
    .kpi-card.warn {{ border-top-color: #f39c12; }}
    .kpi-card.bad {{ border-top-color: #e74c3c; }}
    .kpi-card .kpi-label {{
      font-size: 13px;
      text-transform: uppercase;
      letter-spacing: 0.5px;
      color: #666;
      margin-bottom: 8px;
      font-weight: 600;
    }}
    .kpi-card .kpi-value {{
      font-size: 36px;
      font-weight: 800;
    }}
    .kpi-card .kpi-detail {{
      font-size: 12px;
      color: #888;
      margin-top: 4px;
    }}
    .kpi-value.good {{ color: #27ae60; }}
    .kpi-value.warn {{ color: #f39c12; }}
    .kpi-value.bad {{ color: #e74c3c; }}
    .section {{
      background: #fff;
      border-radius: 10px;
      padding: 24px;
      margin-bottom: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.06);
    }}
    .section h2 {{
      font-size: 18px;
      font-weight: 700;
      margin-bottom: 16px;
      padding-bottom: 12px;
      border-bottom: 2px solid #f0f2f5;
      color: #1a1a2e;
    }}
    .section h2 .count {{
      font-size: 14px;
      font-weight: 400;
      color: #888;
      margin-left: 8px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }}
    thead th {{
      background: #1a1a2e;
      color: #fff;
      padding: 10px 12px;
      text-align: left;
      font-weight: 600;
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: 0.3px;
      white-space: nowrap;
    }}
    thead th.center {{ text-align: center; }}
    tbody td {{
      padding: 10px 12px;
      border-bottom: 1px solid #eee;
      vertical-align: middle;
    }}
    tbody tr:hover {{ background: #f8f9fc; }}
    tbody tr:last-child td {{ border-bottom: none; }}
    .center {{ text-align: center; }}
    .team-name {{ font-weight: 600; white-space: nowrap; }}
    .title-col {{ max-width: 350px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
    .note-col {{ font-size: 12px; color: #555; }}
    .badge {{
      display: inline-block;
      padding: 3px 10px;
      border-radius: 12px;
      font-size: 11px;
      font-weight: 700;
      text-transform: uppercase;
    }}
    .badge-yes {{ background: #d4edda; color: #155724; }}
    .badge-no {{ background: #f8d7da; color: #721c24; }}
    .score {{ font-weight: 700; }}
    .score.good {{ color: #27ae60; }}
    .score.warn {{ color: #f39c12; }}
    .score.bad {{ color: #e74c3c; }}
    .score.na {{ color: #aaa; }}
    .bar-container {{
      width: 100%;
      height: 20px;
      background: #eee;
      border-radius: 4px;
      overflow: hidden;
    }}
    .bar {{
      height: 100%;
      border-radius: 4px;
    }}
    .bar.good {{ background: linear-gradient(90deg, #27ae60, #2ecc71); }}
    .bar.warn {{ background: linear-gradient(90deg, #f39c12, #f1c40f); }}
    .bar.bad {{ background: linear-gradient(90deg, #e74c3c, #c0392b); }}
    a {{ color: #2980b9; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    footer {{
      text-align: center;
      padding: 16px;
      font-size: 12px;
      color: #999;
      margin-top: 8px;
    }}
    @media (max-width: 768px) {{
      .kpi-grid {{ grid-template-columns: 1fr; }}
      body {{ padding: 12px; }}
      header {{ padding: 20px; }}
      .section {{ padding: 16px; }}
      table {{ font-size: 11px; }}
    }}
  </style>
</head>
<body>
  <div class="container">
    <header>
      <h1>DoR Scanner Report</h1>
      <div class="subtitle">SRPOL Teams - Definition of Ready Analysis | Scan Date: {scan_date} | Teams: {total_teams}</div>
    </header>

    <div class="kpi-grid">
      <div class="kpi-card {kpi_teams_class}">
        <div class="kpi-label">Teams with DoR Defined</div>
        <div class="kpi-value {kpi_teams_class}">{pct_teams_with_dor}%</div>
        <div class="kpi-detail">{teams_with_dor} of {total_teams} teams</div>
      </div>
      <div class="kpi-card {kpi_tasks_class}">
        <div class="kpi-label">Tasks Fitting DoR</div>
        <div class="kpi-value {kpi_tasks_class}">{pct_tasks_fitting}%</div>
        <div class="kpi-detail">{total_passing} of {total_tasks} tasks compliant</div>
      </div>
      <div class="kpi-card {kpi_quality_class}">
        <div class="kpi-label">DoR Quality Level</div>
        <div class="kpi-value {kpi_quality_class}">{avg_quality}/100</div>
        <div class="kpi-detail">Across {len(quality_totals)} assessed teams</div>
      </div>
    </div>

    <div class="section">
      <h2>Team Overview</h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th class="center">DoR</th>
            <th class="center">Jira Tasks in progress</th>
            <th class="center">% Tasks fitting DoR</th>
          </tr>
        </thead>
        <tbody>
{summary_rows}        </tbody>
      </table>
    </div>

    <div class="section">
      <h2>DoR Compliance - Failed Issues<span class="count">({total_failing} issues)</span></h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th>Issue Key</th>
            <th>Issue Type</th>
            <th>Title</th>
            <th>Status</th>
            <th>Assignee</th>
            <th>Note</th>
          </tr>
        </thead>
        <tbody>
{compliance_rows}        </tbody>
      </table>
    </div>

    <div class="section">
      <h2>DoR Quality Assessment</h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th class="center">Coverage</th>
            <th>Score</th>
            <th>Note</th>
          </tr>
        </thead>
        <tbody>
{quality_rows}        </tbody>
      </table>
    </div>

    <footer>
      DoR Scanner v2.0 | Generated by wow-dor-scanner skill | Data source: SRPOL Teams Confluence
    </footer>
  </div>
</body>
</html>"""

    report_path = os.path.join(OUTPUT_DIR, "Report-DoR.html")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML report saved: {report_path}")
    return report_path


def main():
    print("=" * 60)
    print("DoR SCANNER - PROCESSING")
    print("=" * 60)

    # Step 1-2: Load and filter Jira data
    print("\n[STEP 1-2] Loading Jira data and filtering by team...")
    team_issues = load_jira_data()
    for team, issues in team_issues.items():
        print(f"  {team}: {len(issues)} issues")

    # Step 3: Save Jira files
    print("\n[STEP 3] Saving Jira files...")
    save_jira_files(team_issues)

    # Step 4: Load DoR files and run compliance analysis
    print("\n[STEP 4] Loading DoR files and running compliance analysis...")
    dor_content = load_dor_files()
    compliance_results = run_compliance_analysis(team_issues, dor_content)
    for team, results in compliance_results.items():
        if results:
            passing = sum(1 for r in results if r["compliance"] == "Pass")
            print(f"  {team}: {passing}/{len(results)} passing")

    # Step 5-6: DoR Quality Assessment
    print("\n[STEP 5-6] Assessing DoR quality...")
    quality_scores = assess_dor_quality(dor_content)
    for team, scores in quality_scores.items():
        if scores.get("total", 0) > 0:
            print(f"  {team}: {scores['total']}/100")

    # Step 5: Generate Excel Report
    print("\n[STEP 5] Generating Report-DoR.xlsx...")
    generate_excel_report(team_issues, compliance_results, quality_scores)

    # Step 7: Update teams.json
    print("\n[STEP 7] Updating teams.json...")
    update_teams_json(team_issues, compliance_results)

    # Step 8: Generate summary
    print("\n[STEP 8] Generating DOR_ANALYSIS_SUMMARY.md...")
    generate_analysis_summary(team_issues, compliance_results, quality_scores)

    # Step 9: Generate HTML dashboard
    print("\n[STEP 9] Generating Report-DoR.html...")
    generate_html_report(team_issues, compliance_results, quality_scores)

    print("\n" + "=" * 60)
    print("DONE - All steps completed successfully")
    print("=" * 60)


if __name__ == "__main__":
    main()
