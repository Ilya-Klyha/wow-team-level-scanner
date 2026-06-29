"""
DoR Scanner - Report Generation for 2026-06-26 scan
No active Jira issues found; generates DoR quality assessment and reports.
"""
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"C:\Users\i.klyha\Desktop\Claude\wow-scanner-tool\assets\teams\20260626 11-18"

TEAM_MAPPING = {
    "Abyss": "PE - WAW - Abyss",
    "Radium": "AE - WAW - Radium",
    "Europium": "AP - WAW - Europium",
    "Copernicium": "AE - WAW - Copernicium",
    "Mouflons": "AS - WAW - Mouflons",
    "Wolves": "AS - WAW - Wolves",
    "Polonium UF": "AS - WAW - Polonium UF",
    "Bigos": "AS - WAW - Bigos",
    "Capybaras": "AS - WAW - Capybaras",
    "ML Serving Sturgeons": "T - WAW - ML Sturgeons",
    "ML Platform Pandas": "T - WAW - ML Pandas",
    "EP Core": "T - WAW - EP Core",
    "Zurek": "Zurek",
    "Igni": "AP - WAW - Igni",
    "SRE": "T - WAW - Embedded SREs SRPOL",
}

TEAM_KEBAB = {
    "Abyss": "abyss",
    "Radium": "radium",
    "Europium": "europium",
    "Copernicium": "copernicium",
    "Mouflons": "mouflons",
    "Wolves": "wolves",
    "Polonium UF": "polonium-uf",
    "Bigos": "bigos",
    "Capybaras": "capybaras",
    "ML Serving Sturgeons": "ml-serving-sturgeons",
    "ML Platform Pandas": "ml-platform-pandas",
    "EP Core": "ep-core",
    "Zurek": "zurek",
    "Igni": "igni",
    "SRE": "sre",
}

NO_DOR_TEAMS = ["ML Serving Sturgeons", "SRE"]

DOR_SOURCES = {
    "Abyss": "direct",
    "Radium": "direct",
    "Europium": "direct",
    "Copernicium": "direct",
    "Mouflons": "direct",
    "Wolves": "direct",
    "Polonium UF": "linked_page",
    "Bigos": "linked_page",
    "Capybaras": "direct",
    "ML Serving Sturgeons": None,
    "ML Platform Pandas": "linked_page",
    "EP Core": "direct",
    "Zurek": "direct",
    "Igni": "direct",
    "SRE": None,
}


def save_jira_txt_files():
    """Save Jira TXT files (all empty since no issues found)."""
    for team_name in TEAM_MAPPING.keys():
        kebab = TEAM_KEBAB[team_name]
        txt_path = os.path.join(OUTPUT_DIR, f"{kebab}-jira.txt")
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(f"JIRA ACTIVE ISSUES - {team_name}\n")
            f.write("=" * 60 + "\n")
            f.write("Query Date: 2026-06-26\n")
            f.write("Total Issues: 0\n")
            f.write("=" * 60 + "\n\n")
            f.write("No active issues found for this team.\n")
            f.write("Note: Both primary (sprint in openSprints()) and fallback queries returned 0 results.\n")
            f.write("The customfield_10114 Team field may require ID-based JQL filtering.\n")
    print(f"Saved Jira TXT files for all {len(TEAM_MAPPING)} teams")


def load_dor_files():
    """Load DoR text files for each team."""
    dor_content = {}
    for team_name in TEAM_MAPPING.keys():
        if team_name in NO_DOR_TEAMS:
            continue
        kebab = TEAM_KEBAB[team_name]
        dor_path = os.path.join(OUTPUT_DIR, f"{kebab}-dor.txt")
        if os.path.exists(dor_path):
            with open(dor_path, "r", encoding="utf-8") as f:
                dor_content[team_name] = f.read()
    return dor_content


def assess_dor_quality(dor_content):
    """Score each team's DoR 0-100 across 7 dimensions."""
    quality_scores = {}

    for team_name, content in dor_content.items():
        if not content.strip() or "not found" in content.lower():
            quality_scores[team_name] = {
                "coverage": 0, "clarity": 0, "measurability": 0,
                "company_alignment": 0, "industry_alignment": 0,
                "actionability": 0, "ac_guidance": 0, "total": 0
            }
            continue

        content_lower = content.lower()

        # 1. COVERAGE (25%) - 10 essential areas
        coverage_areas = 0
        if any(kw in content_lower for kw in ["user story", "requirement", "description", "clearly defined", "clear problem"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["acceptance criteria", "ac ", "ac:"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["estimat", "story point", "sizing", "size"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["dependenc", "blocked", "blocker"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["design", "mockup", "wireframe", "technical spec", "techspec", "ux", "ui", "figma"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["scope", "sprint fit", "fits in sprint", "single sprint", "fit within a sprint"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["risk", "blocker", "impediment"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["product owner", "po ", "stakeholder", "approval", "reviewed"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["feasib", "technical feasibility", "spike", "investigation", "investigated"]):
            coverage_areas += 1
        if any(kw in content_lower for kw in ["test", "testing", "qa", "quality", "testcase", "testable"]):
            coverage_areas += 1
        coverage_score = int((coverage_areas / 10) * 100)

        # 2. CLARITY (20%)
        specificity_indicators = 0
        total_checks = 5
        if any(kw in content_lower for kw in ["must", "shall", "required", "mandatory", "should be"]):
            specificity_indicators += 1
        if any(kw in content_lower for kw in ["defined", "documented", "written", "specified", "provided"]):
            specificity_indicators += 1
        if any(kw in content_lower for kw in ["verif", "check", "confirm", "validated", "approved"]):
            specificity_indicators += 1
        vague_terms = sum(1 for kw in ["appropriate", "sufficient", "if needed", "as needed", "if applicable"]
                         if kw in content_lower)
        if vague_terms <= 1:
            specificity_indicators += 1
        if len(content) > 200:
            specificity_indicators += 1
        clarity_score = int((specificity_indicators / total_checks) * 100)

        # 3. MEASURABILITY (15%)
        measurability_indicators = 0
        if any(kw in content_lower for kw in ["pass/fail", "yes/no", "checklist", "checkbox"]):
            measurability_indicators += 2
        if any(kw in content_lower for kw in ["verified", "confirmed", "approved", "signed off"]):
            measurability_indicators += 1
        if any(kw in content_lower for kw in ["all team", "team agrees", "consensus", "reviewed by", "team agree"]):
            measurability_indicators += 1
        measurability_score = min(100, int((measurability_indicators / 3) * 100))

        # 4. COMPANY STANDARD ALIGNMENT (15%)
        company_indicators = 0
        if "acceptance criteria" in content_lower or "ac " in content_lower:
            company_indicators += 1
        if "dependenc" in content_lower:
            company_indicators += 1
        if "scope" in content_lower or "priority" in content_lower:
            company_indicators += 1
        if "confidence" in content_lower or "clarity" in content_lower:
            company_indicators += 1
        company_score = int((company_indicators / 4) * 100)

        # 5. INDUSTRY BEST PRACTICES (10%)
        industry_indicators = 0
        if any(kw in content_lower for kw in ["independent", "negotiable", "valuable", "estimable", "small", "testable", "invest"]):
            industry_indicators += 1
        if any(kw in content_lower for kw in ["sprint", "iteration", "increment"]):
            industry_indicators += 1
        if any(kw in content_lower for kw in ["refinement", "grooming", "backlog"]):
            industry_indicators += 1
        industry_score = int((industry_indicators / 3) * 100)

        # 6. ACTIONABILITY (10%)
        actionability_indicators = 0
        if any(kw in content_lower for kw in ["responsible", "owner", "who", "role", "reporter"]):
            actionability_indicators += 1
        if any(kw in content_lower for kw in ["before", "prior to", "gate", "block", "not ready"]):
            actionability_indicators += 1
        if any(kw in content_lower for kw in ["process", "workflow", "step", "procedure", "planning"]):
            actionability_indicators += 1
        actionability_score = int((actionability_indicators / 3) * 100)

        # 7. AC GUIDANCE (5%)
        ac_indicators = 0
        if any(kw in content_lower for kw in ["given/when/then", "given when then", "gherkin"]):
            ac_indicators += 2
        if any(kw in content_lower for kw in ["format", "structured", "template", "form of"]):
            ac_indicators += 1
        if "acceptance criteria" in content_lower and any(kw in content_lower for kw in ["specific", "measurable", "testable"]):
            ac_indicators += 1
        ac_score = min(100, int((ac_indicators / 3) * 100))

        # Weighted total
        total = int(
            coverage_score * 0.25 +
            clarity_score * 0.20 +
            measurability_score * 0.15 +
            company_score * 0.15 +
            industry_score * 0.10 +
            actionability_score * 0.10 +
            ac_score * 0.05
        )

        quality_scores[team_name] = {
            "coverage": coverage_score,
            "clarity": clarity_score,
            "measurability": measurability_score,
            "company_alignment": company_score,
            "industry_alignment": industry_score,
            "actionability": actionability_score,
            "ac_guidance": ac_score,
            "total": total,
        }

    return quality_scores


def generate_excel_report(quality_scores):
    """Generate Report-DoR.xlsx with 3 sheets."""
    wb = Workbook()

    header_font = Font(bold=True, size=12)
    kpi_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    total_teams = len(TEAM_MAPPING)
    teams_with_dor = total_teams - len(NO_DOR_TEAMS)
    pct_teams_with_dor = round((teams_with_dor / total_teams) * 100, 1)

    # ============================================================
    # SHEET 1: Summary
    # ============================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # KPI rows 1-2
    ws_summary["A1"] = "DoR Scanner Report - SRPOL Teams"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["C1"] = f"% Teams with DoR: {pct_teams_with_dor}%"
    ws_summary["C1"].font = kpi_font
    ws_summary["A2"] = "Scan Date: 2026-06-26"
    ws_summary["C2"] = "% Tasks fitting DoR: N/A (no active issues found)"
    ws_summary["A3"] = ""

    # Table from row 4
    row = 4
    headers = ["Team", "DoR Defined", "DoR Source", "Jira Tasks In Progress", "% Tasks Fitting DoR"]
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row = 5
    for team_name in TEAM_MAPPING.keys():
        has_dor = "Yes" if team_name not in NO_DOR_TEAMS else "No"
        source = DOR_SOURCES.get(team_name) or "N/A"
        pct_str = "N/A (no active issues)"

        ws_summary.cell(row=row, column=1, value=team_name).border = thin_border
        c2 = ws_summary.cell(row=row, column=2, value=has_dor)
        c2.border = thin_border
        if has_dor == "Yes":
            c2.fill = pass_fill
        else:
            c2.fill = fail_fill
        ws_summary.cell(row=row, column=3, value=source).border = thin_border
        ws_summary.cell(row=row, column=4, value=0).border = thin_border
        ws_summary.cell(row=row, column=5, value=pct_str).border = thin_border
        row += 1

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 14
    ws_summary.column_dimensions["C"].width = 40
    ws_summary.column_dimensions["D"].width = 22
    ws_summary.column_dimensions["E"].width = 28

    # ============================================================
    # SHEET 2: DoR Compliance
    # ============================================================
    ws_compliance = wb.create_sheet("DoR Compliance")
    headers = ["Team", "Issue Key", "Issue Type", "URL", "Title", "Status", "Assignee", "DoR Compliance", "Note"]
    for col, h in enumerate(headers, 1):
        cell = ws_compliance.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    ws_compliance.cell(row=2, column=1, value="(No active issues found for any team)")
    ws_compliance.cell(row=2, column=8, value="N/A")
    ws_compliance.cell(row=2, column=9, value="Jira queries returned 0 results for all 15 teams. Teams may be between sprints or customfield_10114 requires ID-based filtering.")

    ws_compliance.column_dimensions["A"].width = 20
    ws_compliance.column_dimensions["B"].width = 14
    ws_compliance.column_dimensions["C"].width = 12
    ws_compliance.column_dimensions["D"].width = 45
    ws_compliance.column_dimensions["E"].width = 50
    ws_compliance.column_dimensions["F"].width = 14
    ws_compliance.column_dimensions["G"].width = 22
    ws_compliance.column_dimensions["H"].width = 16
    ws_compliance.column_dimensions["I"].width = 55

    # ============================================================
    # SHEET 3: DoR Quality
    # ============================================================
    ws_quality = wb.create_sheet("DoR quality")

    # KPI row 1
    valid_scores = [s["total"] for s in quality_scores.values() if s["total"] > 0]
    avg_quality = round(sum(valid_scores) / len(valid_scores)) if valid_scores else 0
    ws_quality["A1"] = "DoR Quality Assessment"
    ws_quality["A1"].font = Font(bold=True, size=14)
    ws_quality["C1"] = f"Average Quality Score: {avg_quality}/100"
    ws_quality["C1"].font = Font(bold=True, size=12)

    # Table from row 3
    quality_headers = [
        "Team", "Coverage (25%)", "Clarity & Specificity (20%)",
        "Measurability (15%)", "Company Standard (15%)",
        "Industry Practices (10%)", "Actionability (10%)",
        "AC Guidance (5%)", "Total Score"
    ]
    for col, h in enumerate(quality_headers, 1):
        cell = ws_quality.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row = 4
    for team_name in TEAM_MAPPING.keys():
        if team_name in NO_DOR_TEAMS:
            continue
        if team_name not in quality_scores:
            continue
        scores = quality_scores[team_name]
        ws_quality.cell(row=row, column=1, value=team_name).border = thin_border
        ws_quality.cell(row=row, column=2, value=scores["coverage"]).border = thin_border
        ws_quality.cell(row=row, column=3, value=scores["clarity"]).border = thin_border
        ws_quality.cell(row=row, column=4, value=scores["measurability"]).border = thin_border
        ws_quality.cell(row=row, column=5, value=scores["company_alignment"]).border = thin_border
        ws_quality.cell(row=row, column=6, value=scores["industry_alignment"]).border = thin_border
        ws_quality.cell(row=row, column=7, value=scores["actionability"]).border = thin_border
        ws_quality.cell(row=row, column=8, value=scores["ac_guidance"]).border = thin_border

        total_cell = ws_quality.cell(row=row, column=9, value=scores["total"])
        total_cell.border = thin_border
        total_cell.font = Font(bold=True)
        if scores["total"] >= 70:
            total_cell.fill = pass_fill
        elif scores["total"] >= 40:
            total_cell.fill = warn_fill
        else:
            total_cell.fill = fail_fill
        row += 1

    for col in range(1, 10):
        ws_quality.column_dimensions[get_column_letter(col)].width = 22

    report_path = os.path.join(OUTPUT_DIR, "Report-DoR.xlsx")
    wb.save(report_path)
    print(f"Excel report saved: {report_path}")
    return report_path


def generate_html_report(quality_scores):
    """Generate self-contained HTML dashboard."""
    scan_date = "2026-06-26"
    total_teams = len(TEAM_MAPPING)
    teams_with_dor = total_teams - len(NO_DOR_TEAMS)
    pct_teams_with_dor = round((teams_with_dor / total_teams) * 100, 1)

    valid_scores = [s["total"] for s in quality_scores.values() if s["total"] > 0]
    avg_quality = round(sum(valid_scores) / len(valid_scores)) if valid_scores else 0

    # KPI color classes
    kpi_teams_class = "good" if pct_teams_with_dor >= 70 else ("warn" if pct_teams_with_dor >= 40 else "bad")
    kpi_quality_class = "good" if avg_quality >= 70 else ("warn" if avg_quality >= 40 else "bad")

    # Build summary rows
    summary_rows = ""
    for team_name in TEAM_MAPPING.keys():
        has_dor = team_name not in NO_DOR_TEAMS
        source = DOR_SOURCES.get(team_name) or "N/A"
        dor_badge = '<span class="badge badge-yes">Yes</span>' if has_dor else '<span class="badge badge-no">No</span>'

        summary_rows += f"""        <tr>
          <td class="team-name">{team_name}</td>
          <td class="center">{dor_badge}</td>
          <td class="center">{source}</td>
          <td class="center">0</td>
          <td class="center"><span class="score na">-</span></td>
        </tr>\n"""

    # Build quality rows sorted by total score descending
    sorted_quality = sorted(
        [(k, v) for k, v in quality_scores.items() if v.get("total", 0) > 0],
        key=lambda x: x[1]["total"], reverse=True
    )
    quality_rows = ""
    for team_name, scores in sorted_quality:
        total = scores["total"]
        bar_class = "good" if total >= 70 else ("warn" if total >= 40 else "bad")
        missing = []
        if scores["coverage"] < 50:
            missing.append("Coverage")
        if scores["clarity"] < 50:
            missing.append("Clarity")
        if scores["measurability"] < 50:
            missing.append("Measurability")
        if scores["company_alignment"] < 50:
            missing.append("Company Alignment")
        if scores["industry_alignment"] < 50:
            missing.append("Industry Practices")
        if scores["actionability"] < 50:
            missing.append("Actionability")
        if scores["ac_guidance"] < 50:
            missing.append("AC Guidance")
        note = f"Weak areas: {', '.join(missing)}" if missing else "All standard criteria covered"

        quality_rows += f"""        <tr>
          <td class="team-name">{team_name}</td>
          <td class="center"><strong>{scores['coverage']}</strong></td>
          <td class="center"><strong>{scores['clarity']}</strong></td>
          <td class="center"><strong>{scores['measurability']}</strong></td>
          <td class="center"><strong>{scores['company_alignment']}</strong></td>
          <td class="center"><strong>{scores['industry_alignment']}</strong></td>
          <td class="center"><strong>{scores['actionability']}</strong></td>
          <td class="center"><strong>{scores['ac_guidance']}</strong></td>
          <td class="center">
            <strong class="score {bar_class}">{total}</strong>
            <div class="bar-container"><div class="bar {bar_class}" style="width: {total}%"></div></div>
          </td>
          <td class="note-col">{note}</td>
        </tr>\n"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>DoR Scanner Report - SRPOL Teams - {scan_date}</title>
  <style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
      background: #f0f2f5;
      color: #1a1a2e;
      line-height: 1.5;
      padding: 24px;
    }}
    .container {{ max-width: 1400px; margin: 0 auto; }}
    header {{
      background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
      color: #fff;
      padding: 32px 40px;
      border-radius: 12px;
      margin-bottom: 24px;
      box-shadow: 0 4px 20px rgba(0,0,0,0.15);
    }}
    header h1 {{ font-size: 28px; font-weight: 700; margin-bottom: 4px; }}
    header .subtitle {{ font-size: 14px; opacity: 0.8; }}
    .kpi-grid {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 20px;
      margin-bottom: 32px;
    }}
    .kpi-card {{
      background: #fff;
      border-radius: 10px;
      padding: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.06);
      border-top: 4px solid #ccc;
    }}
    .kpi-card.good {{ border-top-color: #27ae60; }}
    .kpi-card.warn {{ border-top-color: #f39c12; }}
    .kpi-card.bad {{ border-top-color: #e74c3c; }}
    .kpi-card.na {{ border-top-color: #95a5a6; }}
    .kpi-card .kpi-label {{
      font-size: 13px;
      text-transform: uppercase;
      letter-spacing: 0.5px;
      color: #666;
      margin-bottom: 8px;
      font-weight: 600;
    }}
    .kpi-card .kpi-value {{
      font-size: 36px;
      font-weight: 800;
    }}
    .kpi-card .kpi-detail {{
      font-size: 12px;
      color: #888;
      margin-top: 4px;
    }}
    .kpi-value.good {{ color: #27ae60; }}
    .kpi-value.warn {{ color: #f39c12; }}
    .kpi-value.bad {{ color: #e74c3c; }}
    .kpi-value.na {{ color: #95a5a6; }}
    .section {{
      background: #fff;
      border-radius: 10px;
      padding: 24px;
      margin-bottom: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.06);
    }}
    .section h2 {{
      font-size: 18px;
      font-weight: 700;
      margin-bottom: 16px;
      padding-bottom: 12px;
      border-bottom: 2px solid #f0f2f5;
      color: #1a1a2e;
    }}
    .section h2 .count {{
      font-size: 14px;
      font-weight: 400;
      color: #888;
      margin-left: 8px;
    }}
    .info-box {{
      background: #f8f9fc;
      border-left: 4px solid #3498db;
      padding: 12px 16px;
      margin-bottom: 16px;
      font-size: 13px;
      color: #555;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }}
    thead th {{
      background: #1a1a2e;
      color: #fff;
      padding: 10px 12px;
      text-align: left;
      font-weight: 600;
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: 0.3px;
      white-space: nowrap;
    }}
    thead th.center {{ text-align: center; }}
    tbody td {{
      padding: 10px 12px;
      border-bottom: 1px solid #eee;
      vertical-align: middle;
    }}
    tbody tr:hover {{ background: #f8f9fc; }}
    tbody tr:last-child td {{ border-bottom: none; }}
    .center {{ text-align: center; }}
    .team-name {{ font-weight: 600; white-space: nowrap; }}
    .note-col {{ font-size: 12px; color: #555; max-width: 300px; }}
    .badge {{
      display: inline-block;
      padding: 3px 10px;
      border-radius: 12px;
      font-size: 11px;
      font-weight: 700;
      text-transform: uppercase;
    }}
    .badge-yes {{ background: #d4edda; color: #155724; }}
    .badge-no {{ background: #f8d7da; color: #721c24; }}
    .score {{ font-weight: 700; }}
    .score.good {{ color: #27ae60; }}
    .score.warn {{ color: #f39c12; }}
    .score.bad {{ color: #e74c3c; }}
    .score.na {{ color: #aaa; }}
    .bar-container {{
      width: 100%;
      height: 8px;
      background: #eee;
      border-radius: 4px;
      overflow: hidden;
      margin-top: 4px;
    }}
    .bar {{
      height: 100%;
      border-radius: 4px;
    }}
    .bar.good {{ background: linear-gradient(90deg, #27ae60, #2ecc71); }}
    .bar.warn {{ background: linear-gradient(90deg, #f39c12, #f1c40f); }}
    .bar.bad {{ background: linear-gradient(90deg, #e74c3c, #c0392b); }}
    footer {{
      text-align: center;
      padding: 16px;
      font-size: 12px;
      color: #999;
      margin-top: 8px;
    }}
    @media (max-width: 768px) {{
      .kpi-grid {{ grid-template-columns: 1fr; }}
      body {{ padding: 12px; }}
    }}
  </style>
</head>
<body>
  <div class="container">
    <header>
      <h1>DoR Scanner Report</h1>
      <div class="subtitle">SRPOL Teams - Definition of Ready Analysis | Scan Date: {scan_date} | Teams: {total_teams}</div>
    </header>

    <div class="kpi-grid">
      <div class="kpi-card {kpi_teams_class}">
        <div class="kpi-label">Teams with DoR Defined</div>
        <div class="kpi-value {kpi_teams_class}">{pct_teams_with_dor}%</div>
        <div class="kpi-detail">{teams_with_dor} of {total_teams} teams</div>
      </div>
      <div class="kpi-card na">
        <div class="kpi-label">Tasks Fitting DoR</div>
        <div class="kpi-value na">N/A</div>
        <div class="kpi-detail">No active issues found in current sprint</div>
      </div>
      <div class="kpi-card {kpi_quality_class}">
        <div class="kpi-label">Avg DoR Quality Score</div>
        <div class="kpi-value {kpi_quality_class}">{avg_quality}/100</div>
        <div class="kpi-detail">Across {len(valid_scores)} assessed teams</div>
      </div>
    </div>

    <div class="section">
      <h2>Team Overview<span class="count">({total_teams} teams)</span></h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th class="center">DoR</th>
            <th class="center">DoR Source</th>
            <th class="center">Jira Tasks in Progress</th>
            <th class="center">% Tasks Fitting DoR</th>
          </tr>
        </thead>
        <tbody>
{summary_rows}        </tbody>
      </table>
    </div>

    <div class="section">
      <h2>DoR Compliance<span class="count">(0 issues analyzed)</span></h2>
      <div class="info-box">
        No active issues found for any of the 15 teams. Jira queries with status IN ("In Progress", "Code Review", "In Development")
        returned 0 results both with sprint filter and without. The customfield_10114 (Team) field uses object-based values
        that may require ID-based JQL filtering rather than text string matching.
      </div>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th>Issue Key</th>
            <th>Issue Type</th>
            <th>URL</th>
            <th>Title</th>
            <th>Status</th>
            <th>Assignee</th>
            <th>DoR Compliance</th>
            <th>Note</th>
          </tr>
        </thead>
        <tbody>
          <tr><td colspan="9" style="text-align:center; color:#888; padding:20px;">No data available</td></tr>
        </tbody>
      </table>
    </div>

    <div class="section">
      <h2>DoR Quality Assessment<span class="count">(avg: {avg_quality}/100)</span></h2>
      <table>
        <thead>
          <tr>
            <th>Team</th>
            <th class="center">Coverage (25%)</th>
            <th class="center">Clarity (20%)</th>
            <th class="center">Measurability (15%)</th>
            <th class="center">Company Std (15%)</th>
            <th class="center">Industry (10%)</th>
            <th class="center">Actionability (10%)</th>
            <th class="center">AC Guidance (5%)</th>
            <th class="center">Total</th>
            <th>Note</th>
          </tr>
        </thead>
        <tbody>
{quality_rows}        </tbody>
      </table>
    </div>

    <footer>
      DoR Scanner v2.0 | Generated by wow-dor-scanner skill | Data sources: SRPOL Teams Confluence + Jira
    </footer>
  </div>
</body>
</html>"""

    report_path = os.path.join(OUTPUT_DIR, "Report-DoR.html")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML report saved: {report_path}")
    return report_path


def main():
    print("=" * 60)
    print("DoR SCANNER - REPORT GENERATION (2026-06-26)")
    print("=" * 60)

    # Step 4: Save Jira TXT files
    print("\n[STEP 4] Saving Jira TXT files (0 issues per team)...")
    save_jira_txt_files()

    # Step 6: DoR Quality Assessment
    print("\n[STEP 6] Loading DoR files and assessing quality...")
    dor_content = load_dor_files()
    quality_scores = assess_dor_quality(dor_content)
    for team, scores in quality_scores.items():
        print(f"  {team}: {scores['total']}/100")

    # Step 7: Generate Excel Report
    print("\n[STEP 7] Generating Report-DoR.xlsx...")
    generate_excel_report(quality_scores)

    # Step 8: Generate HTML Report
    print("\n[STEP 8] Generating Report-DoR.html...")
    generate_html_report(quality_scores)

    # Save quality scores JSON
    quality_path = os.path.join(OUTPUT_DIR, "dor-quality-scores.json")
    with open(quality_path, "w", encoding="utf-8") as f:
        json.dump(quality_scores, f, indent=2)
    print(f"\nQuality scores saved: {quality_path}")

    print("\n" + "=" * 60)
    print("DONE - All reports generated successfully")
    print("=" * 60)


if __name__ == "__main__":
    main()
